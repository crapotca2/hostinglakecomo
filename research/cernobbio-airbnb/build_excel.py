#!/usr/bin/env python3
"""
Turn july_listings.json into a polished multi-sheet Excel report.

Sheets:
    1. Riepilogo            — high-level KPIs for the month
    2. Listing               — flat comp set, one row per unique listing
    3. Prezzi settimanali   — listing × week matrix with heat map
    4. Trend settimanale    — aggregates per week
    5. Per tipologia        — group stats by type/comune
    6. Top & Bottom 20      — most expensive / cheapest

Style: Host Como brand. Teal headers #0C7489, banded rows, frozen panes,
column autofit, Italian labels.
"""
from __future__ import annotations

import json
import pathlib
import re
import sys
from datetime import datetime
from statistics import median, mean

# Windows console encoding fix
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paths ─────────────────────────────────────────────────────────────────
OUT_DIR = pathlib.Path(__file__).parent
IN_FILE = OUT_DIR / "july_listings.json"
XLSX = OUT_DIR / "Cernobbio_Luglio_2026.xlsx"

# ── Style tokens (Host Como brand) ────────────────────────────────────────
BRAND_PRIMARY = "0C7489"     # teal (header fill)
BRAND_LIGHT   = "119DB0"     # bright teal (accent)
BAND_LIGHT    = "F0F7F9"     # subtle alternating row
BAND_DARK     = "FFFFFF"
INK           = "0A2540"     # body text
GRID          = "DCE7EB"     # cell borders

THIN = Side(style="thin", color=GRID)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill(start_color=BRAND_PRIMARY, end_color=BRAND_PRIMARY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=BRAND_PRIMARY)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="555555")

KPI_LABEL_FONT = Font(name="Calibri", size=10, color="555555")
KPI_VALUE_FONT = Font(name="Calibri", size=20, bold=True, color=BRAND_PRIMARY)

CELL_FONT = Font(name="Calibri", size=10, color=INK)


# ── Number parsing (US format: 1,000.61) ──────────────────────────────────
def _parse_us_number(tok: str) -> float | None:
    if tok is None:
        return None
    clean = str(tok).replace(",", "")
    try:
        return float(clean)
    except (ValueError, TypeError):
        return None


def _per_night_from_search_row(row: dict, nights: int) -> float | None:
    """Extract per-night EUR from a single search_all result."""
    sp = row.get("price") or {}
    if not isinstance(sp, dict):
        return None

    # First try the break_down "N nights x €X" — that's the rack rate
    for ln in sp.get("break_down", []) or []:
        if not isinstance(ln, dict):
            continue
        desc = str(ln.get("description") or "")
        m = re.search(r"(?:nights?|notti)\s*x\s*\D*([\d.,]+)", desc, re.I)
        if m:
            val = _parse_us_number(m.group(1))
            if val:
                return val

    # Fallback: total amount / nights
    unit = sp.get("unit") or {}
    amt = unit.get("amount") if isinstance(unit, dict) else None
    if isinstance(amt, (int, float)) and amt > 0 and nights > 0:
        return round(float(amt) / nights, 2)
    return None


def _comune_from_title(title: str) -> str:
    """Title like "Apartment in Blevio" / "Villa in Cernobbio"."""
    if not title:
        return ""
    m = re.search(r"\bin\s+([\w'\-àèìòùÀÈÌÒÙ]+(?:\s+[\w'\-]+)?)\s*$", title.strip())
    return m.group(1) if m else ""


def _tipologia_from_title(title: str) -> str:
    """Drop the "in X" tail to keep just the type ("Apartment", "Villa", ...)."""
    if not title:
        return ""
    return re.sub(r"\s+in\s+.+$", "", title.strip())


def _bed_bath(search_row: dict) -> dict:
    out = {"camere": None, "letti": None, "bagni": None, "ospiti": None}
    sc = search_row.get("structuredContent") or {}
    if not isinstance(sc, dict):
        return out
    for line in (sc.get("primaryLine") or []) + (sc.get("secondaryLine") or []):
        if not isinstance(line, dict):
            continue
        body = str(line.get("body") or "")
        m = re.search(r"([\d.]+)", body)
        if not m:
            continue
        try:
            num = float(m.group(1))
        except ValueError:
            continue
        bl = body.lower()
        if "bedroom" in bl or "camera" in bl:
            out["camere"] = num
        elif "bath" in bl or "bagn" in bl:
            out["bagni"] = num
        elif "bed" in bl or "letto" in bl:
            out["letti"] = num
        elif "guest" in bl or "sleeps" in bl or "ospit" in bl:
            out["ospiti"] = num
    return out


def _host_flags(search_row: dict) -> dict:
    out = {"superhost": 0, "host_pro": 0, "host_blurb": ""}
    sc = search_row.get("structuredContent") or {}
    blurbs = []
    for line in (sc.get("secondaryLine") or []):
        if not isinstance(line, dict):
            continue
        body = str(line.get("body") or "")
        blurbs.append(body)
        bl = body.lower()
        if "superhost" in bl:
            out["superhost"] = 1
        if "business host" in bl or "host professionale" in bl:
            out["host_pro"] = 1
    out["host_blurb"] = " | ".join(blurbs)
    return out


def _amenity_flags(name: str, title: str) -> dict:
    """Search rows don't expose the amenities tree, so we sniff name+title."""
    text = f"{name} {title}".lower()
    return {
        "vista_lago":   int(any(s in text for s in ["lake view", "vista lago", "lakefront", "lake como"])),
        "piscina":      int("pool" in text or "piscina" in text),
        "jacuzzi":      int("jacuzzi" in text or "hot tub" in text or "idromass" in text),
        "aria_cond":    int(re.search(r"\b(a/c|ac|air condition|aria condizion)\b", text) is not None),
        "panoramic":    int("panoram" in text or "breathtaking" in text or "spectacular" in text),
    }


# ── Main aggregation ──────────────────────────────────────────────────────
def aggregate(payload: dict) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Returns (df_listings, df_weekly_matrix, weekly_stats_dict)
    """
    windows = payload.get("windows") or []
    week_labels = [w["label"] for w in windows]

    # Build per-room dict aggregating across windows
    per_room: dict[int, dict] = {}
    for w in windows:
        ci = w["check_in"]; co = w["check_out"]
        # Compute nights for this window (typically 7)
        nights = 7
        try:
            nights = (datetime.fromisoformat(co) - datetime.fromisoformat(ci)).days or 7
        except Exception:
            pass

        for row in w.get("listings") or []:
            rid = row.get("room_id")
            if rid is None:
                continue
            price = _per_night_from_search_row(row, nights=nights)
            rec = per_room.setdefault(rid, {
                "room_id": rid,
                "_row": row,
                "weekly": {lab: None for lab in week_labels},
                "_seen_in": 0,
            })
            rec["weekly"][w["label"]] = price
            rec["_seen_in"] += 1

    # Flatten
    rows = []
    matrix_rows = []
    for rid, rec in per_room.items():
        row = rec["_row"]
        coords = row.get("coordinates") or {}
        lat = coords.get("latitude") if isinstance(coords, dict) else None
        lng = (coords.get("longitude") or coords.get("longitud")) if isinstance(coords, dict) else None

        bb = _bed_bath(row)
        hf = _host_flags(row)
        af = _amenity_flags(row.get("name") or "", row.get("title") or "")

        rating_obj = row.get("rating") or {}
        rating_value = rating_obj.get("value") if isinstance(rating_obj, dict) else None
        review_count = rating_obj.get("reviewCount") if isinstance(rating_obj, dict) else None
        if rating_value == 0:
            rating_value = None
        if review_count == 0:
            review_count = None

        weekly_prices = [rec["weekly"][lab] for lab in week_labels]
        seen = [p for p in weekly_prices if p is not None]
        avg_price = round(mean(seen), 2) if seen else None
        med_price = round(median(seen), 2) if seen else None

        rows.append({
            "Nome": (row.get("name") or "")[:120],
            "Tipologia": _tipologia_from_title(row.get("title") or ""),
            "Comune": _comune_from_title(row.get("title") or ""),
            "Camere": bb["camere"],
            "Letti": bb["letti"],
            "Bagni": bb["bagni"],
            "Ospiti": bb["ospiti"],
            "Prezzo medio €/notte": avg_price,
            "Prezzo mediano €/notte": med_price,
            "N. settimane disponibili": len(seen),
            "Rating": rating_value,
            "N. recensioni": review_count,
            "Superhost": "Sì" if hf["superhost"] else "",
            "Host professionale": "Sì" if hf["host_pro"] else "",
            "Vista lago": "Sì" if af["vista_lago"] else "",
            "Piscina": "Sì" if af["piscina"] else "",
            "Jacuzzi": "Sì" if af["jacuzzi"] else "",
            "Aria condizionata": "Sì" if af["aria_cond"] else "",
            "Panoramico": "Sì" if af["panoramic"] else "",
            "Latitudine": lat,
            "Longitudine": lng,
            "ID Airbnb": rid,
            "Link": f"https://www.airbnb.com/rooms/{rid}",
        })

        matrix_rows.append({
            "Nome": (row.get("name") or "")[:60],
            "Tipologia": _tipologia_from_title(row.get("title") or ""),
            "Comune": _comune_from_title(row.get("title") or ""),
            **{lab: rec["weekly"][lab] for lab in week_labels},
            "Media": avg_price,
            "ID Airbnb": rid,
        })

    df_listings = pd.DataFrame(rows).sort_values(
        by="Prezzo medio €/notte", ascending=True, na_position="last"
    ).reset_index(drop=True)

    df_matrix = pd.DataFrame(matrix_rows).sort_values(
        by="Media", ascending=True, na_position="last"
    ).reset_index(drop=True)

    # Weekly stats
    weekly_stats = {}
    for w in windows:
        lab = w["label"]
        col_prices = []
        for row in w.get("listings") or []:
            try:
                ci = w["check_in"]; co = w["check_out"]
                nights = (datetime.fromisoformat(co) - datetime.fromisoformat(ci)).days or 7
            except Exception:
                nights = 7
            p = _per_night_from_search_row(row, nights=nights)
            if p is not None:
                col_prices.append(p)
        col_prices.sort()
        n = len(col_prices)
        weekly_stats[lab] = {
            "check_in": w["check_in"],
            "check_out": w["check_out"],
            "count": n,
            "min":    col_prices[0] if n else None,
            "p25":    col_prices[n//4] if n else None,
            "median": col_prices[n//2] if n else None,
            "mean":   round(sum(col_prices)/n, 2) if n else None,
            "p75":    col_prices[3*n//4] if n else None,
            "max":    col_prices[-1] if n else None,
        }

    return df_listings, df_matrix, weekly_stats


# ── Excel write helpers ───────────────────────────────────────────────────
def _style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = 28


def _band_rows(ws, start_row: int, end_row: int, n_cols: int) -> None:
    fill_alt = PatternFill(start_color=BAND_LIGHT, end_color=BAND_LIGHT, fill_type="solid")
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fill_type is None or cell.fill.start_color.rgb in (None, "00000000"):
                    cell.fill = fill_alt
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if cell.font.color is None or cell.font.color.rgb is None:
                cell.font = CELL_FONT


def _autosize_columns(ws, df: pd.DataFrame, start_col: int = 1) -> None:
    for idx, col in enumerate(df.columns):
        max_len = max(
            [len(str(col))] +
            [len(str(v)) if v is not None else 0 for v in df[col].head(50).tolist()]
        )
        max_len = min(max_len + 3, 50)
        ws.column_dimensions[get_column_letter(start_col + idx)].width = max_len


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, title: str | None = None,
                    subtitle: str | None = None) -> int:
    """Write df starting at start_row with optional title/subtitle. Returns
    next free row index."""
    cursor = start_row
    if title:
        ws.cell(row=cursor, column=1, value=title).font = TITLE_FONT
        ws.merge_cells(start_row=cursor, end_row=cursor,
                       start_column=1, end_column=max(len(df.columns), 4))
        cursor += 1
    if subtitle:
        ws.cell(row=cursor, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=cursor, end_row=cursor,
                       start_column=1, end_column=max(len(df.columns), 4))
        cursor += 1
    if title or subtitle:
        cursor += 1  # blank gap

    # Header row
    header_row = cursor
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=j, value=col)
    _style_header_row(ws, header_row, len(df.columns))
    cursor += 1

    # Data rows
    body_start = cursor
    for _, r in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if pd.isna(v):
                v = None
            ws.cell(row=cursor, column=j, value=v)
        cursor += 1
    body_end = cursor - 1

    _band_rows(ws, body_start, body_end, len(df.columns))
    _autosize_columns(ws, df)

    # Freeze the header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return cursor + 1


def main() -> int:
    if not IN_FILE.exists():
        print(f"missing {IN_FILE} — run scrape_july.py first", file=sys.stderr)
        return 1
    payload = json.loads(IN_FILE.read_text(encoding="utf-8"))
    df_listings, df_matrix, weekly_stats = aggregate(payload)
    n_total = len(df_listings)

    print(f"  {n_total} unique listings across {len(weekly_stats)} weekly windows")

    wb = Workbook()

    # ── Sheet 1: Riepilogo ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Riepilogo"
    ws.cell(row=1, column=1, value="Cernobbio + Zona — Mercato Affitti Brevi").font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws.cell(row=2, column=1, value=f"Luglio 2026 · scraping {payload['meta']['scraped_at'][:10]}").font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")

    # Big KPI tiles
    all_prices = [df_listings["Prezzo medio €/notte"].dropna().tolist()]
    prices_flat = sorted([p for sub in all_prices for p in sub])
    kpis = [
        ("Listing unici",        f"{n_total}"),
        ("Mediana €/notte",      f"€ {int(median(prices_flat)) if prices_flat else 0:,}".replace(",", ".")),
        ("Mean €/notte",         f"€ {int(mean(prices_flat)) if prices_flat else 0:,}".replace(",", ".")),
        ("Min €/notte",          f"€ {int(prices_flat[0]) if prices_flat else 0:,}".replace(",", ".")),
        ("Max €/notte",          f"€ {int(prices_flat[-1]) if prices_flat else 0:,}".replace(",", ".")),
    ]
    for i, (label, value) in enumerate(kpis):
        c = i + 1
        ws.cell(row=4, column=c, value=label).font = KPI_LABEL_FONT
        ws.cell(row=4, column=c).alignment = Alignment(horizontal="left")
        ws.cell(row=5, column=c, value=value).font = KPI_VALUE_FONT
        ws.cell(row=5, column=c).alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.row_dimensions[5].height = 30

    # Weekly stats table
    ws.cell(row=8, column=1, value="Trend per settimana di luglio").font = Font(
        name="Calibri", size=14, bold=True, color=BRAND_PRIMARY
    )

    weekly_df = pd.DataFrame([
        {
            "Settimana": lab,
            "Check-in": st["check_in"],
            "Check-out": st["check_out"],
            "N. listing": st["count"],
            "Mediana €/n": st["median"],
            "Mean €/n": st["mean"],
            "p25 €/n": st["p25"],
            "p75 €/n": st["p75"],
            "Min €/n": st["min"],
            "Max €/n": st["max"],
        }
        for lab, st in weekly_stats.items()
    ])
    write_dataframe(ws, weekly_df, start_row=10)

    # ── Sheet 2: Listing — vista completa ─────────────────────────────────
    ws2 = wb.create_sheet("Listing")
    write_dataframe(
        ws2, df_listings, start_row=1,
        title="Comp set completo Cernobbio + Zona",
        subtitle=f"{n_total} listing unici · prezzo medio aggregato sulle {len(weekly_stats)} settimane di luglio in cui ciascuno è apparso",
    )

    # ── Sheet 3: Prezzi settimanali (matrix) + heat map ──────────────────
    ws3 = wb.create_sheet("Prezzi settimanali")
    write_dataframe(
        ws3, df_matrix, start_row=1,
        title="Prezzi per settimana di luglio (matrix listing × settimana)",
        subtitle="Celle vuote = listing non disponibile / min-stay > 7 notti per quella settimana. Colore: rosso = caro, verde = economico.",
    )
    # Apply 3-color heat map to the 5 weekly columns + Media
    week_cols = [c for c in df_matrix.columns if c.startswith("Sett.") or c == "Media"]
    # Find header row (we wrote title+subtitle+blank → header is at row 4)
    header_row_matrix = 4
    data_start = header_row_matrix + 1
    data_end = data_start + len(df_matrix) - 1
    for col_name in week_cols:
        col_idx = list(df_matrix.columns).index(col_name) + 1
        rng = f"{get_column_letter(col_idx)}{data_start}:{get_column_letter(col_idx)}{data_end}"
        rule = ColorScaleRule(
            start_type="percentile", start_value=10, start_color="C6E6CB",
            mid_type="percentile",   mid_value=50,   mid_color="FFFFFF",
            end_type="percentile",   end_value=90,   end_color="F4C7C3",
        )
        ws3.conditional_formatting.add(rng, rule)

    # ── Sheet 4: Trend settimanale (full version) ─────────────────────────
    ws4 = wb.create_sheet("Trend settimanale")
    write_dataframe(
        ws4, weekly_df, start_row=1,
        title="Andamento prezzi settimana per settimana — Luglio 2026",
        subtitle="Mediana, mean e percentili calcolati su tutti i listing disponibili in ogni finestra",
    )

    # ── Sheet 5: Per tipologia + comune ──────────────────────────────────
    ws5 = wb.create_sheet("Per tipologia")
    grp = df_listings.groupby(["Tipologia", "Comune"])["Prezzo medio €/notte"].agg(
        ["count", "median", "mean", "min", "max"]
    ).round(0).reset_index()
    grp.columns = ["Tipologia", "Comune", "N. listing", "Mediana €/n", "Mean €/n", "Min €/n", "Max €/n"]
    grp = grp.sort_values(by="N. listing", ascending=False).reset_index(drop=True)
    write_dataframe(
        ws5, grp, start_row=1,
        title="Mediana prezzo per tipologia × comune",
        subtitle="Aggregato sul prezzo medio mensile di ogni listing. Ordinato per numerosità.",
    )

    # ── Sheet 6: Top 20 / Bottom 20 ──────────────────────────────────────
    ws6 = wb.create_sheet("Top & Bottom 20")
    cols_pretty = ["Nome", "Tipologia", "Comune", "Camere", "Letti",
                   "Prezzo medio €/notte", "Rating", "N. recensioni", "Vista lago", "Piscina", "Link"]
    top20 = df_listings.dropna(subset=["Prezzo medio €/notte"]).nlargest(20, "Prezzo medio €/notte")[cols_pretty]
    bot20 = df_listings.dropna(subset=["Prezzo medio €/notte"]).nsmallest(20, "Prezzo medio €/notte")[cols_pretty]
    next_row = write_dataframe(
        ws6, top20, start_row=1,
        title="Top 20 più cari",
        subtitle="Listing con prezzo medio più alto in luglio 2026",
    )
    write_dataframe(
        ws6, bot20, start_row=next_row + 2,
        title="Top 20 più economici",
        subtitle="Listing con prezzo medio più basso (esclusi i Null)",
    )

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(XLSX)
    print(f"\n✅  Excel scritto: {XLSX}")
    print(f"   Sheets: Riepilogo · Listing · Prezzi settimanali · Trend settimanale · Per tipologia · Top & Bottom 20")
    return 0


if __name__ == "__main__":
    sys.exit(main())
