#!/usr/bin/env python3
"""
Owner-pitch Excel v3 — 3 corrected scenarios, June + July data, market
distribution histograms, and surrounding hotel landscape.

Scenari (corrected from owner's brief):
    1. Solo piano terra      → mini-studio (cucina+bagno+ripostiglio)
                                ~0-1 camere, 2 ospiti.
    2. Solo mansarda          → 2 camere matrimoniali, 4 ospiti.
    3. Casa intera            → piano 1 (1-2 camere) + mansarda (2 cam.)
                                = 3-4 camere, 6-8 ospiti.

Output: Cernobbio_3_Scenari.xlsx with 9 sheets including embedded
matplotlib histograms for room/guest distribution comparison.
"""
from __future__ import annotations

import json
import pathlib
import re
import sys
from datetime import datetime, date, timedelta
from statistics import median, mean

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XLImage

# ── Paths ─────────────────────────────────────────────────────────────────
OUT_DIR = pathlib.Path(__file__).parent
WEEKLY_JUL  = OUT_DIR / "july_listings.json"
WEEKLY_JUN  = OUT_DIR / "june_listings.json"
DAILY_JUL   = OUT_DIR / "july_daily.json"
DAILY_JUN   = OUT_DIR / "june_daily.json"
HOTELS_JSON = OUT_DIR / "hotels.json"
XLSX        = OUT_DIR / "Cernobbio_3_Scenari.xlsx"
TMP_DIR     = OUT_DIR / ".charts"
TMP_DIR.mkdir(exist_ok=True)

# ── Brand tokens (Host Como teal) ─────────────────────────────────────────
BRAND   = "0C7489"
BRAND_2 = "119DB0"
BAND    = "F0F7F9"
INK     = "0A2540"
GRID    = "DCE7EB"
ACCENT  = "F4A261"

# Scenario palette (used in histograms)
COLOR_S1 = "#4E9CC9"   # light blue — piano terra
COLOR_S2 = "#0C7489"   # teal brand — mansarda
COLOR_S3 = "#073B4C"   # navy — casa intera
COLOR_OTHER = "#D5DEE3"

THIN = Side(style="thin", color=GRID)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=BRAND)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="555555")
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color=BRAND)
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="555555")
KPI_VALUE_FONT = Font(name="Calibri", size=22, bold=True, color=BRAND)
CELL_FONT = Font(name="Calibri", size=10, color=INK)

# ── Parse helpers (shared with scrape) ────────────────────────────────────
def _parse_us(tok):
    if tok is None: return None
    try: return float(str(tok).replace(",", ""))
    except (ValueError, TypeError): return None


def _per_night(row: dict, nights: int) -> float | None:
    sp = row.get("price") or {}
    if not isinstance(sp, dict):
        return None
    for ln in sp.get("break_down", []) or []:
        if not isinstance(ln, dict): continue
        desc = str(ln.get("description") or "")
        m = re.search(r"(?:nights?|notti)\s*x\s*\D*([\d.,]+)", desc, re.I)
        if m:
            v = _parse_us(m.group(1))
            if v: return v
    unit = sp.get("unit") or {}
    amt = unit.get("amount") if isinstance(unit, dict) else None
    if isinstance(amt, (int, float)) and amt > 0 and nights > 0:
        return round(amt / nights, 2)
    return None


def _comune(title):
    if not title: return ""
    m = re.search(r"\bin\s+([\w'\-àèìòùÀÈÌÒÙ]+(?:\s+[\w'\-]+)?)\s*$", title.strip())
    return m.group(1) if m else ""


def _tipologia(title):
    if not title: return ""
    return re.sub(r"\s+in\s+.+$", "", title.strip())


def _bed_bath(sr):
    out = {"camere": None, "letti": None, "bagni": None, "ospiti": None}
    sc = sr.get("structuredContent") or {}
    if not isinstance(sc, dict): return out
    for line in (sc.get("primaryLine") or []) + (sc.get("secondaryLine") or []):
        if not isinstance(line, dict): continue
        body = str(line.get("body") or "")
        m = re.search(r"([\d.]+)", body)
        if not m: continue
        try: num = float(m.group(1))
        except ValueError: continue
        bl = body.lower()
        if "bedroom" in bl or "camera" in bl: out["camere"] = num
        elif "bath" in bl or "bagn" in bl:    out["bagni"]  = num
        elif "bed" in bl or "letto" in bl:    out["letti"]  = num
        elif "guest" in bl or "sleeps" in bl: out["ospiti"] = num
    return out


TIPICITY_KEYWORDS = [
    "stone", "pietra", "rustic", "rustico",
    "historic", "antico", "storica", "storico",
    "traditional", "tipica", "tipico",
    "village", "borgo", "old town", "centro storico",
    "wood beams", "wooden beams", "travi", "legno",
    "century", "secolo", "1700", "1800", "1900",
    "cottage", "casale", "casa indipendente",
    "courtyard", "corte", "garden", "giardino",
    "panoram", "breathtaking", "lake view", "vista lago",
]


def _typical_score(name, title):
    text = f"{name} {title}".lower()
    hits = [k for k in TIPICITY_KEYWORDS if k in text]
    return len(hits), hits


# ── Build aggregated comp set across weekly windows of one month ──────────
def comp_set_from_weekly(weekly_path: pathlib.Path, month_label: str) -> pd.DataFrame:
    if not weekly_path.exists():
        return pd.DataFrame()
    payload = json.loads(weekly_path.read_text(encoding="utf-8"))
    windows = payload["windows"]
    per_room: dict[int, dict] = {}
    for w in windows:
        try:
            nights = (datetime.fromisoformat(w["check_out"]) - datetime.fromisoformat(w["check_in"])).days or 7
        except Exception:
            nights = 7
        for row in w.get("listings") or []:
            rid = row.get("room_id")
            if not rid: continue
            rec = per_room.setdefault(rid, {"_row": row, "prices": []})
            p = _per_night(row, nights=nights)
            if p is not None:
                rec["prices"].append(p)
    rows = []
    for rid, rec in per_room.items():
        sr = rec["_row"]
        bb = _bed_bath(sr)
        name = sr.get("name") or ""
        title = sr.get("title") or ""
        score, hits = _typical_score(name, title)
        rating_obj = sr.get("rating") or {}
        rating_value = rating_obj.get("value") if isinstance(rating_obj, dict) else None
        if rating_value == 0: rating_value = None
        review_count = rating_obj.get("reviewCount") if isinstance(rating_obj, dict) else None
        if review_count == 0: review_count = None

        prices = rec["prices"]
        rows.append({
            "Mese": month_label,
            "ID Airbnb": rid,
            "Nome": name[:120],
            "Tipologia": _tipologia(title),
            "Comune": _comune(title),
            "Camere": bb["camere"],
            "Letti": bb["letti"],
            "Bagni": bb["bagni"],
            "Ospiti": bb["ospiti"],
            "Prezzo medio €/n": round(mean(prices), 2) if prices else None,
            "Prezzo mediano €/n": round(median(prices), 2) if prices else None,
            "N. settimane": len(prices),
            "Rating": rating_value,
            "Recensioni": review_count,
            "Affinità tipica": score,
            "Indizi": ", ".join(hits[:5]),
            "Link": f"https://www.airbnb.com/rooms/{rid}",
        })
    return pd.DataFrame(rows)


# ── Build daily matrix for a month ────────────────────────────────────────
def daily_matrix(daily_path: pathlib.Path) -> pd.DataFrame:
    if not daily_path.exists():
        return pd.DataFrame()
    payload = json.loads(daily_path.read_text(encoding="utf-8"))
    days = payload.get("days") or []
    day_labels = [d["start"] for d in days]
    per_room: dict[int, dict] = {}
    for d in days:
        for row in d.get("listings") or []:
            rid = row.get("room_id")
            if not rid: continue
            p = _per_night(row, nights=2)
            rec = per_room.setdefault(rid, {"_row": row,
                                            "daily": {lab: None for lab in day_labels}})
            rec["daily"][d["start"]] = p

    rows = []
    for rid, rec in per_room.items():
        sr = rec["_row"]
        bb = _bed_bath(sr)
        name = sr.get("name") or ""
        title = sr.get("title") or ""
        score, _ = _typical_score(name, title)
        seen = [v for v in rec["daily"].values() if v is not None]
        out = {
            "ID Airbnb": rid,
            "Nome": name[:80],
            "Tipologia": _tipologia(title),
            "Comune": _comune(title),
            "Camere": bb["camere"],
            "Ospiti": bb["ospiti"],
            "Affinità tipica": score,
            "Media giorni": round(mean(seen), 2) if seen else None,
            "Giorni visti": len(seen),
        }
        out.update({lab: rec["daily"][lab] for lab in day_labels})
        rows.append(out)
    return pd.DataFrame(rows)


# ── Scenari ───────────────────────────────────────────────────────────────
SCEN_1 = {
    "key":           "scen1",
    "label":         "Scen. 1 — Solo piano terra",
    "subtitle":      "Mini-studio: cucina + bagno + ampio ripostiglio. Sviluppo orizzontale, 0-1 letto matrimoniale, 2 ospiti.",
    "ospiti_range":  (1, 3),
    "camere_range":  (0, 1),
    "tipi_ammessi":  ["Apartment", "Room", "Condo", "Place to stay", "Vacation home", "Cottage"],
    "comuni":        ["Cernobbio", "Moltrasio", "Blevio", "Como"],
    "color":         COLOR_S1,
    "exclude_villas": True,
}
SCEN_2 = {
    "key":           "scen2",
    "label":         "Scen. 2 — Solo mansarda",
    "subtitle":      "Solo mansarda affittata: 2 camere matrimoniali, 4 ospiti. Soffitti in legno, ambiente raccolto.",
    "ospiti_range":  (3, 5),
    "camere_range":  (2, 2),
    "tipi_ammessi":  ["Apartment", "Home", "Condo", "Vacation home", "Cottage", "Townhouse"],
    "comuni":        ["Cernobbio", "Moltrasio", "Blevio", "Como"],
    "color":         COLOR_S2,
    "exclude_villas": True,
}
SCEN_3 = {
    "key":           "scen3",
    "label":         "Scen. 3 — Casa intera",
    "subtitle":      "Piano terra + piano 1 + mansarda. 3-4 camere, 6-8 ospiti. Casa tipica completa con corte privata.",
    "ospiti_range":  (6, 9),
    "camere_range":  (3, 5),
    "tipi_ammessi":  ["Home", "Townhouse", "Apartment", "Vacation home", "Cottage"],
    "comuni":        ["Cernobbio", "Moltrasio", "Blevio", "Como"],
    "color":         COLOR_S3,
    "exclude_villas": True,  # owner's house is "tipica", not luxury villa
}
SCENARIOS = [SCEN_1, SCEN_2, SCEN_3]


def filter_scenario(df: pd.DataFrame, sc: dict) -> pd.DataFrame:
    cmin, cmax = sc["camere_range"]
    omin, omax = sc["ospiti_range"]
    f = df.copy()
    if f.empty:
        return f
    cond_camere = f["Camere"].between(cmin, cmax, inclusive="both")
    cond_ospiti = f["Ospiti"].between(omin, omax, inclusive="both")
    # If both known → camere primary; if camere unknown, fall back to ospiti
    cond_cap = cond_camere.fillna(False) | (f["Camere"].isna() & cond_ospiti.fillna(False))
    cond_type = f["Tipologia"].isin(sc["tipi_ammessi"])
    cond_comune = f["Comune"].isin(sc["comuni"])
    if sc.get("exclude_villas"):
        cond_type = cond_type & (f["Tipologia"] != "Villa")
    f = f[cond_cap & cond_type & cond_comune].copy()
    f = f.sort_values(by=["Affinità tipica", "Prezzo mediano €/n"],
                      ascending=[False, True]).reset_index(drop=True)
    return f


# ── Histograms ────────────────────────────────────────────────────────────
def plot_distribution(df_jun: pd.DataFrame, df_jul: pd.DataFrame, field: str,
                       title: str, xlabel: str, out_path: pathlib.Path,
                       max_val: int = 10):
    """Stacked-style side-by-side bars for June vs July on the given int field."""
    def _counts(df):
        if df.empty or field not in df.columns:
            return {i: 0 for i in range(max_val + 1)}
        s = df[field].dropna().astype(int).tolist()
        c = {i: 0 for i in range(max_val + 1)}
        for v in s:
            if v > max_val:
                v = max_val
            c[v] = c.get(v, 0) + 1
        return c

    cj = _counts(df_jun)
    cl = _counts(df_jul)
    xs = sorted(cj.keys())
    width = 0.4

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar([x - width/2 for x in xs], [cj[x] for x in xs], width,
           label="Giugno 2026", color=COLOR_S1, edgecolor="white")
    ax.bar([x + width/2 for x in xs], [cl[x] for x in xs], width,
           label="Luglio 2026", color=COLOR_S2, edgecolor="white")

    # Highlight scenario bands as colored background spans
    band_alpha = 0.10
    ax.axvspan(SCEN_1["camere_range"][0] - 0.5,
               SCEN_1["camere_range"][1] + 0.5, alpha=band_alpha,
               color=COLOR_S1, label=f"{SCEN_1['label']}") if field == "Camere" else None

    ax.set_xticks(xs)
    ax.set_xticklabels([f"{x}+" if x == max_val else str(x) for x in xs])
    ax.set_xlabel(xlabel)
    ax.set_ylabel("Numero listing")
    ax.set_title(title, fontsize=13, fontweight="bold", color="#073B4C")
    ax.grid(axis="y", linestyle=":", alpha=0.4)
    ax.legend(loc="upper right", framealpha=0.95)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()
    plt.savefig(out_path, dpi=130, bbox_inches="tight")
    plt.close()


def plot_grouping(df_jun: pd.DataFrame, df_jul: pd.DataFrame,
                   out_path: pathlib.Path):
    """Stacked horizontal bars: per-scenario count, June vs July."""
    rows = []
    for sc in SCENARIOS:
        n_jun = len(filter_scenario(df_jun, sc))
        n_jul = len(filter_scenario(df_jul, sc))
        rows.append((sc["label"], n_jun, n_jul, sc["color"]))

    labels = [r[0] for r in rows]
    jun_vals = [r[1] for r in rows]
    jul_vals = [r[2] for r in rows]
    colors_l = [r[3] for r in rows]
    y = list(range(len(rows)))
    height = 0.4

    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.barh([yi - height/2 for yi in y], jun_vals, height,
            label="Giugno 2026", color=COLOR_S1, edgecolor="white")
    ax.barh([yi + height/2 for yi in y], jul_vals, height,
            label="Luglio 2026", color=COLOR_S2, edgecolor="white")
    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.set_xlabel("Listing comparabili nel mercato (bbox Cernobbio + zona)")
    ax.set_title("Profondità del comp set per scenario", fontsize=13,
                 fontweight="bold", color="#073B4C")
    for i, (jn, jl) in enumerate(zip(jun_vals, jul_vals)):
        ax.text(jn + 1, i - height/2, str(jn), va="center", fontsize=9, color="#0A2540")
        ax.text(jl + 1, i + height/2, str(jl), va="center", fontsize=9, color="#0A2540")
    ax.grid(axis="x", linestyle=":", alpha=0.4)
    ax.legend(loc="lower right", framealpha=0.95)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()
    plt.savefig(out_path, dpi=130, bbox_inches="tight")
    plt.close()


def plot_daily_trend(df_daily_jun: pd.DataFrame,
                      df_daily_jul: pd.DataFrame,
                      out_path: pathlib.Path):
    """Daily median price comparison giugno vs luglio."""
    def _series(df, month_idx):
        if df.empty:
            return [], []
        dates = sorted([c for c in df.columns if c.startswith("2026-")])
        days, meds = [], []
        for d in dates:
            vals = df[d].dropna().tolist()
            if not vals: continue
            days.append(int(d[-2:]))  # day of month
            meds.append(median(vals))
        return days, meds

    djun, mjun = _series(df_daily_jun, 6)
    djul, mjul = _series(df_daily_jul, 7)

    fig, ax = plt.subplots(figsize=(11, 5))
    if djun:
        ax.plot(djun, mjun, "o-", label="Giugno 2026", color=COLOR_S1,
                linewidth=2, markersize=5)
    if djul:
        ax.plot(djul, mjul, "s-", label="Luglio 2026", color=COLOR_S2,
                linewidth=2, markersize=5)
    ax.set_xlabel("Giorno del mese")
    ax.set_ylabel("Mediana €/notte")
    ax.set_title("Andamento mediana prezzo daily — Giugno vs Luglio 2026",
                 fontsize=13, fontweight="bold", color="#073B4C")
    ax.grid(linestyle=":", alpha=0.4)
    ax.legend(loc="best", framealpha=0.95)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlim(0.5, 30.5)
    plt.tight_layout()
    plt.savefig(out_path, dpi=130, bbox_inches="tight")
    plt.close()


# ── Excel write helpers ───────────────────────────────────────────────────
def _style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = 28


def _band_rows(ws, start_row, end_row, n_cols):
    fill_alt = PatternFill(start_color=BAND, end_color=BAND, fill_type="solid")
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fill_type is None or (cell.fill.start_color.rgb in (None, "00000000")):
                    cell.fill = fill_alt
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if cell.font is None or cell.font.color is None or cell.font.color.rgb is None:
                cell.font = CELL_FONT


def _autosize(ws, df, start_col=1, max_width=42):
    for idx, col in enumerate(df.columns):
        col_strs = [str(col)]
        for v in df[col].head(50).tolist():
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            col_strs.append(str(v))
        ws.column_dimensions[get_column_letter(start_col + idx)].width = min(max(len(s) for s in col_strs) + 3, max_width)


def write_df(ws, df, start_row=1, title=None, subtitle=None):
    cursor = start_row
    if title:
        ws.cell(row=cursor, column=1, value=title).font = TITLE_FONT
        ws.merge_cells(start_row=cursor, end_row=cursor, start_column=1,
                       end_column=max(len(df.columns), 4))
        cursor += 1
    if subtitle:
        ws.cell(row=cursor, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=cursor, end_row=cursor, start_column=1,
                       end_column=max(len(df.columns), 4))
        cursor += 1
    if title or subtitle:
        cursor += 1
    header_row = cursor
    if df.empty:
        return cursor + 1, header_row, header_row + 1, header_row
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=j, value=col)
    _style_header_row(ws, header_row, len(df.columns))
    cursor += 1
    body_start = cursor
    for _, r in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if pd.isna(v): v = None
            ws.cell(row=cursor, column=j, value=v)
        cursor += 1
    body_end = cursor - 1
    _band_rows(ws, body_start, body_end, len(df.columns))
    _autosize(ws, df)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return cursor + 1, header_row, body_start, body_end


# ── Revenue model (per scenario) ──────────────────────────────────────────
def _revenue_table(adr_jul_median: float) -> pd.DataFrame:
    if not adr_jul_median or adr_jul_median <= 0:
        return pd.DataFrame()
    months = [
        ("Gennaio",   31, 0.48),
        ("Febbraio",  28, 0.48),
        ("Marzo",     31, 0.55),
        ("Aprile",    30, 0.62),
        ("Maggio",    31, 0.78),
        ("Giugno",    30, 0.92),
        ("Luglio",    31, 1.00),
        ("Agosto",    31, 1.00),
        ("Settembre", 30, 0.78),
        ("Ottobre",   31, 0.62),
        ("Novembre",  30, 0.48),
        ("Dicembre",  31, 0.68),
    ]
    rows, totals = [], {"50%": 0.0, "65%": 0.0, "75%": 0.0}
    for name, n_days, mult in months:
        month_adr = adr_jul_median * mult
        for label, occ in [("50%", 0.50), ("65%", 0.65), ("75%", 0.75)]:
            totals[label] += month_adr * n_days * occ
        rows.append({
            "Mese": name, "Giorni": n_days, "ADR stimato €": round(month_adr, 0),
            "Ricavo @ 50%": round(month_adr * n_days * 0.50, 0),
            "Ricavo @ 65%": round(month_adr * n_days * 0.65, 0),
            "Ricavo @ 75%": round(month_adr * n_days * 0.75, 0),
        })
    rows.append({"Mese": "TOTALE ANNO", "Giorni": 365,
                 "ADR stimato €": round(adr_jul_median, 0),
                 "Ricavo @ 50%": round(totals["50%"], 0),
                 "Ricavo @ 65%": round(totals["65%"], 0),
                 "Ricavo @ 75%": round(totals["75%"], 0)})
    return pd.DataFrame(rows)


def _kpi(df: pd.DataFrame) -> dict:
    if df.empty:
        return {"n": 0, "median": None, "mean": None, "p25": None, "p75": None, "min": None, "max": None}
    prices = df["Prezzo mediano €/n"].dropna().tolist()
    if not prices:
        return {"n": len(df), "median": None, "mean": None, "p25": None, "p75": None, "min": None, "max": None}
    prices.sort()
    n = len(prices)
    return {
        "n": len(df),
        "median": median(prices), "mean": mean(prices),
        "p25": prices[n//4], "p75": prices[3*n//4],
        "min": prices[0], "max": prices[-1],
    }


# ── Main ──────────────────────────────────────────────────────────────────
def main() -> int:
    if not WEEKLY_JUL.exists():
        print(f"missing {WEEKLY_JUL}", file=sys.stderr)
        return 1

    print("→ Loading weekly data…")
    df_jun = comp_set_from_weekly(WEEKLY_JUN, "giugno")
    df_jul = comp_set_from_weekly(WEEKLY_JUL, "luglio")
    print(f"  giugno: {len(df_jun)} listings · luglio: {len(df_jul)} listings")

    print("→ Loading daily data…")
    df_d_jun = daily_matrix(DAILY_JUN)
    df_d_jul = daily_matrix(DAILY_JUL)
    print(f"  daily: giugno {len(df_d_jun)} · luglio {len(df_d_jul)}")

    print("→ Filtering 3 scenarios…")
    sc_jun = {sc["key"]: filter_scenario(df_jun, sc) for sc in SCENARIOS}
    sc_jul = {sc["key"]: filter_scenario(df_jul, sc) for sc in SCENARIOS}
    for sc in SCENARIOS:
        print(f"  {sc['label']:<35} giugno: {len(sc_jun[sc['key']]):>3} · luglio: {len(sc_jul[sc['key']]):>3}")

    print("→ Plotting histograms + trends…")
    chart_rooms     = TMP_DIR / "hist_camere.png"
    chart_guests    = TMP_DIR / "hist_ospiti.png"
    chart_grouping  = TMP_DIR / "scenario_grouping.png"
    chart_trend     = TMP_DIR / "trend_daily.png"
    plot_distribution(df_jun, df_jul, "Camere",
                      "Distribuzione listing per numero CAMERE — Giugno vs Luglio 2026",
                      "Numero camere", chart_rooms, max_val=7)
    plot_distribution(df_jun, df_jul, "Ospiti",
                      "Distribuzione listing per numero OSPITI — Giugno vs Luglio 2026",
                      "Numero ospiti", chart_guests, max_val=10)
    plot_grouping(df_jun, df_jul, chart_grouping)
    plot_daily_trend(df_d_jun, df_d_jul, chart_trend)

    print("→ Loading hotels…")
    hotels_payload = json.loads(HOTELS_JSON.read_text(encoding="utf-8")) if HOTELS_JSON.exists() else {"hotels": []}
    hotels_rows = []
    for h in hotels_payload.get("hotels", []):
        hotels_rows.append({
            "Hotel": h["name"],
            "Tipo": h.get("primary_type_display") or h.get("primary_type") or "",
            "Rating": h.get("rating"),
            "Recensioni": h.get("reviews"),
            "Indirizzo": h.get("address"),
            "Sito web": h.get("website"),
            "Telefono": h.get("phone"),
            "Maps": h.get("google_maps_url"),
            "Booking giugno (15-22)": h.get("booking_june"),
            "Booking luglio (15-22)": h.get("booking_july"),
        })
    df_hotels = pd.DataFrame(hotels_rows).sort_values(
        by="Rating", ascending=False, na_position="last"
    ).reset_index(drop=True) if hotels_rows else pd.DataFrame()

    print("→ Building workbook…")
    wb = Workbook()

    # ── Sheet 1: Brief ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Brief"
    ws.cell(row=1, column=1, value="Casa tipica Cernobbio — 3 Scenari").font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws.cell(row=2, column=1,
            value=f"Mercato affitti brevi · Giugno + Luglio 2026 baseline · scraping {datetime.now().strftime('%Y-%m-%d')}"
            ).font = SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    ws.cell(row=4, column=1, value="Profilo immobile").font = SECTION_FONT
    profile = [
        ["Località", "Cernobbio, zona La Vignetta (pre-collinare)"],
        ["Caratteristiche", "Casa indipendente in corte privata, pietra, soffitti legno massiccio, parquet, cementine"],
        ["Sviluppo", "Piano terra (cucina+bagno+ripostiglio) · Piano 1 (1 cam + soggiorno o 2 cam + terrazzo) · Mansarda (2 cam)"],
    ]
    for i, (k, v) in enumerate(profile, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, color=INK)
        ws.cell(row=i, column=2, value=v).font = CELL_FONT
        ws.merge_cells(start_row=i, end_row=i, start_column=2, end_column=7)

    ws.cell(row=9, column=1, value="I 3 scenari valutati").font = SECTION_FONT
    scen_rows = []
    for sc in SCENARIOS:
        kj = _kpi(sc_jun[sc["key"]])
        kl = _kpi(sc_jul[sc["key"]])
        scen_rows.append({
            "Scenario": sc["label"],
            "Layout": sc["subtitle"],
            "Comp. Giu": kj["n"],
            "Mediana Giu €/n": round(kj["median"], 0) if kj["median"] else None,
            "Comp. Lug": kl["n"],
            "Mediana Lug €/n": round(kl["median"], 0) if kl["median"] else None,
            "p25-p75 Lug": (
                f"{int(kl['p25'])}-{int(kl['p75'])}"
                if kl["p25"] and kl["p75"] else ""
            ),
            "Δ Giu→Lug": (
                f"{int((kl['median']-kj['median'])/kj['median']*100):+d}%"
                if kj["median"] and kl["median"] else ""
            ),
        })
    write_df(ws, pd.DataFrame(scen_rows), start_row=10)

    notes = [
        "• Scen. 1 (piano terra) è il segmento più ampio in mercato — mini-studio in Cernobbio è la categoria più affollata. Margine: ADR contenuto ma alta rotazione.",
        "• Scen. 2 (mansarda) intercetta coppie/piccole famiglie 4 ospiti — sweet spot per posizionamento premium se enfatizzi la tipicità (legno+cementine).",
        "• Scen. 3 (casa intera) ha pochi comparabili NON-villa: il segmento 3-4 camere a Cernobbio è dominato dalle ville luxury, voi sareste in una fascia distinta (€500-900/n vs €1500+/n delle ville).",
        "• Il mercato cresce ~6-8% da giugno a luglio (vedi colonna Δ Giu→Lug). Pricing dinamico: alzare ADR del 15-20% per il periodo 15 lug – 25 ago.",
    ]
    for i, n in enumerate(notes, start=16):
        c = ws.cell(row=i, column=1, value=n)
        c.font = CELL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=7)
        ws.row_dimensions[i].height = 32
    for col, w in [("A", 20), ("B", 50), ("C", 12), ("D", 16), ("E", 12), ("F", 16), ("G", 14)]:
        ws.column_dimensions[col].width = w

    # ── Sheet 2-4: Scen. 1/2/3 — Comparabili ──────────────────────────────
    display_cols = [
        "Mese", "Nome", "Tipologia", "Comune", "Camere", "Letti", "Bagni", "Ospiti",
        "Prezzo mediano €/n", "Prezzo medio €/n",
        "Rating", "Recensioni",
        "Affinità tipica", "Indizi",
        "Link",
    ]
    sheet_names = ["Scen. 1 piano terra", "Scen. 2 mansarda", "Scen. 3 casa intera"]
    for sc, sheet_name in zip(SCENARIOS, sheet_names):
        ws_s = wb.create_sheet(sheet_name)
        combined = pd.concat([sc_jun[sc["key"]], sc_jul[sc["key"]]],
                             ignore_index=True)
        if combined.empty:
            ws_s.cell(row=1, column=1, value=sc["label"]).font = TITLE_FONT
            ws_s.cell(row=3, column=1, value="Nessun listing comparabile trovato in questa fascia.").font = CELL_FONT
            continue
        kj = _kpi(sc_jun[sc["key"]])
        kl = _kpi(sc_jul[sc["key"]])
        write_df(
            ws_s, combined[display_cols].head(60), start_row=1,
            title=sc["label"],
            subtitle=(
                f"{sc['subtitle']} · "
                f"Giugno: {kj['n']} comp. · mediana €{int(kj['median'] or 0)}/n · "
                f"Luglio: {kl['n']} comp. · mediana €{int(kl['median'] or 0)}/n"
            ),
        )

    # ── Sheet 5: Distribuzione mercato (istogrammi) ──────────────────────
    ws_h = wb.create_sheet("Distribuzione mercato")
    ws_h.cell(row=1, column=1, value="Distribuzione del mercato — Camere e Ospiti").font = TITLE_FONT
    ws_h.merge_cells("A1:I1")
    ws_h.cell(row=2, column=1,
              value="Quanti listing esistono per ogni numero di camere/ospiti nel bbox di Cernobbio, mese per mese."
              ).font = SUBTITLE_FONT
    ws_h.merge_cells("A2:I2")

    # Embed histograms
    ws_h.cell(row=4, column=1, value="Per numero CAMERE").font = SECTION_FONT
    img = XLImage(str(chart_rooms))
    img.anchor = "A6"
    ws_h.add_image(img)

    ws_h.cell(row=33, column=1, value="Per numero OSPITI").font = SECTION_FONT
    img2 = XLImage(str(chart_guests))
    img2.anchor = "A35"
    ws_h.add_image(img2)

    ws_h.cell(row=62, column=1, value="Profondità del comp set per scenario").font = SECTION_FONT
    img3 = XLImage(str(chart_grouping))
    img3.anchor = "A64"
    ws_h.add_image(img3)

    # Tabular grouping below the charts
    cur = 90
    ws_h.cell(row=cur, column=1, value="Tabella di raggruppamento per scenario").font = SECTION_FONT
    cur += 1
    rows_g = []
    for sc in SCENARIOS:
        kj = _kpi(sc_jun[sc["key"]])
        kl = _kpi(sc_jul[sc["key"]])
        rows_g.append({
            "Scenario": sc["label"],
            "Range camere": f"{sc['camere_range'][0]}-{sc['camere_range'][1]}",
            "Range ospiti": f"{sc['ospiti_range'][0]}-{sc['ospiti_range'][1]}",
            "Listing Giugno": kj["n"],
            "Listing Luglio": kl["n"],
            "Mediana €/n Giugno": round(kj["median"], 0) if kj["median"] else None,
            "Mediana €/n Luglio": round(kl["median"], 0) if kl["median"] else None,
        })
    write_df(ws_h, pd.DataFrame(rows_g), start_row=cur)

    # ── Sheet 6: Trend daily giugno + luglio ─────────────────────────────
    ws_t = wb.create_sheet("Trend giornaliero")
    ws_t.cell(row=1, column=1, value="Andamento giornaliero — Giugno e Luglio 2026").font = TITLE_FONT
    ws_t.merge_cells("A1:I1")
    ws_t.cell(row=2, column=1, value="Mediana €/notte per ogni giorno del mese, su 2-night sliding windows.").font = SUBTITLE_FONT
    ws_t.merge_cells("A2:I2")

    img4 = XLImage(str(chart_trend))
    img4.anchor = "A4"
    ws_t.add_image(img4)

    # Below: combined trend table
    cur = 30
    ws_t.cell(row=cur, column=1, value="Tabella trend daily").font = SECTION_FONT
    cur += 1

    def _trend_rows(df, month_idx):
        rows = []
        if df.empty:
            return rows
        dates = sorted([c for c in df.columns if c.startswith("2026-")])
        for d in dates:
            vals = df[d].dropna().tolist()
            if not vals: continue
            vals.sort()
            n = len(vals)
            try:
                dt = date.fromisoformat(d)
                dow = ["Lun","Mar","Mer","Gio","Ven","Sab","Dom"][dt.weekday()]
            except Exception:
                dow = ""
            rows.append({
                "Mese":  ["Giugno","Luglio"][month_idx],
                "Data":  d,
                "Giorno": dow,
                "N. listing": n,
                "Mediana €/n": round(median(vals), 0),
                "Mean €/n":    round(mean(vals), 0),
                "p25 €/n":     round(vals[n//4], 0),
                "p75 €/n":     round(vals[3*n//4], 0),
            })
        return rows

    trend_combined = pd.DataFrame(_trend_rows(df_d_jun, 0) + _trend_rows(df_d_jul, 1))
    if not trend_combined.empty:
        write_df(ws_t, trend_combined, start_row=cur)

    # ── Sheet 7: Hotels ───────────────────────────────────────────────────
    ws_ho = wb.create_sheet("Hotel zona Cernobbio")
    if not df_hotels.empty:
        write_df(
            ws_ho, df_hotels, start_row=1,
            title=f"Hotel a Cernobbio + zona ({len(df_hotels)} strutture)",
            subtitle="Fonte: Google Places API · Nightly rates non disponibili da Google; i link Booking aprono la search corretta per la data. Ordinati per rating.",
        )
    else:
        ws_ho.cell(row=1, column=1, value="Hotel zona Cernobbio").font = TITLE_FONT
        ws_ho.cell(row=3, column=1, value="Run scrape_hotels.py per popolare questa scheda.").font = CELL_FONT

    # ── Sheet 8: Ricavi annui per scenario ────────────────────────────────
    ws_rev = wb.create_sheet("Ricavi annui")
    ws_rev.cell(row=1, column=1, value="Stima ricavi annui — proiezione su 3 occupancy").font = TITLE_FONT
    ws_rev.merge_cells("A1:G1")
    ws_rev.cell(row=2, column=1, value="Modello: ADR luglio (mediana) × moltiplicatore stagionale × giorni × occupancy. Lordo, no commissioni piattaforma, no imposte.").font = SUBTITLE_FONT
    ws_rev.merge_cells("A2:G2")

    next_row = 4
    for sc in SCENARIOS:
        kl = _kpi(sc_jul[sc["key"]])
        adr = kl["median"]
        if not adr:
            ws_rev.cell(row=next_row, column=1, value=f"{sc['label']} — comp set vuoto, modello non calcolato.").font = SECTION_FONT
            next_row += 2
            continue
        ws_rev.cell(row=next_row, column=1,
                    value=f"{sc['label']} — ADR baseline luglio €{int(adr)}").font = SECTION_FONT
        nr, _, _, _ = write_df(ws_rev, _revenue_table(adr), start_row=next_row + 1)
        next_row = nr + 2

    # ── Sheet 9: Metodologia ──────────────────────────────────────────────
    ws_m = wb.create_sheet("Metodologia")
    ws_m.cell(row=1, column=1, value="Metodologia").font = TITLE_FONT
    ws_m.merge_cells("A1:E1")
    notes_m = [
        "Fonti dati:",
        "  • Airbnb StaysSearch (API interna) via pyairbnb (TLS impersonation Chrome).",
        "  • Google Places API (New) places:searchText per gli hotel.",
        "",
        "Coperture temporali:",
        "  • Weekly: 4 finestre giugno (3 giu-1 lug, Mer→Mer) + 5 finestre luglio (1 lug-5 ago).",
        "  • Daily: 30 sliding 2-night windows per ciascun mese (start day = ogni giorno 1-30).",
        "",
        "Filtri Scenari:",
        "  Scen. 1 (Solo piano terra) — Camere 0-1, Ospiti 1-3, no Villa, Cernobbio + zona.",
        "  Scen. 2 (Solo mansarda)   — Camere 2, Ospiti 3-5, no Villa.",
        "  Scen. 3 (Casa intera)     — Camere 3-5, Ospiti 6-9, no Villa (la casa NON è villa luxury).",
        "",
        "Punteggio affinità casa tipica: conta hit di parole chiave (stone, pietra, rustic, historic, antico, tipica, traditional, village, borgo, wood beams, travi, legno, century, cottage, courtyard, corte, garden, panoram, vista lago…) nel titolo/nome listing.",
        "",
        "Modello ricavi:",
        "  Mensile = ADR Luglio × moltiplicatore stagionale × giorni × occupancy.",
        "  Moltiplicatori: alta (giu/lug/ago) 0.92/1.00/1.00, shoulder+ (mag/set) 0.78, shoulder (apr/ott) 0.62, dicembre weighted 0.68, bassa (gen/feb/nov) 0.48.",
        "",
        "Caveat:",
        "  • Prezzi = rack rate Airbnb dinamici al momento dello scraping. Non confermati come booking.",
        "  • Hotel prices = non disponibili via Google Places. Operator deve aprire i link Booking per i numeri reali.",
        "  • Per gli scenari: il filtro 'Tipologia' usa il tag Airbnb auto-assegnato. Falsi positivi possibili.",
        "  • Occupancy 50/65/75% sono ipotesi standard di mercato; storico Smoobu o AirDNA validerebbero il dato.",
    ]
    for i, n in enumerate(notes_m, start=3):
        c = ws_m.cell(row=i, column=1, value=n)
        c.font = CELL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws_m.merge_cells(start_row=i, end_row=i, start_column=1, end_column=5)
        ws_m.row_dimensions[i].height = 18
    ws_m.column_dimensions["A"].width = 100

    wb.save(XLSX)
    print(f"\n✅  Excel scritto: {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
