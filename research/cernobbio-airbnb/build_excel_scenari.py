#!/usr/bin/env python3
"""
Owner-pitch Excel for the Cernobbio "casa tipica La Vignetta" prospect.

Two scenarios:
  A — Mansarda chiusa (~2-3 camere, 4-6 ospiti)
  B — Mansarda aperta (~4-5 camere, 6-8 ospiti)

For each scenario we:
  - filter the comp set on capacity / bedrooms / type
  - boost ranking of listings that look "casa tipica" (stone/historic/wood
    language in the name)
  - present median ADR, range, and the top 10 most-comparable listings
  - extrapolate annual revenue under three occupancy assumptions

Plus a daily-price heatmap across July 2026 and a closing brief.

Output: Cernobbio_Scenari_CasaTipica.xlsx
"""
from __future__ import annotations

import json
import pathlib
import re
import sys
from datetime import datetime, date, timedelta
from statistics import median, mean

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Paths ─────────────────────────────────────────────────────────────────
OUT_DIR = pathlib.Path(__file__).parent
WEEKLY  = OUT_DIR / "july_listings.json"
DAILY   = OUT_DIR / "july_daily.json"
XLSX    = OUT_DIR / "Cernobbio_Scenari_CasaTipica.xlsx"

# ── Brand tokens ──────────────────────────────────────────────────────────
BRAND   = "0C7489"
BRAND_2 = "119DB0"
BAND    = "F0F7F9"
INK     = "0A2540"
GRID    = "DCE7EB"
ACCENT  = "F4A261"   # warm orange for owner-facing emphasis

THIN = Side(style="thin", color=GRID)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=BRAND)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="555555")
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color=BRAND)
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="555555")
KPI_VALUE_FONT = Font(name="Calibri", size=22, bold=True, color=BRAND)
CELL_FONT = Font(name="Calibri", size=10, color=INK)
EMPH_FONT = Font(name="Calibri", size=11, bold=True, color=ACCENT)

# ── Number parsing ────────────────────────────────────────────────────────
def _parse_us(tok):
    if tok is None:
        return None
    try:
        return float(str(tok).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _per_night_from_row(row: dict, nights: int) -> float | None:
    sp = row.get("price") or {}
    if not isinstance(sp, dict):
        return None
    for ln in sp.get("break_down", []) or []:
        if not isinstance(ln, dict):
            continue
        desc = str(ln.get("description") or "")
        m = re.search(r"(?:nights?|notti)\s*x\s*\D*([\d.,]+)", desc, re.I)
        if m:
            v = _parse_us(m.group(1))
            if v:
                return v
    unit = sp.get("unit") or {}
    amt = unit.get("amount") if isinstance(unit, dict) else None
    if isinstance(amt, (int, float)) and amt > 0 and nights > 0:
        return round(amt / nights, 2)
    return None


def _comune(title):
    if not title: return ""
    m = re.search(r"\bin\s+([\w'\-àèìòùÀÈÌÒÙ]+(?:\s+[\w'\-]+)?)\s*$", title.strip())
    return m.group(1) if m else ""


def _tipologia(title):
    if not title: return ""
    return re.sub(r"\s+in\s+.+$", "", title.strip())


def _bed_bath(sr):
    out = {"camere": None, "letti": None, "bagni": None, "ospiti": None}
    sc = sr.get("structuredContent") or {}
    if not isinstance(sc, dict):
        return out
    for line in (sc.get("primaryLine") or []) + (sc.get("secondaryLine") or []):
        if not isinstance(line, dict): continue
        body = str(line.get("body") or "")
        m = re.search(r"([\d.]+)", body)
        if not m: continue
        try: num = float(m.group(1))
        except ValueError: continue
        bl = body.lower()
        if "bedroom" in bl or "camera" in bl: out["camere"] = num
        elif "bath" in bl or "bagn" in bl:    out["bagni"]  = num
        elif "bed" in bl or "letto" in bl:    out["letti"]  = num
        elif "guest" in bl or "sleeps" in bl: out["ospiti"] = num
    return out


# ── "Casa tipica" heuristic ───────────────────────────────────────────────
# Keywords that strongly suggest the listing is positioned as a historic /
# stone / characterful property — the same lane as the owner's "casa
# tipica in pietra, soffitti in legno, cementine". Each hit gives a +1
# affinity score; ≥2 = likely casa tipica, ≥3 = strong match.
TIPICITY_KEYWORDS = [
    "stone", "pietra", "rustic", "rustico",
    "historic", "antico", "storica", "storico",
    "traditional", "tipica", "tipico",
    "village", "borgo", "old town", "centro storico",
    "wood beams", "wooden beams", "travi", "legno",
    "century", "secolo", "1700", "1800", "1900",
    "cottage", "casale", "casa indipendente",
    "courtyard", "corte", "garden", "giardino",
    "panoram", "breathtaking", "lake view", "vista lago",
]


def _typical_score(name: str, title: str) -> tuple[int, list[str]]:
    text = f"{name} {title}".lower()
    hits = [k for k in TIPICITY_KEYWORDS if k in text]
    return len(hits), hits


# ── Aggregation: weekly comp set ──────────────────────────────────────────
def build_weekly_comp_set() -> pd.DataFrame:
    payload = json.loads(WEEKLY.read_text(encoding="utf-8"))
    windows = payload["windows"]
    week_labels = [w["label"] for w in windows]
    per_room: dict[int, dict] = {}

    for w in windows:
        try:
            nights = (datetime.fromisoformat(w["check_out"]) - datetime.fromisoformat(w["check_in"])).days or 7
        except Exception:
            nights = 7
        for row in w.get("listings") or []:
            rid = row.get("room_id")
            if not rid: continue
            rec = per_room.setdefault(rid, {
                "_row": row,
                "weekly": {l: None for l in week_labels},
            })
            rec["weekly"][w["label"]] = _per_night_from_row(row, nights=nights)

    rows = []
    for rid, rec in per_room.items():
        sr = rec["_row"]
        bb = _bed_bath(sr)
        name = sr.get("name") or ""
        title = sr.get("title") or ""
        score, hits = _typical_score(name, title)
        coords = sr.get("coordinates") or {}
        lat = coords.get("latitude") if isinstance(coords, dict) else None
        lng = (coords.get("longitude") or coords.get("longitud")) if isinstance(coords, dict) else None
        rating_obj = sr.get("rating") or {}
        rating_value = rating_obj.get("value") if isinstance(rating_obj, dict) else None
        review_count = rating_obj.get("reviewCount") if isinstance(rating_obj, dict) else None
        if rating_value == 0: rating_value = None
        if review_count == 0: review_count = None

        weekly_prices = [rec["weekly"][lab] for lab in week_labels]
        seen = [p for p in weekly_prices if p is not None]
        rows.append({
            "ID Airbnb": rid,
            "Nome": name[:120],
            "Tipologia": _tipologia(title),
            "Comune": _comune(title),
            "Camere": bb["camere"],
            "Letti": bb["letti"],
            "Bagni": bb["bagni"],
            "Ospiti": bb["ospiti"],
            "Prezzo medio €/notte (luglio)": round(mean(seen), 2) if seen else None,
            "Prezzo mediano €/notte (luglio)": round(median(seen), 2) if seen else None,
            "N. settimane disponibili": len(seen),
            "Rating": rating_value,
            "N. recensioni": review_count,
            "Affinità casa tipica": score,
            "Indizi (parole chiave)": ", ".join(hits[:6]),
            "Latitudine": lat,
            "Longitudine": lng,
            "Link": f"https://www.airbnb.com/rooms/{rid}",
        })

    return pd.DataFrame(rows)


# ── Daily matrix ──────────────────────────────────────────────────────────
def build_daily_matrix() -> pd.DataFrame:
    """
    One row per listing × 30 columns (Jul 1 → Jul 30 starting days).
    Each cell = per-night price as quoted by Airbnb for a 2-night stay
    starting that day.
    """
    if not DAILY.exists():
        print("⚠️  july_daily.json missing — daily matrix will be empty. "
              "Run `python scrape_daily.py` first.", file=sys.stderr)
        return pd.DataFrame()

    payload = json.loads(DAILY.read_text(encoding="utf-8"))
    days = payload.get("days") or []
    day_labels = [d["start"] for d in days]  # ISO dates as column names

    per_room: dict[int, dict] = {}
    for d in days:
        for row in d.get("listings") or []:
            rid = row.get("room_id")
            if not rid: continue
            p = _per_night_from_row(row, nights=2)
            rec = per_room.setdefault(rid, {
                "_row": row,
                "daily": {lab: None for lab in day_labels},
            })
            rec["daily"][d["start"]] = p

    rows = []
    for rid, rec in per_room.items():
        sr = rec["_row"]
        bb = _bed_bath(sr)
        name = sr.get("name") or ""
        title = sr.get("title") or ""
        score, _ = _typical_score(name, title)
        out = {
            "ID Airbnb": rid,
            "Nome": name[:80],
            "Tipologia": _tipologia(title),
            "Comune": _comune(title),
            "Camere": bb["camere"],
            "Ospiti": bb["ospiti"],
            "Affinità tipica": score,
        }
        out.update({lab: rec["daily"][lab] for lab in day_labels})
        seen = [v for v in rec["daily"].values() if v is not None]
        out["Media giorni"] = round(mean(seen), 2) if seen else None
        out["Giorni visti"] = len(seen)
        rows.append(out)

    return pd.DataFrame(rows)


# ── Scenario filters ──────────────────────────────────────────────────────
SCENARIO_A = {
    "label":            "Scenario A — Mansarda chiusa",
    "subtitle":         "Layout: piano terra (cucina+bagno+ripostiglio) + piano 1 (1 camera + 1 soggiorno o 2ª camera). 4 ospiti tipici.",
    "camere_range":     (2, 3),
    "ospiti_range":     (3, 6),
    "tipi_ammessi":     ["Apartment", "Home", "Condo", "Vacation home", "Cottage", "Townhouse"],
    "comuni_priority":  ["Cernobbio", "Moltrasio", "Como", "Blevio"],
}
SCENARIO_B = {
    "label":            "Scenario B — Mansarda aperta",
    "subtitle":         "Layout pieno: piano terra + piano 1 + mansarda 2 camere. 6-8 ospiti.",
    "camere_range":     (4, 6),
    "ospiti_range":     (6, 10),
    "tipi_ammessi":     ["Home", "Apartment", "Vacation home", "Cottage", "Townhouse", "Villa"],
    "comuni_priority":  ["Cernobbio", "Moltrasio", "Como", "Blevio"],
}


def filter_for_scenario(df: pd.DataFrame, sc: dict) -> pd.DataFrame:
    cmin, cmax = sc["camere_range"]
    omin, omax = sc["ospiti_range"]
    f = df.copy()

    # Camere filter — accept if camere falls in range OR camere unknown but
    # ospiti falls in range (we trust ospiti as proxy)
    cond_camere = f["Camere"].between(cmin, cmax, inclusive="both")
    cond_ospiti = f["Ospiti"].between(omin, omax, inclusive="both")
    cond_cap = cond_camere.fillna(False) | (f["Camere"].isna() & cond_ospiti.fillna(False))

    cond_type = f["Tipologia"].isin(sc["tipi_ammessi"])
    cond_comune = f["Comune"].isin(sc["comuni_priority"])

    f = f[cond_cap & cond_type & cond_comune].copy()
    # Rank by affinity score first, then price ascending (cheapest "tipica"
    # houses are the best illustrative comp)
    f = f.sort_values(
        by=["Affinità casa tipica", "Prezzo mediano €/notte (luglio)"],
        ascending=[False, True],
    ).reset_index(drop=True)
    return f


# ── Excel write helpers (reused style) ────────────────────────────────────
def _style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = 28


def _band_rows(ws, start_row, end_row, n_cols):
    fill_alt = PatternFill(start_color=BAND, end_color=BAND, fill_type="solid")
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fill_type is None or (cell.fill.start_color.rgb in (None, "00000000")):
                    cell.fill = fill_alt
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL
            if cell.font is None or cell.font.color is None or cell.font.color.rgb is None:
                cell.font = CELL_FONT


def _autosize(ws, df, start_col=1, max_width=42):
    for idx, col in enumerate(df.columns):
        col_str = [str(col)]
        for v in df[col].head(50).tolist():
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            col_str.append(str(v))
        max_len = max(len(s) for s in col_str)
        ws.column_dimensions[get_column_letter(start_col + idx)].width = min(max_len + 3, max_width)


def write_df(ws, df, start_row=1, title=None, subtitle=None):
    cursor = start_row
    if title:
        ws.cell(row=cursor, column=1, value=title).font = TITLE_FONT
        ws.merge_cells(start_row=cursor, end_row=cursor, start_column=1,
                       end_column=max(len(df.columns), 4))
        cursor += 1
    if subtitle:
        ws.cell(row=cursor, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=cursor, end_row=cursor, start_column=1,
                       end_column=max(len(df.columns), 4))
        cursor += 1
    if title or subtitle:
        cursor += 1
    header_row = cursor
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=j, value=col)
    _style_header_row(ws, header_row, len(df.columns))
    cursor += 1
    body_start = cursor
    for _, r in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if pd.isna(v): v = None
            ws.cell(row=cursor, column=j, value=v)
        cursor += 1
    body_end = cursor - 1
    _band_rows(ws, body_start, body_end, len(df.columns))
    _autosize(ws, df)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return cursor + 1, header_row, body_start, body_end


# ── Revenue model ─────────────────────────────────────────────────────────
def _revenue_table(adr_median: float) -> pd.DataFrame:
    """
    Crude 12-month revenue model:
      - July ADR = adr_median (peak baseline)
      - Seasonality multipliers vs July baseline (italian lake market norms):
          high (Jun/Jul/Aug):    1.00
          shoulder+ (May/Sep):   0.78
          shoulder (Apr/Oct):    0.62
          xmas/NY (last 10 days Dec): 0.92
          low (rest):            0.48
      - 3 occupancy scenarios: 50%, 65%, 75%
    """
    if not adr_median or adr_median <= 0:
        return pd.DataFrame()

    months = [
        ("Gennaio",   31, 0.48),
        ("Febbraio",  28, 0.48),
        ("Marzo",     31, 0.55),
        ("Aprile",    30, 0.62),
        ("Maggio",    31, 0.78),
        ("Giugno",    30, 1.00),
        ("Luglio",    31, 1.00),
        ("Agosto",    31, 1.00),
        ("Settembre", 30, 0.78),
        ("Ottobre",   31, 0.62),
        ("Novembre",  30, 0.48),
        ("Dicembre",  31, 0.68),  # weighted xmas
    ]

    rows = []
    totals = {"50%": 0.0, "65%": 0.0, "75%": 0.0}
    for name, n_days, mult in months:
        month_adr = adr_median * mult
        for occ_label, occ in [("50%", 0.50), ("65%", 0.65), ("75%", 0.75)]:
            totals[occ_label] += month_adr * n_days * occ
        rows.append({
            "Mese": name,
            "Giorni": n_days,
            "ADR stimato €": round(month_adr, 0),
            "Ricavo @ occ. 50%": round(month_adr * n_days * 0.50, 0),
            "Ricavo @ occ. 65%": round(month_adr * n_days * 0.65, 0),
            "Ricavo @ occ. 75%": round(month_adr * n_days * 0.75, 0),
        })
    rows.append({
        "Mese": "TOTALE ANNO",
        "Giorni": 365,
        "ADR stimato €": round(adr_median, 0),
        "Ricavo @ occ. 50%": round(totals["50%"], 0),
        "Ricavo @ occ. 65%": round(totals["65%"], 0),
        "Ricavo @ occ. 75%": round(totals["75%"], 0),
    })
    return pd.DataFrame(rows)


# ── Main ──────────────────────────────────────────────────────────────────
def main() -> int:
    if not WEEKLY.exists():
        print(f"missing {WEEKLY} — run scrape_july.py first", file=sys.stderr)
        return 1

    print("→ Building weekly comp set…")
    df_all = build_weekly_comp_set()
    print(f"  {len(df_all)} unique listings in weekly data")

    print("→ Building daily matrix…")
    df_daily = build_daily_matrix()
    print(f"  {len(df_daily)} unique listings in daily data")

    print("→ Filtering scenarios…")
    df_a = filter_for_scenario(df_all, SCENARIO_A)
    df_b = filter_for_scenario(df_all, SCENARIO_B)
    print(f"  Scenario A: {len(df_a)} comparable listings")
    print(f"  Scenario B: {len(df_b)} comparable listings")

    # Compute scenario KPIs
    def _kpi(df):
        prices = df["Prezzo mediano €/notte (luglio)"].dropna().tolist()
        prices.sort()
        n = len(prices)
        return {
            "n": len(df),
            "median": median(prices) if prices else None,
            "mean":   mean(prices) if prices else None,
            "p25":    prices[n//4] if n else None,
            "p75":    prices[3*n//4] if n else None,
            "min":    prices[0] if n else None,
            "max":    prices[-1] if n else None,
        }

    kpi_a = _kpi(df_a)
    kpi_b = _kpi(df_b)

    # ── Build workbook ────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Brief proprietario
    ws = wb.active
    ws.title = "Brief"
    ws.cell(row=1, column=1, value="Casa tipica Cernobbio — Brief di mercato").font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1, value=f"Inquadramento competitivo per i due scenari · luglio 2026 baseline").font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    ws.cell(row=4, column=1, value="Profilo immobile").font = SECTION_FONT
    profile_rows = [
        ["Località",          "Cernobbio, zona La Vignetta (pre-collinare)"],
        ["Caratteristiche",   "Casa indipendente in corte privata, pietra, soffitti legno massiccio, parquet, cementine originali"],
        ["Sviluppo",          "2 piani + mansarda"],
        ["Scenario A",        "Mansarda chiusa → 2-3 camere, 4 ospiti"],
        ["Scenario B",        "Mansarda aperta → 4-5 camere, 6-8 ospiti"],
    ]
    for i, (k, v) in enumerate(profile_rows, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, color=INK)
        ws.cell(row=i, column=2, value=v).font = CELL_FONT
        ws.merge_cells(start_row=i, end_row=i, start_column=2, end_column=6)

    ws.cell(row=11, column=1, value="Sintesi prezzi luglio 2026 — listing comparabili").font = SECTION_FONT
    kpi_df = pd.DataFrame([
        {
            "Scenario": SCENARIO_A["label"],
            "N. listing comparabili": kpi_a["n"],
            "Mediana €/notte": round(kpi_a["median"], 0) if kpi_a["median"] else None,
            "Mean €/notte":    round(kpi_a["mean"], 0) if kpi_a["mean"] else None,
            "p25 €/notte":     round(kpi_a["p25"], 0) if kpi_a["p25"] else None,
            "p75 €/notte":     round(kpi_a["p75"], 0) if kpi_a["p75"] else None,
            "Range €/notte":   f"{int(kpi_a['min'] or 0)} – {int(kpi_a['max'] or 0)}",
        },
        {
            "Scenario": SCENARIO_B["label"],
            "N. listing comparabili": kpi_b["n"],
            "Mediana €/notte": round(kpi_b["median"], 0) if kpi_b["median"] else None,
            "Mean €/notte":    round(kpi_b["mean"], 0) if kpi_b["mean"] else None,
            "p25 €/notte":     round(kpi_b["p25"], 0) if kpi_b["p25"] else None,
            "p75 €/notte":     round(kpi_b["p75"], 0) if kpi_b["p75"] else None,
            "Range €/notte":   f"{int(kpi_b['min'] or 0)} – {int(kpi_b['max'] or 0)}",
        },
    ])
    write_df(ws, kpi_df, start_row=12)

    ws.cell(row=19, column=1, value="Lettura della stima").font = SECTION_FONT
    notes = [
        "• I dati partono da uno snapshot luglio 2026 (alta stagione). Per l'anno intero applicare la stagionalità nel foglio “Ricavi annui”.",
        "• La mediana è il riferimento più solido (ignora gli outlier luxury). Il p25 indica un posizionamento prudente, il p75 quello premium.",
        f"• Scenario A → fascia €{int(kpi_a['p25'] or 0)}-{int(kpi_a['p75'] or 0)}/n in alta stagione. Lo posizionerei verso il p75 grazie alla tipicità della casa (pietra + cementine + corte privata = forte differenziale).",
        f"• Scenario B → fascia €{int(kpi_b['p25'] or 0)}-{int(kpi_b['p75'] or 0)}/n. La mansarda aperta permette di catturare gruppi 6-8 (matrimoni Villa d'Este, weekend di luglio/agosto). +50-70% ricavi vs scenario A in alta stagione.",
        "• I “Listing comparabili” nei due fogli successivi sono ordinati per affinità (pietra/legno/storico nel nome) e poi per prezzo — quelli in alto sono il vero comp set.",
    ]
    for i, n in enumerate(notes, start=20):
        c = ws.cell(row=i, column=1, value=n)
        c.font = CELL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=6)
        ws.row_dimensions[i].height = 32
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    for c in "CDEF":
        ws.column_dimensions[c].width = 22

    # Sheet 2: Scenario A — Comparabili
    ws_a = wb.create_sheet("Scen. A — Comparabili")
    display_cols = [
        "Nome", "Tipologia", "Comune", "Camere", "Letti", "Bagni", "Ospiti",
        "Prezzo mediano €/notte (luglio)", "Prezzo medio €/notte (luglio)",
        "Rating", "N. recensioni",
        "Affinità casa tipica", "Indizi (parole chiave)",
        "Link",
    ]
    df_a_show = df_a[display_cols].head(40)
    write_df(
        ws_a, df_a_show, start_row=1,
        title=SCENARIO_A["label"],
        subtitle=f"{SCENARIO_A['subtitle']} · {kpi_a['n']} listing comparabili trovati · mediana €{int(kpi_a['median'] or 0)}/notte",
    )

    # Sheet 3: Scenario B — Comparabili
    ws_b = wb.create_sheet("Scen. B — Comparabili")
    df_b_show = df_b[display_cols].head(40)
    write_df(
        ws_b, df_b_show, start_row=1,
        title=SCENARIO_B["label"],
        subtitle=f"{SCENARIO_B['subtitle']} · {kpi_b['n']} listing comparabili trovati · mediana €{int(kpi_b['median'] or 0)}/notte",
    )

    # Sheet 4: Ricavi annui (entrambi gli scenari)
    ws_rev = wb.create_sheet("Ricavi annui")
    ws_rev.cell(row=1, column=1, value="Stima ricavi annui — proiezione su occupancy 50/65/75%").font = TITLE_FONT
    ws_rev.merge_cells("A1:G1")
    ws_rev.cell(row=2, column=1, value=(
        "Modello: ADR luglio × moltiplicatore stagionale × giorni × occupancy. "
        "Lordo, no commissioni piattaforma, no tasse di soggiorno."
    )).font = SUBTITLE_FONT
    ws_rev.merge_cells("A2:G2")

    if kpi_a["median"]:
        ws_rev.cell(row=4, column=1, value=f"{SCENARIO_A['label']} — ADR baseline €{int(kpi_a['median'])}").font = SECTION_FONT
        write_df(ws_rev, _revenue_table(kpi_a["median"]), start_row=5)
    if kpi_b["median"]:
        next_start = 4 + 16  # after scenario A table
        ws_rev.cell(row=next_start, column=1, value=f"{SCENARIO_B['label']} — ADR baseline €{int(kpi_b['median'])}").font = SECTION_FONT
        write_df(ws_rev, _revenue_table(kpi_b["median"]), start_row=next_start + 1)

    # Sheet 5: Prezzi giornalieri luglio (full matrix)
    if not df_daily.empty:
        ws_d = wb.create_sheet("Giornaliero luglio")
        # Order columns: meta first, then dates sorted, then totals
        meta_cols = ["Nome", "Tipologia", "Comune", "Camere", "Ospiti",
                     "Affinità tipica", "Media giorni", "Giorni visti"]
        date_cols = sorted([c for c in df_daily.columns if c.startswith("2026-")])
        ordered = [c for c in meta_cols if c in df_daily.columns] + date_cols
        df_d_sorted = df_daily[ordered].sort_values(
            by="Media giorni", ascending=True, na_position="last"
        ).head(80)
        _, hr, bstart, bend = write_df(
            ws_d, df_d_sorted, start_row=1,
            title="Prezzi giornalieri €/notte — Luglio 2026",
            subtitle=f"{len(df_daily)} listing osservati su 30 sliding 2-night windows (1-30 lug). Top 80 per prezzo medio mostrati. Heatmap su tutte le date.",
        )
        # Heatmap on date columns
        for col_name in date_cols:
            col_idx = ordered.index(col_name) + 1
            letter = get_column_letter(col_idx)
            rng = f"{letter}{bstart}:{letter}{bend}"
            rule = ColorScaleRule(
                start_type="percentile", start_value=10, start_color="C6E6CB",
                mid_type="percentile",   mid_value=50,   mid_color="FFFFFF",
                end_type="percentile",   end_value=90,   end_color="F4C7C3",
            )
            ws_d.conditional_formatting.add(rng, rule)

        # Sheet 6: daily trend (aggregate)
        ws_t = wb.create_sheet("Trend giornaliero")
        trend_rows = []
        for d_col in date_cols:
            vals = df_daily[d_col].dropna().tolist()
            if not vals: continue
            vals.sort()
            n = len(vals)
            try:
                dt = date.fromisoformat(d_col)
                dow = ["Lun","Mar","Mer","Gio","Ven","Sab","Dom"][dt.weekday()]
            except Exception:
                dow = ""
            trend_rows.append({
                "Data": d_col,
                "Giorno": dow,
                "N. listing": n,
                "Mediana €/n": round(median(vals), 0),
                "Mean €/n":    round(mean(vals), 0),
                "p25 €/n":     round(vals[n//4], 0),
                "p75 €/n":     round(vals[3*n//4], 0),
                "Min €/n":     round(vals[0], 0),
                "Max €/n":     round(vals[-1], 0),
            })
        write_df(
            ws_t, pd.DataFrame(trend_rows), start_row=1,
            title="Andamento giornaliero prezzo / notte — luglio 2026",
            subtitle="Aggregato su tutti i listing visti per ciascuna data di check-in (2-night sliding window).",
        )

    # Sheet 7: Metodologia
    ws_m = wb.create_sheet("Metodologia")
    ws_m.cell(row=1, column=1, value="Metodologia").font = TITLE_FONT
    ws_m.merge_cells("A1:E1")
    notes_meth = [
        "Fonte dati: Airbnb StaysSearch (API interna), accesso via pyairbnb (TLS impersonation Chrome).",
        "Bbox geografico: Cernobbio + frazioni primo bacino (lat 45.825-45.870 × lng 9.060-9.110). Include parti di Blevio, Moltrasio, Brunate, Como.",
        "Snapshot weekly: 5 finestre Mer-Mer coprenti tutto luglio 2026 + prima settimana di agosto (153, 135, 113, 117, 145 listing per finestra).",
        "Snapshot daily: 30 finestre di 2 notti (sliding, ogni giorno di luglio come start day).",
        "",
        "Filtro Scenario A (mansarda chiusa):",
        "  • Camere: 2-3 (oppure ospiti 3-6 se camere non note)",
        "  • Ospiti: 3-6",
        "  • Tipologia ammesse: Apartment, Home, Condo, Vacation home, Cottage, Townhouse",
        "  • Comune: Cernobbio, Moltrasio, Como, Blevio (zona competitiva primo bacino)",
        "",
        "Filtro Scenario B (mansarda aperta):",
        "  • Camere: 4-6 (oppure ospiti 6-10)",
        "  • Ospiti: 6-10",
        "  • Tipologia ammesse: Home, Apartment, Vacation home, Cottage, Townhouse, Villa",
        "  • Comune: stessi del A",
        "",
        "Punteggio affinità casa tipica (0-N):",
        "  Conta le parole chiave “stone, pietra, rustic, historic, antico, traditional, tipica, village, borgo,",
        "  wood beams, travi, legno, century, cottage, casa indipendente, courtyard, corte, garden, panoram,",
        "  lake view…” presenti nel titolo/nome del listing. ≥2 hit = candidato “casa tipica”.",
        "",
        "Modello ricavi annui:",
        "  Mensile: ADR baseline luglio × moltiplicatore stagionale × giorni × occupancy.",
        "  Moltiplicatori: alta stagione (giu/lug/ago) = 1.00, shoulder+ (mag/set) = 0.78,",
        "    shoulder (apr/ott) = 0.62, dicembre (weighted xmas) = 0.68, bassa (gen/feb/nov) = 0.48.",
        "  Occupancy presentate: 50% (anno 1 conservativo), 65% (cruise), 75% (top performer).",
        "",
        "Caveat onesti:",
        "  • I prezzi sono rack rate Airbnb dinamici al momento dello scraping — non confermati come booking.",
        "  • Il filtro “tipologia” usa il tag che Airbnb assegna automaticamente (Apartment/Home/Villa). Falsi positivi possibili.",
        "  • Il giornaliero rappresenta il prezzo di una stay che inizia in quel giorno — non il prezzo applicato a quel singolo giorno in una stay più lunga.",
        "  • Occupancy 50/65/75% sono ipotesi; il dato vero deve venire da AirDNA o da uno storico Smoobu.",
    ]
    for i, n in enumerate(notes_meth, start=3):
        c = ws_m.cell(row=i, column=1, value=n)
        c.font = CELL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws_m.merge_cells(start_row=i, end_row=i, start_column=1, end_column=5)
        if not n.startswith("  "):
            c.font = Font(name="Calibri", size=10, color=INK, bold=n[:1].isupper() and ":" in n)
    ws_m.column_dimensions["A"].width = 30
    for c in "BCDE":
        ws_m.column_dimensions[c].width = 22

    # Save
    wb.save(XLSX)
    print(f"\n✅  Excel scritto: {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
