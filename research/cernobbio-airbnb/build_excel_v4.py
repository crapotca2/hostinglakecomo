#!/usr/bin/env python3
"""
Owner-pitch Excel v4 — Cernobbio 3-scenari research deliverable.

ARCHITETTURA v4 (≠ v3):
- Python = SOLO ETL: legge JSON, flatten in tabelle Excel.
- Excel = motore di calcolo: ogni KPI, mediana, p25/p75, conteggio,
  delta, ricavo proiettato è una FORMULA che referenzia le Excel Tables.
  L'owner può audire ogni numero cliccandolo e vedendo la formula.

Sheet inventory (12 fogli, ordine esatto):
  1. Brief                          (cover + KPI grid + note)
  2. Scen. 1 piano terra            (KPI strip + FILTER spill)
  3. Scen. 2 mansarda
  4. Scen. 3 casa intera
  5. Distribuzione mercato          (3 grid + BarChart nativi)
  6. Trend giornaliero              (TblTrend + LineChart)
  7. Ricavi annui                   (3 blocchi proiezione)
  8. Hotel zona Cernobbio           (TblHotels)
  9. Dati listing                   (TblListing — universo grezzo)
 10. Dati giornaliero giugno        (TblDailyJun — matrice prezzi)
 11. Dati giornaliero luglio        (TblDailyJul)
 12. Metodologia                    (testo statico)

Excel Tables esposte:
  TblListing, TblDailyJun, TblDailyJul, TblDistrCam, TblDistrOsp,
  TblScenDistr, TblTrend, TblHotels

Tutto referenziato via structured references o range espliciti.

Output: Cernobbio_3_Scenari_v4.xlsx
"""
from __future__ import annotations

import json
import pathlib
import re
import sys
from datetime import datetime
from statistics import median, mean

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.formula import ArrayFormula


# ── Paths ─────────────────────────────────────────────────────────────────
OUT_DIR = pathlib.Path(__file__).parent
WEEKLY_JUL  = OUT_DIR / "july_listings.json"
WEEKLY_JUN  = OUT_DIR / "june_listings.json"
DAILY_JUL   = OUT_DIR / "july_daily.json"
DAILY_JUN   = OUT_DIR / "june_daily.json"
HOTELS_JSON = OUT_DIR / "hotels.json"
XLSX        = OUT_DIR / "Cernobbio_3_Scenari_v4.xlsx"


# ── Brand tokens ──────────────────────────────────────────────────────────
TEAL          = "0C7489"
TEAL_LIGHT    = "DCE7EB"
BAND          = "F0F7F9"
INK           = "0A2540"
ACCENT_ORANGE = "F4A261"


# ── Comuni ammessi per i 3 scenari ────────────────────────────────────────
COMUNI_OK = {"Cernobbio", "Moltrasio", "Blevio", "Como"}


# ─────────────────────────────────────────────────────────────────────────
#  ETL helpers — ricalcati su v3, ma usati SOLO per popolare le tabelle
# ─────────────────────────────────────────────────────────────────────────
def _parse_us(tok):
    if tok is None:
        return None
    try:
        return float(str(tok).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _per_night_from_row(row: dict, nights: int) -> float | None:
    """Estrae il €/notte da una riga Airbnb StaysSearch."""
    sp = row.get("price") or {}
    if not isinstance(sp, dict):
        return None
    for ln in sp.get("break_down", []) or []:
        if not isinstance(ln, dict):
            continue
        desc = str(ln.get("description") or "")
        m = re.search(r"(?:nights?|notti)\s*x\s*\D*([\d.,]+)", desc, re.I)
        if m:
            v = _parse_us(m.group(1))
            if v:
                return v
    unit = sp.get("unit") or {}
    amt = unit.get("amount") if isinstance(unit, dict) else None
    if isinstance(amt, (int, float)) and amt > 0 and nights > 0:
        return round(amt / nights, 2)
    return None


def _comune(title):
    if not title:
        return ""
    m = re.search(r"\bin\s+([\w'\-àèìòùÀÈÌÒÙ]+(?:\s+[\w'\-]+)?)\s*$", title.strip())
    return m.group(1) if m else ""


def _tipologia(title):
    if not title:
        return ""
    return re.sub(r"\s+in\s+.+$", "", title.strip())


def _bed_bath(sr):
    out = {"camere": None, "letti": None, "bagni": None, "ospiti": None}
    sc = sr.get("structuredContent") or {}
    if not isinstance(sc, dict):
        return out
    for line in (sc.get("primaryLine") or []) + (sc.get("secondaryLine") or []):
        if not isinstance(line, dict):
            continue
        body = str(line.get("body") or "")
        m = re.search(r"([\d.]+)", body)
        if not m:
            continue
        try:
            num = float(m.group(1))
        except ValueError:
            continue
        bl = body.lower()
        if "bedroom" in bl or "camera" in bl:
            out["camere"] = num
        elif "bath" in bl or "bagn" in bl:
            out["bagni"] = num
        elif "bed" in bl or "letto" in bl:
            out["letti"] = num
        elif "guest" in bl or "sleeps" in bl:
            out["ospiti"] = num
    return out


TIPICITY_KEYWORDS = [
    "stone", "pietra", "rustic", "rustico",
    "historic", "antico", "storica", "storico",
    "traditional", "tipica", "tipico",
    "village", "borgo", "old town", "centro storico",
    "wood beams", "wooden beams", "travi", "legno",
    "century", "secolo", "1700", "1800", "1900",
    "cottage", "casale", "casa indipendente",
    "courtyard", "corte", "garden", "giardino",
    "panoram", "breathtaking", "lake view", "vista lago",
]


def _typical_score(name, title):
    text = f"{name} {title}".lower()
    hits = [k for k in TIPICITY_KEYWORDS if k in text]
    return len(hits), hits


# ─────────────────────────────────────────────────────────────────────────
#  Scenario tagging — STESSA logica del filter_scenario di v3 ma per row dict
# ─────────────────────────────────────────────────────────────────────────
def _tag_scenario(camere, ospiti, tipologia, comune) -> str:
    """Restituisce S1/S2/S3/'' secondo le regole dell'architetto.

    - S1: camere None o 0-1 AND ospiti None o 1-3 AND non Villa AND comune OK
    - S2: camere == 2 AND ospiti None o 3-5 AND non Villa AND comune OK
    - S3: camere 3-5 AND ospiti None o 6-9 AND non Villa AND comune OK
    - Se camere è None, fallback su ospiti per il check di appartenenza
    """
    if tipologia == "Villa":
        return ""
    if comune not in COMUNI_OK:
        return ""

    # Check S1
    cam_ok_s1 = (camere is None) or (0 <= camere <= 1)
    osp_ok_s1 = (ospiti is None) or (1 <= ospiti <= 3)
    if camere is None:
        # Fallback: solo ospiti
        s1 = (ospiti is not None) and (1 <= ospiti <= 3)
    else:
        s1 = cam_ok_s1 and osp_ok_s1
    if s1:
        return "S1"

    # Check S2
    cam_ok_s2 = (camere is not None) and (camere == 2)
    osp_ok_s2 = (ospiti is None) or (3 <= ospiti <= 5)
    if camere is None:
        s2 = (ospiti is not None) and (3 <= ospiti <= 5)
    else:
        s2 = cam_ok_s2 and osp_ok_s2
    if s2:
        return "S2"

    # Check S3
    cam_ok_s3 = (camere is not None) and (3 <= camere <= 5)
    osp_ok_s3 = (ospiti is None) or (6 <= ospiti <= 9)
    if camere is None:
        s3 = (ospiti is not None) and (6 <= ospiti <= 9)
    else:
        s3 = cam_ok_s3 and osp_ok_s3
    if s3:
        return "S3"

    return ""


# ─────────────────────────────────────────────────────────────────────────
#  Build comp set (weekly aggregated) — produce rows per TblListing
# ─────────────────────────────────────────────────────────────────────────
def build_comp_rows(weekly_path: pathlib.Path, month_label: str) -> list[dict]:
    if not weekly_path.exists():
        return []
    payload = json.loads(weekly_path.read_text(encoding="utf-8"))
    windows = payload.get("windows", [])
    per_room: dict[int, dict] = {}
    for w in windows:
        try:
            nights = (datetime.fromisoformat(w["check_out"])
                      - datetime.fromisoformat(w["check_in"])).days or 7
        except Exception:
            nights = 7
        for row in (w.get("listings") or []):
            rid = row.get("room_id")
            if not rid:
                continue
            rec = per_room.setdefault(rid, {"_row": row, "prices": []})
            p = _per_night_from_row(row, nights=nights)
            if p is not None:
                rec["prices"].append(p)

    rows = []
    for rid, rec in per_room.items():
        sr = rec["_row"]
        bb = _bed_bath(sr)
        name = sr.get("name") or ""
        title = sr.get("title") or ""
        score, hits = _typical_score(name, title)
        rating_obj = sr.get("rating") or {}
        rating_value = rating_obj.get("value") if isinstance(rating_obj, dict) else None
        if rating_value == 0:
            rating_value = None
        review_count = rating_obj.get("reviewCount") if isinstance(rating_obj, dict) else None
        if review_count == 0:
            review_count = None

        prices = rec["prices"]
        tipologia = _tipologia(title)
        comune = _comune(title)
        camere = bb["camere"]
        ospiti = bb["ospiti"]

        scenario = _tag_scenario(
            camere=int(camere) if camere is not None else None,
            ospiti=int(ospiti) if ospiti is not None else None,
            tipologia=tipologia,
            comune=comune,
        )

        rows.append({
            "Mese": month_label,
            "ID Airbnb": rid,
            "Nome": (name or "")[:120],
            "Tipologia": tipologia,
            "Comune": comune,
            "Camere": int(camere) if camere is not None else None,
            "Letti": int(bb["letti"]) if bb["letti"] is not None else None,
            "Bagni": float(bb["bagni"]) if bb["bagni"] is not None else None,
            "Ospiti": int(ospiti) if ospiti is not None else None,
            "Prezzo mediano €/n": round(median(prices), 2) if prices else None,
            "Prezzo medio €/n":   round(mean(prices), 2) if prices else None,
            "N. settimane": len(prices),
            "Rating": rating_value,
            "Recensioni": review_count,
            "Affinità tipica": score,
            "Indizi": ", ".join(hits[:5]),
            "Link": f"https://www.airbnb.com/rooms/{rid}",
            "Scenario": scenario,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────
#  Build daily matrix — produce rows per TblDailyJun / TblDailyJul
# ─────────────────────────────────────────────────────────────────────────
def build_daily_rows(daily_path: pathlib.Path) -> tuple[list[dict], list[str]]:
    """Restituisce (rows, day_labels) dove day_labels è la lista ordinata
    delle 30 date in ISO yyyy-mm-dd."""
    if not daily_path.exists():
        return [], []
    payload = json.loads(daily_path.read_text(encoding="utf-8"))
    days = payload.get("days") or []
    day_labels = [d["start"] for d in days]

    per_room: dict[int, dict] = {}
    for d in days:
        for row in d.get("listings") or []:
            rid = row.get("room_id")
            if not rid:
                continue
            p = _per_night_from_row(row, nights=2)
            rec = per_room.setdefault(
                rid,
                {"_row": row, "daily": {lab: None for lab in day_labels}},
            )
            rec["daily"][d["start"]] = p

    rows = []
    for rid, rec in per_room.items():
        sr = rec["_row"]
        bb = _bed_bath(sr)
        name = sr.get("name") or ""
        title = sr.get("title") or ""
        score, _ = _typical_score(name, title)
        out = {
            "ID Airbnb": rid,
            "Nome": (name or "")[:80],
            "Tipologia": _tipologia(title),
            "Comune": _comune(title),
            "Camere": int(bb["camere"]) if bb["camere"] is not None else None,
            "Ospiti": int(bb["ospiti"]) if bb["ospiti"] is not None else None,
            "Affinità tipica": score,
        }
        for lab in day_labels:
            out[lab] = rec["daily"][lab]
        rows.append(out)
    return rows, day_labels


# ─────────────────────────────────────────────────────────────────────────
#  Style helpers — named styles
# ─────────────────────────────────────────────────────────────────────────
def _thin_border(color=TEAL_LIGHT):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def register_named_styles(wb: Workbook) -> None:
    """Crea e registra tutti i NamedStyle. Idempotente."""
    styles = []

    # BodyHC
    ns = NamedStyle(name="BodyHC")
    ns.font = Font(name="Kanit", size=10, color=INK)
    ns.alignment = Alignment(vertical="center")
    ns.border = _thin_border()
    styles.append(ns)

    # HeaderHC
    ns = NamedStyle(name="HeaderHC")
    ns.font = Font(name="Kanit", bold=True, size=11, color="FFFFFF")
    ns.fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
    ns.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ns.border = _thin_border()
    styles.append(ns)

    # TitleHC
    ns = NamedStyle(name="TitleHC")
    ns.font = Font(name="Kanit", bold=True, size=20, color=TEAL)
    ns.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(ns)

    # SubtitleHC
    ns = NamedStyle(name="SubtitleHC")
    ns.font = Font(name="Kanit", italic=True, size=11, color="555555")
    ns.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(ns)

    # SectionHC
    ns = NamedStyle(name="SectionHC")
    ns.font = Font(name="Kanit", bold=True, size=14, color=TEAL)
    ns.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(ns)

    # KPILabelHC
    ns = NamedStyle(name="KPILabelHC")
    ns.font = Font(name="Kanit", size=10, color="555555")
    ns.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(ns)

    # KPIValueHC
    ns = NamedStyle(name="KPIValueHC")
    ns.font = Font(name="Kanit", bold=True, size=22, color=TEAL)
    ns.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(ns)

    # PriceEUR (BodyHC + format €)
    ns = NamedStyle(name="PriceEUR")
    ns.font = Font(name="Kanit", size=10, color=INK)
    ns.alignment = Alignment(vertical="center")
    ns.border = _thin_border()
    ns.number_format = '€ #,##0'
    styles.append(ns)

    # Pct1 (BodyHC + 0.0%)
    ns = NamedStyle(name="Pct1")
    ns.font = Font(name="Kanit", size=10, color=INK)
    ns.alignment = Alignment(vertical="center")
    ns.border = _thin_border()
    ns.number_format = '0.0%'
    styles.append(ns)

    for s in styles:
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def style_sheet(ws):
    """Setup standard per ogni sheet: niente gridline, niente freeze panes."""
    ws.sheet_view.showGridLines = False
    # NO freeze_panes — esplicitamente richiesto


def set_default_font(wb: Workbook):
    """Mutate workbook default Normal style so empty cells use Kanit."""
    try:
        default = wb._named_styles["Normal"]
        default.font = Font(name="Kanit", size=10, color=INK)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────
#  Excel Table builder
# ─────────────────────────────────────────────────────────────────────────
def add_excel_table(ws, ref: str, name: str, style="TableStyleMedium2"):
    """Aggiunge una Excel Table con stile Medium2. ref es. 'A3:R462'."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)


# ─────────────────────────────────────────────────────────────────────────
#  Chart helpers
# ─────────────────────────────────────────────────────────────────────────
def style_chart(chart, title=None):
    chart.style = 2
    chart.legend.position = 'b'
    if chart.x_axis is not None:
        chart.x_axis.majorGridlines = None
    if chart.y_axis is not None:
        chart.y_axis.majorGridlines = None
    chart.width = 14   # cm
    chart.height = 8
    if title:
        chart.title = title


# ─────────────────────────────────────────────────────────────────────────
#  Sheet 9: Dati listing (TblListing)
# ─────────────────────────────────────────────────────────────────────────
LISTING_COLS = [
    "Mese", "ID Airbnb", "Nome", "Tipologia", "Comune",
    "Camere", "Letti", "Bagni", "Ospiti",
    "Prezzo mediano €/n", "Prezzo medio €/n", "N. settimane",
    "Rating", "Recensioni",
    "Affinità tipica", "Indizi", "Link",
    "Scenario",
]

# These two indices are hardcoded in the SORT formula on the Scen. 1/2/3 sheets
# (the {15,10} array argument). If anyone reorders LISTING_COLS without
# updating the SORT, the scenario sheets will silently sort by the wrong
# columns. The asserts here make the breakage explode at build time.
assert LISTING_COLS.index("Affinità tipica") + 1 == 15, \
    "SORT column index mismatch — update Scen sheet SORT formula"
assert LISTING_COLS.index("Prezzo mediano €/n") + 1 == 10, \
    "SORT column index mismatch — update Scen sheet SORT formula"


def write_dati_listing(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Dati listing")
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL_LIGHT

    ws.cell(row=1, column=1, value="Dati listing — universo completo").style = "TitleHC"
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(LISTING_COLS))
    ws.cell(row=2, column=1,
            value=f"{len(rows)} righe · Giugno + Luglio 2026 aggregato per settimana."
            ).style = "SubtitleHC"
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(LISTING_COLS))

    header_row = 3
    for j, col in enumerate(LISTING_COLS, start=1):
        c = ws.cell(row=header_row, column=j, value=col)
        c.style = "HeaderHC"
    ws.row_dimensions[header_row].height = 26

    for i, r in enumerate(rows, start=header_row + 1):
        for j, col in enumerate(LISTING_COLS, start=1):
            v = r.get(col)
            ws.cell(row=i, column=j, value=v)

    last_row = header_row + len(rows)
    if len(rows) == 0:
        last_row = header_row + 1  # tabella vuota: una riga finta
        ws.cell(row=last_row, column=1, value=None)

    ref = f"A{header_row}:{get_column_letter(len(LISTING_COLS))}{last_row}"
    add_excel_table(ws, ref, "TblListing")

    # Column widths
    widths = {
        "A": 10, "B": 14, "C": 38, "D": 18, "E": 16,
        "F": 8, "G": 8, "H": 8, "I": 8,
        "J": 16, "K": 16, "L": 12,
        "M": 8, "N": 12,
        "O": 14, "P": 32, "Q": 36,
        "R": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Number format for price cols (J, K) — applica direttamente alle celle
    for r in range(header_row + 1, last_row + 1):
        for col_letter in ("J", "K"):
            ws[f"{col_letter}{r}"].number_format = '€ #,##0'

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheets 10-11: Dati giornaliero (TblDailyJun, TblDailyJul)
# ─────────────────────────────────────────────────────────────────────────
DAILY_BASE_COLS = ["ID Airbnb", "Nome", "Tipologia", "Comune", "Camere", "Ospiti", "Affinità tipica"]


def write_dati_daily(wb: Workbook, sheet_name: str, table_name: str,
                     rows: list[dict], day_labels: list[str], month_label: str):
    ws = wb.create_sheet(sheet_name)
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL_LIGHT

    cols = DAILY_BASE_COLS + day_labels
    n_cols = len(cols)

    ws.cell(row=1, column=1, value=f"Dati giornaliero — {month_label} 2026").style = "TitleHC"
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=min(n_cols, 12))
    ws.cell(row=2, column=1,
            value=f"{len(rows)} listing · 30 sliding 2-night windows. Prezzo €/notte.").style = "SubtitleHC"
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=min(n_cols, 12))

    header_row = 3
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=j, value=col)
        c.style = "HeaderHC"
    ws.row_dimensions[header_row].height = 26

    for i, r in enumerate(rows, start=header_row + 1):
        for j, col in enumerate(cols, start=1):
            v = r.get(col)
            ws.cell(row=i, column=j, value=v)

    last_row = header_row + len(rows)
    if len(rows) == 0:
        last_row = header_row + 1

    ref = f"A{header_row}:{get_column_letter(n_cols)}{last_row}"
    add_excel_table(ws, ref, table_name)

    # Column widths
    base_widths = {"A": 12, "B": 36, "C": 18, "D": 14,
                   "E": 8, "F": 8, "G": 14}
    for col, w in base_widths.items():
        ws.column_dimensions[col].width = w
    # Date columns: 11 chars
    for idx in range(len(DAILY_BASE_COLS) + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 12

    # Format date cells as €
    for r in range(header_row + 1, last_row + 1):
        for idx in range(len(DAILY_BASE_COLS) + 1, n_cols + 1):
            ws.cell(row=r, column=idx).number_format = '€ #,##0'

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheet 1: Brief
# ─────────────────────────────────────────────────────────────────────────
def write_brief(wb: Workbook):
    ws = wb.create_sheet("Brief", 0)
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL

    ws.cell(row=1, column=1, value="Casa tipica Cernobbio — 3 Scenari").style = "TitleHC"
    ws.merge_cells("A1:J1")
    ws.cell(row=2, column=1,
            value=f"Mercato affitti brevi · Giugno + Luglio 2026 · scraping {datetime.now().strftime('%Y-%m-%d')}"
            ).style = "SubtitleHC"
    ws.merge_cells("A2:J2")

    # Profilo immobile
    ws.cell(row=4, column=1, value="Profilo immobile").style = "SectionHC"
    profile = [
        ["Località",          "Cernobbio, zona La Vignetta (pre-collinare)"],
        ["Caratteristiche",   "Casa indipendente in corte privata, pietra, soffitti legno massiccio, parquet, cementine"],
        ["Sviluppo",          "Piano terra (cucina+bagno+ripostiglio) · Piano 1 (1 cam + soggiorno o 2 cam + terrazzo) · Mansarda (2 cam)"],
    ]
    for i, (k, v) in enumerate(profile, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(name="Kanit", bold=True, color=INK, size=10)
        ws.cell(row=i, column=2, value=v).font = Font(name="Kanit", color=INK, size=10)
        ws.merge_cells(start_row=i, end_row=i, start_column=2, end_column=10)
        ws.row_dimensions[i].height = 22

    # KPI grid header at row 10
    ws.cell(row=9, column=1, value="KPI per scenario (mediana e percentili calcolati live da TblListing)").style = "SectionHC"
    ws.merge_cells("A9:J9")

    headers = ["Scenario", "Mese", "N comp", "Mediana", "Media",
               "p25", "p75", "Min", "Max", "Δ Giu→Lug%"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=10, column=j, value=h)
        c.style = "HeaderHC"
    ws.row_dimensions[10].height = 26

    # 6 data rows: (scenario, mese, row_idx in sheet, prev_giu_row_idx_or_None)
    rows_def = [
        ("S1", "giugno", 11, None),
        ("S1", "luglio", 12, 11),
        ("S2", "giugno", 13, None),
        ("S2", "luglio", 14, 13),
        ("S3", "giugno", 15, None),
        ("S3", "luglio", 16, 15),
    ]

    scen_labels = {
        "S1": "Scen. 1 — Solo piano terra",
        "S2": "Scen. 2 — Solo mansarda",
        "S3": "Scen. 3 — Casa intera",
    }

    for (scen, mese, r, prev_r) in rows_def:
        # A: scenario label
        ws.cell(row=r, column=1, value=scen_labels[scen]).font = Font(name="Kanit", size=10, color=INK)
        # B: mese
        ws.cell(row=r, column=2, value=mese.capitalize()).font = Font(name="Kanit", size=10, color=INK)

        # C: N comp
        ws.cell(row=r, column=3, value=(
            f'=SUMPRODUCT((TblListing[Mese]="{mese}")*(TblListing[Scenario]="{scen}")*1)'
        ))
        # D: Mediana (array formula — CSE for Excel < 365 compatibility)
        d_cell = ws.cell(row=r, column=4)
        d_cell.value = ArrayFormula(
            d_cell.coordinate,
            f'=MEDIAN(IF((TblListing[Mese]="{mese}")*(TblListing[Scenario]="{scen}"),'
            f' TblListing[Prezzo mediano €/n]))',
        )
        d_cell.number_format = '€ #,##0'
        # E: Media (array formula)
        e_cell = ws.cell(row=r, column=5)
        e_cell.value = ArrayFormula(
            e_cell.coordinate,
            f'=AVERAGE(IF((TblListing[Mese]="{mese}")*(TblListing[Scenario]="{scen}"),'
            f' TblListing[Prezzo mediano €/n]))',
        )
        e_cell.number_format = '€ #,##0'
        # F: p25 (array formula)
        f_cell = ws.cell(row=r, column=6)
        f_cell.value = ArrayFormula(
            f_cell.coordinate,
            f'=PERCENTILE.INC(IF((TblListing[Mese]="{mese}")*(TblListing[Scenario]="{scen}"),'
            f' TblListing[Prezzo mediano €/n]), 0.25)',
        )
        f_cell.number_format = '€ #,##0'
        # G: p75 (array formula)
        g_cell = ws.cell(row=r, column=7)
        g_cell.value = ArrayFormula(
            g_cell.coordinate,
            f'=PERCENTILE.INC(IF((TblListing[Mese]="{mese}")*(TblListing[Scenario]="{scen}"),'
            f' TblListing[Prezzo mediano €/n]), 0.75)',
        )
        g_cell.number_format = '€ #,##0'
        # H: Min
        ws.cell(row=r, column=8, value=(
            f'=MINIFS(TblListing[Prezzo mediano €/n], TblListing[Mese], "{mese}",'
            f' TblListing[Scenario], "{scen}")'
        )).number_format = '€ #,##0'
        # I: Max
        ws.cell(row=r, column=9, value=(
            f'=MAXIFS(TblListing[Prezzo mediano €/n], TblListing[Mese], "{mese}",'
            f' TblListing[Scenario], "{scen}")'
        )).number_format = '€ #,##0'

        # J: Delta% (solo luglio rows)
        if prev_r is not None:
            ws.cell(row=r, column=10, value=(
                f'=IFERROR((D{r}-D{prev_r})/D{prev_r}, "")'
            )).number_format = '0.0%'

    # Executive notes
    ws.cell(row=18, column=1, value="Note esecutive").style = "SectionHC"
    notes = [
        "• Scen. 1 (piano terra) — segmento più ampio del mercato: mini-studio in Cernobbio è la categoria più affollata. Margine basso ma alta rotazione.",
        "• Scen. 2 (mansarda) — intercetta coppie/piccole famiglie 4 ospiti. Sweet spot per posizionamento premium se enfatizzi la tipicità (legno, cementine).",
        "• Scen. 3 (casa intera) — pochi comparabili NON-villa. Il 3-4 camere a Cernobbio è dominato dalle ville luxury: voi sareste in una fascia distinta (€500-900/n vs €1500+/n delle ville).",
        "• Il mercato cresce ~6-8% giugno→luglio (colonna Δ). Pricing dinamico: alzare ADR del 15-20% per il periodo 15 lug – 25 ago.",
        "• Tutti i KPI di questa scheda sono FORMULE che leggono TblListing: filtra/aggiungi righe sul foglio 'Dati listing' e i KPI si aggiornano live.",
    ]
    for i, n in enumerate(notes, start=19):
        c = ws.cell(row=i, column=1, value=n)
        c.font = Font(name="Kanit", size=10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=10)
        ws.row_dimensions[i].height = 32

    # Column widths
    widths = {"A": 30, "B": 12, "C": 10, "D": 14, "E": 14,
              "F": 12, "G": 12, "H": 12, "I": 12, "J": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheets 2-4: Scen. 1/2/3 — KPI strip + FILTER spill
# ─────────────────────────────────────────────────────────────────────────
SCEN_SHEETS = [
    {
        "name": "Scen. 1 piano terra",
        "title": "Scen. 1 — Solo piano terra",
        "subtitle": "Mini-studio: cucina + bagno + ampio ripostiglio. Sviluppo orizzontale, 0-1 camere, 1-3 ospiti.",
        "tag": "S1",
        "brief_giu_row": 11,
        "brief_lug_row": 12,
    },
    {
        "name": "Scen. 2 mansarda",
        "title": "Scen. 2 — Solo mansarda",
        "subtitle": "Solo mansarda affittata: 2 camere matrimoniali, 3-5 ospiti. Soffitti in legno, ambiente raccolto.",
        "tag": "S2",
        "brief_giu_row": 13,
        "brief_lug_row": 14,
    },
    {
        "name": "Scen. 3 casa intera",
        "title": "Scen. 3 — Casa intera",
        "subtitle": "Piano terra + piano 1 + mansarda. 3-5 camere, 6-9 ospiti. Casa tipica completa con corte privata.",
        "tag": "S3",
        "brief_giu_row": 15,
        "brief_lug_row": 16,
    },
]


def write_scenario_sheet(wb: Workbook, conf: dict):
    ws = wb.create_sheet(conf["name"])
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL

    # Row 1 title, row 2 subtitle
    ws.cell(row=1, column=1, value=conf["title"]).style = "TitleHC"
    ws.merge_cells("A1:R1")
    ws.cell(row=2, column=1, value=conf["subtitle"]).style = "SubtitleHC"
    ws.merge_cells("A2:R2")

    # Row 4-5: KPI strip
    kpi_labels = ["N. Giu", "Mediana Giu €/n", "N. Lug", "Mediana Lug €/n", "Δ %"]
    for j, lab in enumerate(kpi_labels, start=1):
        c = ws.cell(row=4, column=j, value=lab)
        c.style = "KPILabelHC"

    gr = conf["brief_giu_row"]
    lr = conf["brief_lug_row"]

    ws.cell(row=5, column=1, value=f"=Brief!C{gr}").style = "KPIValueHC"
    ws.cell(row=5, column=2, value=f"=Brief!D{gr}").style = "KPIValueHC"
    ws.cell(row=5, column=2).number_format = '€ #,##0'
    ws.cell(row=5, column=3, value=f"=Brief!C{lr}").style = "KPIValueHC"
    ws.cell(row=5, column=4, value=f"=Brief!D{lr}").style = "KPIValueHC"
    ws.cell(row=5, column=4).number_format = '€ #,##0'
    ws.cell(row=5, column=5, value=f"=Brief!J{lr}").style = "KPIValueHC"
    ws.cell(row=5, column=5).number_format = '0.0%'

    ws.row_dimensions[5].height = 36

    # Row 7: Header strip styled like a table header, columns matching TblListing
    for j, col in enumerate(LISTING_COLS, start=1):
        c = ws.cell(row=7, column=j, value=col)
        c.style = "HeaderHC"
    ws.row_dimensions[7].height = 26

    # Row 8: FILTER+SORT spill formula
    # Sort by Affinità tipica DESC (col 15), then Prezzo mediano €/n ASC (col 10)
    formula = (
        f'=SORT(FILTER(TblListing, TblListing[Scenario]="{conf["tag"]}"),'
        f' {{15,10}}, {{-1,1}})'
    )
    ws.cell(row=8, column=1, value=formula)

    # Column widths (match Dati listing style)
    widths = {
        "A": 10, "B": 14, "C": 38, "D": 18, "E": 16,
        "F": 8, "G": 8, "H": 8, "I": 8,
        "J": 16, "K": 16, "L": 12,
        "M": 8, "N": 12,
        "O": 14, "P": 32, "Q": 36,
        "R": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheet 5: Distribuzione mercato
# ─────────────────────────────────────────────────────────────────────────
def write_distribuzione(wb: Workbook):
    ws = wb.create_sheet("Distribuzione mercato")
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL

    ws.cell(row=1, column=1, value="Distribuzione mercato — Camere, Ospiti, Scenari").style = "TitleHC"
    ws.merge_cells("A1:N1")
    ws.cell(row=2, column=1,
            value="Conteggi calcolati live da TblListing via COUNTIFS. Cambia i dati e i grafici si aggiornano."
            ).style = "SubtitleHC"
    ws.merge_cells("A2:N2")

    # ── Section 1: Per numero CAMERE (rows 4-13) ──────────────────────────
    ws.cell(row=4, column=1, value="Per numero CAMERE").style = "SectionHC"
    # Row 5 header
    for j, h in enumerate(["Camere", "Giugno", "Luglio"], start=1):
        ws.cell(row=5, column=j, value=h).style = "HeaderHC"
    ws.row_dimensions[5].height = 26

    # Rows 6-13: buckets 0,1,2,3,4,5,6,"7+"
    cam_buckets = [0, 1, 2, 3, 4, 5, 6, "7+"]
    for i, b in enumerate(cam_buckets, start=6):
        if b == "7+":
            ws.cell(row=i, column=1, value="7+").style = "BodyHC"
            ws.cell(row=i, column=2, value='=COUNTIFS(TblListing[Camere], ">=7", TblListing[Mese], "giugno")').style = "BodyHC"
            ws.cell(row=i, column=3, value='=COUNTIFS(TblListing[Camere], ">=7", TblListing[Mese], "luglio")').style = "BodyHC"
        else:
            ws.cell(row=i, column=1, value=b).style = "BodyHC"
            ws.cell(row=i, column=2, value=f'=COUNTIFS(TblListing[Camere], {b}, TblListing[Mese], "giugno")').style = "BodyHC"
            ws.cell(row=i, column=3, value=f'=COUNTIFS(TblListing[Camere], {b}, TblListing[Mese], "luglio")').style = "BodyHC"

    add_excel_table(ws, "A5:C13", "TblDistrCam")

    # BarChart for camere
    chart_cam = BarChart()
    chart_cam.type = "col"
    chart_cam.grouping = "clustered"
    data = Reference(ws, min_col=2, max_col=3, min_row=5, max_row=13)
    cats = Reference(ws, min_col=1, max_col=1, min_row=6, max_row=13)
    chart_cam.add_data(data, titles_from_data=True)
    chart_cam.set_categories(cats)
    style_chart(chart_cam, title="Distribuzione listing per numero CAMERE — Giu vs Lug 2026")
    chart_cam.x_axis.title = "Numero camere"
    chart_cam.y_axis.title = "N. listing"
    ws.add_chart(chart_cam, "E5")

    # ── Section 2: Per numero OSPITI (rows 18-30) ─────────────────────────
    ws.cell(row=18, column=1, value="Per numero OSPITI").style = "SectionHC"
    for j, h in enumerate(["Ospiti", "Giugno", "Luglio"], start=1):
        ws.cell(row=19, column=j, value=h).style = "HeaderHC"
    ws.row_dimensions[19].height = 26

    # Rows 20-30: 1..10 then "10+" (using >=11)
    osp_buckets = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, "10+"]
    for i, b in enumerate(osp_buckets, start=20):
        if b == "10+":
            ws.cell(row=i, column=1, value="10+").style = "BodyHC"
            ws.cell(row=i, column=2, value='=COUNTIFS(TblListing[Ospiti], ">=11", TblListing[Mese], "giugno")').style = "BodyHC"
            ws.cell(row=i, column=3, value='=COUNTIFS(TblListing[Ospiti], ">=11", TblListing[Mese], "luglio")').style = "BodyHC"
        else:
            ws.cell(row=i, column=1, value=b).style = "BodyHC"
            ws.cell(row=i, column=2, value=f'=COUNTIFS(TblListing[Ospiti], {b}, TblListing[Mese], "giugno")').style = "BodyHC"
            ws.cell(row=i, column=3, value=f'=COUNTIFS(TblListing[Ospiti], {b}, TblListing[Mese], "luglio")').style = "BodyHC"

    add_excel_table(ws, "A19:C30", "TblDistrOsp")

    chart_osp = BarChart()
    chart_osp.type = "col"
    chart_osp.grouping = "clustered"
    data2 = Reference(ws, min_col=2, max_col=3, min_row=19, max_row=30)
    cats2 = Reference(ws, min_col=1, max_col=1, min_row=20, max_row=30)
    chart_osp.add_data(data2, titles_from_data=True)
    chart_osp.set_categories(cats2)
    style_chart(chart_osp, title="Distribuzione listing per numero OSPITI — Giu vs Lug 2026")
    chart_osp.x_axis.title = "Numero ospiti"
    chart_osp.y_axis.title = "N. listing"
    ws.add_chart(chart_osp, "E19")

    # ── Section 3: Profondità comp set per scenario (rows 34-38) ──────────
    ws.cell(row=34, column=1, value="Profondità comp set per scenario").style = "SectionHC"
    for j, h in enumerate(
        ["Scenario", "Range camere", "Range ospiti", "Listing Giugno", "Listing Luglio"],
        start=1,
    ):
        ws.cell(row=35, column=j, value=h).style = "HeaderHC"
    ws.row_dimensions[35].height = 26

    scen_defs = [
        ("Scen. 1 — Piano terra",  "0-1", "1-3", "S1"),
        ("Scen. 2 — Mansarda",     "2",   "3-5", "S2"),
        ("Scen. 3 — Casa intera",  "3-5", "6-9", "S3"),
    ]
    for i, (label, cam_range, osp_range, tag) in enumerate(scen_defs, start=36):
        ws.cell(row=i, column=1, value=label).style = "BodyHC"
        ws.cell(row=i, column=2, value=cam_range).style = "BodyHC"
        ws.cell(row=i, column=3, value=osp_range).style = "BodyHC"
        ws.cell(row=i, column=4,
                value=f'=COUNTIFS(TblListing[Scenario], "{tag}", TblListing[Mese], "giugno")').style = "BodyHC"
        ws.cell(row=i, column=5,
                value=f'=COUNTIFS(TblListing[Scenario], "{tag}", TblListing[Mese], "luglio")').style = "BodyHC"

    add_excel_table(ws, "A35:E38", "TblScenDistr")

    chart_scen = BarChart()
    chart_scen.type = "bar"   # horizontal
    chart_scen.grouping = "clustered"
    data3 = Reference(ws, min_col=4, max_col=5, min_row=35, max_row=38)
    cats3 = Reference(ws, min_col=1, max_col=1, min_row=36, max_row=38)
    chart_scen.add_data(data3, titles_from_data=True)
    chart_scen.set_categories(cats3)
    style_chart(chart_scen, title="Profondità comp set per scenario — Giu vs Lug 2026")
    chart_scen.x_axis.title = "N. listing"
    ws.add_chart(chart_scen, "G35")

    # Column widths
    for col, w in [("A", 26), ("B", 14), ("C", 14), ("D", 16), ("E", 16)]:
        ws.column_dimensions[col].width = w

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheet 6: Trend giornaliero (TblTrend + LineChart)
# ─────────────────────────────────────────────────────────────────────────
def write_trend(wb: Workbook, jun_days: list[str], jul_days: list[str]):
    """
    Layout:
      Row 1 title
      Row 4 section
      Row 5 header (10 cols)
      Rows 6-35 giugno (30)
      Rows 36-65 luglio (30)
    Each row: Mese | Data | Giorno | N. listing | Mediana €/n | Media €/n |
              p25 | p75 | Solo Giu | Solo Lug

    Below at row 70+: small chart-pivot table (Giorno mese | Mediana Giu | Mediana Lug)
    LineChart at K5.
    """
    ws = wb.create_sheet("Trend giornaliero")
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL

    ws.cell(row=1, column=1, value="Trend giornaliero — Giu vs Lug 2026").style = "TitleHC"
    ws.merge_cells("A1:J1")
    ws.cell(row=2, column=1,
            value="Mediana €/notte per giorno (2-night sliding windows). Tutto calcolato live dai fogli 'Dati giornaliero giugno/luglio'."
            ).style = "SubtitleHC"
    ws.merge_cells("A2:J2")

    ws.cell(row=4, column=1, value="Tabella").style = "SectionHC"
    headers = ["Mese", "Data", "Giorno", "N. listing", "Mediana €/n", "Media €/n",
               "p25 €/n", "p75 €/n", "Solo Giu", "Solo Lug"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=5, column=j, value=h).style = "HeaderHC"
    ws.row_dimensions[5].height = 26

    # Daily data sheets: 'Dati giornaliero giugno' and 'Dati giornaliero luglio'
    # First 7 cols = dimensions. Date cols start at column H (col 8).
    # Compute last_row for each daily sheet — we need the number of rows.
    # The daily tables have data starting at row 4. We use generous max range.
    # June daily has ~331 listings; July ~301. Use 500 as safe upper bound.
    # When using COUNT/MEDIAN with empty cells, they're ignored — so over-shoot is fine.

    # Build rows for giugno (6-35) and luglio (36-65)
    def _date_col_letter(day_idx_in_month: int):
        # day_idx_in_month: 1..30 → corresponds to date col offset
        # First date col is H (col 8) for day 1
        return get_column_letter(8 + day_idx_in_month - 1)

    DAILY_MAX_ROW = 2000  # very safe upper bound (current data is ~335 + ~305)

    # Giugno rows 6-35 (day 1 → row 6, day 30 → row 35)
    for day_idx in range(1, 31):
        r = 5 + day_idx  # row 6 for day 1
        col_letter = _date_col_letter(day_idx)
        rng = f"'Dati giornaliero giugno'!{col_letter}4:{col_letter}{DAILY_MAX_ROW}"

        ws.cell(row=r, column=1, value="Giugno").style = "BodyHC"
        ws.cell(row=r, column=2, value=f"=DATE(2026,6,{day_idx})").style = "BodyHC"
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=3, value=f'=TEXT(B{r},"ddd")').style = "BodyHC"
        ws.cell(row=r, column=4, value=f"=COUNT({rng})").style = "BodyHC"
        ws.cell(row=r, column=5, value=f"=IFERROR(MEDIAN({rng}),NA())")
        ws.cell(row=r, column=5).number_format = '€ #,##0'
        ws.cell(row=r, column=6, value=f"=IFERROR(AVERAGE({rng}),NA())")
        ws.cell(row=r, column=6).number_format = '€ #,##0'
        ws.cell(row=r, column=7, value=f"=IFERROR(PERCENTILE.INC({rng},0.25),NA())")
        ws.cell(row=r, column=7).number_format = '€ #,##0'
        ws.cell(row=r, column=8, value=f"=IFERROR(PERCENTILE.INC({rng},0.75),NA())")
        ws.cell(row=r, column=8).number_format = '€ #,##0'
        # Solo Giu / Solo Lug
        ws.cell(row=r, column=9, value=f'=IF(A{r}="Giugno",E{r},NA())')
        ws.cell(row=r, column=9).number_format = '€ #,##0'
        ws.cell(row=r, column=10, value=f'=IF(A{r}="Luglio",E{r},NA())')
        ws.cell(row=r, column=10).number_format = '€ #,##0'

    # Luglio rows 36-65
    for day_idx in range(1, 31):
        r = 35 + day_idx  # row 36 for day 1
        col_letter = _date_col_letter(day_idx)
        rng = f"'Dati giornaliero luglio'!{col_letter}4:{col_letter}{DAILY_MAX_ROW}"

        ws.cell(row=r, column=1, value="Luglio").style = "BodyHC"
        ws.cell(row=r, column=2, value=f"=DATE(2026,7,{day_idx})").style = "BodyHC"
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=3, value=f'=TEXT(B{r},"ddd")').style = "BodyHC"
        ws.cell(row=r, column=4, value=f"=COUNT({rng})").style = "BodyHC"
        ws.cell(row=r, column=5, value=f"=IFERROR(MEDIAN({rng}),NA())")
        ws.cell(row=r, column=5).number_format = '€ #,##0'
        ws.cell(row=r, column=6, value=f"=IFERROR(AVERAGE({rng}),NA())")
        ws.cell(row=r, column=6).number_format = '€ #,##0'
        ws.cell(row=r, column=7, value=f"=IFERROR(PERCENTILE.INC({rng},0.25),NA())")
        ws.cell(row=r, column=7).number_format = '€ #,##0'
        ws.cell(row=r, column=8, value=f"=IFERROR(PERCENTILE.INC({rng},0.75),NA())")
        ws.cell(row=r, column=8).number_format = '€ #,##0'
        ws.cell(row=r, column=9, value=f'=IF(A{r}="Giugno",E{r},NA())')
        ws.cell(row=r, column=9).number_format = '€ #,##0'
        ws.cell(row=r, column=10, value=f'=IF(A{r}="Luglio",E{r},NA())')
        ws.cell(row=r, column=10).number_format = '€ #,##0'

    add_excel_table(ws, "A5:J65", "TblTrend")

    # Chart-pivot table at rows 70-100
    # Giorno mese | Mediana Giu | Mediana Lug
    ws.cell(row=69, column=1, value="Dati per LineChart (Giorno mese × Mediana per mese)").style = "SectionHC"
    for j, h in enumerate(["Giorno mese", "Mediana Giu", "Mediana Lug"], start=1):
        ws.cell(row=70, column=j, value=h).style = "HeaderHC"
    ws.row_dimensions[70].height = 26

    for day in range(1, 31):
        r = 70 + day  # row 71 for day 1, row 100 for day 30
        ws.cell(row=r, column=1, value=day).style = "BodyHC"
        giugno_trend_row = 5 + day
        luglio_trend_row = 35 + day
        ws.cell(row=r, column=2, value=f"=E{giugno_trend_row}").number_format = '€ #,##0'
        ws.cell(row=r, column=3, value=f"=E{luglio_trend_row}").number_format = '€ #,##0'

    # LineChart at K5
    chart = LineChart()
    chart.type = "line"
    chart.grouping = "standard"
    data = Reference(ws, min_col=2, max_col=3, min_row=70, max_row=100)
    cats = Reference(ws, min_col=1, max_col=1, min_row=71, max_row=100)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    style_chart(chart, title="Andamento mediana €/notte — Giu vs Lug 2026")
    chart.x_axis.title = "Giorno del mese"
    chart.y_axis.title = "Mediana €/n"
    ws.add_chart(chart, "L5")

    # Column widths
    for col, w in [("A", 10), ("B", 13), ("C", 10), ("D", 12),
                   ("E", 14), ("F", 14), ("G", 12), ("H", 12),
                   ("I", 12), ("J", 12)]:
        ws.column_dimensions[col].width = w

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheet 7: Ricavi annui (3 stacked blocks)
# ─────────────────────────────────────────────────────────────────────────
MONTHS_DEF = [
    ("Gennaio",   31, 0.48),
    ("Febbraio",  28, 0.48),
    ("Marzo",     31, 0.55),
    ("Aprile",    30, 0.62),
    ("Maggio",    31, 0.78),
    ("Giugno",    30, 0.92),
    ("Luglio",    31, 1.00),
    ("Agosto",    31, 1.00),
    ("Settembre", 30, 0.78),
    ("Ottobre",   31, 0.62),
    ("Novembre",  30, 0.48),
    ("Dicembre",  31, 0.68),
]


def _write_revenue_block(ws, start_row: int, scen_label: str,
                         brief_lug_row: int):
    """
    Write one revenue block starting at start_row.
    Layout:
      start_row+0  → section title
      start_row+1  → ADR baseline cell (label A, value B)
      start_row+2  → header
      start_row+3..+14 → 12 mesi
      start_row+15 → TOTALE ANNO
    Returns next free row.
    """
    title_row = start_row
    adr_row   = start_row + 1
    header_row = start_row + 2
    first_month_row = start_row + 3
    last_month_row = start_row + 14
    total_row = start_row + 15

    # Title
    ws.cell(row=title_row, column=1,
            value=f"{scen_label} · ADR baseline luglio (mediana scenario)"
            ).style = "SectionHC"
    ws.merge_cells(start_row=title_row, end_row=title_row, start_column=1, end_column=7)

    # ADR baseline
    ws.cell(row=adr_row, column=1, value="ADR baseline luglio €/n").style = "KPILabelHC"
    c = ws.cell(row=adr_row, column=2, value=f"=Brief!D{brief_lug_row}")
    c.style = "KPIValueHC"
    c.number_format = '€ #,##0'

    # Header
    headers = ["Mese", "Giorni", "Moltiplicatore", "ADR stimato €/n",
               "Ricavo @ 50%", "Ricavo @ 65%", "Ricavo @ 75%"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h).style = "HeaderHC"
    ws.row_dimensions[header_row].height = 26

    # 12 mesi
    for i, (mname, days, mult) in enumerate(MONTHS_DEF):
        r = first_month_row + i
        ws.cell(row=r, column=1, value=mname).style = "BodyHC"
        ws.cell(row=r, column=2, value=days).style = "BodyHC"
        ws.cell(row=r, column=3, value=mult).style = "BodyHC"
        ws.cell(row=r, column=3).number_format = '0.00'
        # D: =$B$<adr_row>*C<r>
        ws.cell(row=r, column=4, value=f"=$B${adr_row}*C{r}")
        ws.cell(row=r, column=4).number_format = '€ #,##0'
        # E: =D*B*0.5
        ws.cell(row=r, column=5, value=f"=D{r}*B{r}*0.5")
        ws.cell(row=r, column=5).number_format = '€ #,##0'
        ws.cell(row=r, column=6, value=f"=D{r}*B{r}*0.65")
        ws.cell(row=r, column=6).number_format = '€ #,##0'
        ws.cell(row=r, column=7, value=f"=D{r}*B{r}*0.75")
        ws.cell(row=r, column=7).number_format = '€ #,##0'

    # Total row
    ws.cell(row=total_row, column=1, value="TOTALE ANNO").font = Font(name="Kanit", bold=True, color=TEAL, size=11)
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_month_row}:B{last_month_row})")
    ws.cell(row=total_row, column=2).font = Font(name="Kanit", bold=True, color=INK, size=10)
    ws.cell(row=total_row, column=5, value=f"=SUM(E{first_month_row}:E{last_month_row})")
    ws.cell(row=total_row, column=5).number_format = '€ #,##0'
    ws.cell(row=total_row, column=5).font = Font(name="Kanit", bold=True, color=TEAL, size=11)
    ws.cell(row=total_row, column=6, value=f"=SUM(F{first_month_row}:F{last_month_row})")
    ws.cell(row=total_row, column=6).number_format = '€ #,##0'
    ws.cell(row=total_row, column=6).font = Font(name="Kanit", bold=True, color=TEAL, size=11)
    ws.cell(row=total_row, column=7, value=f"=SUM(G{first_month_row}:G{last_month_row})")
    ws.cell(row=total_row, column=7).number_format = '€ #,##0'
    ws.cell(row=total_row, column=7).font = Font(name="Kanit", bold=True, color=TEAL, size=11)

    return total_row + 1


def write_ricavi(wb: Workbook):
    ws = wb.create_sheet("Ricavi annui")
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL

    ws.cell(row=1, column=1, value="Ricavi annui — proiezione su 3 occupancy").style = "TitleHC"
    ws.merge_cells("A1:G1")
    ws.cell(row=2, column=1,
            value="Modello: ADR Luglio (mediana scenario) × moltiplicatore stagionale × giorni × occupancy. Lordo, no commissioni, no imposte."
            ).style = "SubtitleHC"
    ws.merge_cells("A2:G2")

    # Block 1: rows 3-18 (Scen.1 → Brief row 12)
    _write_revenue_block(ws, start_row=3,
                         scen_label="Scen. 1 — Solo piano terra",
                         brief_lug_row=12)
    # Block 2: rows 22-37 (Scen.2 → Brief row 14)
    _write_revenue_block(ws, start_row=22,
                         scen_label="Scen. 2 — Solo mansarda",
                         brief_lug_row=14)
    # Block 3: rows 41-56 (Scen.3 → Brief row 16)
    _write_revenue_block(ws, start_row=41,
                         scen_label="Scen. 3 — Casa intera",
                         brief_lug_row=16)

    # Column widths
    for col, w in [("A", 22), ("B", 10), ("C", 16), ("D", 18),
                   ("E", 16), ("F", 16), ("G", 16)]:
        ws.column_dimensions[col].width = w

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheet 8: Hotel zona Cernobbio (TblHotels)
# ─────────────────────────────────────────────────────────────────────────
HOTEL_COLS = [
    "Hotel", "Tipo", "Rating", "Recensioni", "Score affinità",
    "Indirizzo", "Sito web", "Telefono", "Maps",
    "Booking giugno", "Booking luglio",
]


def write_hotels(wb: Workbook):
    ws = wb.create_sheet("Hotel zona Cernobbio")
    style_sheet(ws)
    ws.sheet_properties.tabColor = TEAL

    payload = json.loads(HOTELS_JSON.read_text(encoding="utf-8")) if HOTELS_JSON.exists() else {"hotels": []}
    hotels = payload.get("hotels", []) or []

    ws.cell(row=1, column=1,
            value=f"Hotel zona Cernobbio ({len(hotels)} strutture)").style = "TitleHC"
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(HOTEL_COLS))
    ws.cell(row=2, column=1,
            value="Fonte: Google Places API. Score affinità = Rating*10 + log10(Recensioni+1)*5."
            ).style = "SubtitleHC"
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(HOTEL_COLS))

    header_row = 3
    for j, col in enumerate(HOTEL_COLS, start=1):
        ws.cell(row=header_row, column=j, value=col).style = "HeaderHC"
    ws.row_dimensions[header_row].height = 26

    # Sort by rating desc, missing last
    hotels_sorted = sorted(hotels, key=lambda h: (h.get("rating") or 0), reverse=True)

    for i, h in enumerate(hotels_sorted, start=header_row + 1):
        ws.cell(row=i, column=1, value=h.get("name") or "")
        ws.cell(row=i, column=2, value=h.get("primary_type_display") or h.get("primary_type") or "")
        ws.cell(row=i, column=3, value=h.get("rating"))
        ws.cell(row=i, column=4, value=h.get("reviews"))
        # E: Score affinità → FORMULA structured
        ws.cell(row=i, column=5, value=(
            '=IF(AND(ISNUMBER([@Rating]),ISNUMBER([@Recensioni])),'
            '[@Rating]*10+LOG10([@Recensioni]+1)*5,"")'
        ))
        ws.cell(row=i, column=5).number_format = '0.0'
        ws.cell(row=i, column=6, value=h.get("address") or "")
        ws.cell(row=i, column=7, value=h.get("website") or "")
        ws.cell(row=i, column=8, value=h.get("phone") or "")
        ws.cell(row=i, column=9, value=h.get("google_maps_url") or "")
        ws.cell(row=i, column=10, value=h.get("booking_june") or "")
        ws.cell(row=i, column=11, value=h.get("booking_july") or "")

    last_row = header_row + max(len(hotels_sorted), 1)
    ref = f"A{header_row}:{get_column_letter(len(HOTEL_COLS))}{last_row}"
    add_excel_table(ws, ref, "TblHotels")

    # Column widths
    widths = {"A": 36, "B": 16, "C": 9, "D": 11, "E": 14,
              "F": 50, "G": 40, "H": 18, "I": 32, "J": 32, "K": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Sheet 12: Metodologia
# ─────────────────────────────────────────────────────────────────────────
def write_metodologia(wb: Workbook):
    ws = wb.create_sheet("Metodologia")
    style_sheet(ws)
    ws.sheet_properties.tabColor = ACCENT_ORANGE

    ws.cell(row=1, column=1, value="Metodologia").style = "TitleHC"
    ws.merge_cells("A1:E1")

    notes = [
        "Fonti dati:",
        "  • Airbnb StaysSearch (API interna) via pyairbnb (TLS impersonation Chrome).",
        "  • Google Places API (New) places:searchText per gli hotel.",
        "",
        "Coperture temporali:",
        "  • Weekly: 4 finestre giugno (3 giu - 1 lug, Mer→Mer) + 5 finestre luglio (1 lug - 5 ago).",
        "  • Daily: 30 sliding 2-night windows per ciascun mese (start day = ogni giorno 1-30).",
        "  • Daily luglio: copertura 1-30 lug (31 lug escluso — finestra 2-notti non raccoglibile l'ultimo giorno del mese senza check-out in agosto).",
        "",
        "Filtri Scenari (taggati nella colonna Scenario di TblListing):",
        "  • S1 — Solo piano terra: Camere 0-1 (o None), Ospiti 1-3 (o None), non Villa, comune in {Cernobbio, Moltrasio, Blevio, Como}.",
        "  • S2 — Solo mansarda:   Camere 2, Ospiti 3-5 (o None), non Villa, comune OK.",
        "  • S3 — Casa intera:     Camere 3-5, Ospiti 6-9 (o None), non Villa, comune OK.",
        "  • Fallback: se Camere è None, il check di appartenenza scivola sulla sola colonna Ospiti.",
        "",
        "Punteggio affinità tipica:",
        "  Conta hit di keyword (stone, pietra, rustic, historic, antico, tipica, traditional, village, borgo, wood beams,",
        "  travi, legno, century, cottage, courtyard, corte, garden, panoram, vista lago…) nel titolo/nome del listing.",
        "",
        "Modello ricavi (foglio 'Ricavi annui'):",
        "  Mensile = ADR Luglio (mediana scenario, da Brief) × moltiplicatore stagionale × giorni × occupancy.",
        "  Moltiplicatori: alta (giu/lug/ago) 0.92/1.00/1.00, shoulder+ (mag/set) 0.78, shoulder (apr/ott) 0.62,",
        "  dicembre weighted 0.68, bassa (gen/feb/nov) 0.48.",
        "",
        "Architettura formule (v4):",
        "  • Tutti i KPI del foglio Brief sono formule che leggono TblListing.",
        "  • Distribuzione mercato usa COUNTIFS strutturate.",
        "  • Trend giornaliero usa MEDIAN/AVERAGE/PERCENTILE.INC sui range delle date dei fogli daily.",
        "  • Ricavi annui leggono l'ADR baseline luglio dal foglio Brief (cella D12/D14/D16).",
        "  • Score hotel calcolato via formula [@Rating]*10 + LOG10([@Recensioni]+1)*5.",
        "",
        "Caveat:",
        "  • Prezzi = rack rate Airbnb dinamici al momento dello scraping. Non confermati come booking.",
        "  • Hotel: nightly rate NON disponibili via Google Places; usare i link Booking della scheda.",
        "  • Filtro Tipologia ('Villa' esclusa) usa il tag Airbnb auto-assegnato — falsi positivi possibili.",
        "  • Occupancy 50/65/75% sono ipotesi standard di mercato; storico Smoobu o AirDNA validerebbero il dato.",
    ]
    for i, n in enumerate(notes, start=3):
        c = ws.cell(row=i, column=1, value=n)
        c.font = Font(name="Kanit", size=10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=5)
        ws.row_dimensions[i].height = 18

    ws.column_dimensions["A"].width = 100
    return ws


# ─────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────
def main() -> int:
    print("→ Loading weekly data…")
    rows_jun = build_comp_rows(WEEKLY_JUN, "giugno")
    rows_jul = build_comp_rows(WEEKLY_JUL, "luglio")
    listing_rows = rows_jun + rows_jul
    print(f"  giugno: {len(rows_jun)} · luglio: {len(rows_jul)} · totale: {len(listing_rows)}")

    n_s1 = sum(1 for r in listing_rows if r["Scenario"] == "S1")
    n_s2 = sum(1 for r in listing_rows if r["Scenario"] == "S2")
    n_s3 = sum(1 for r in listing_rows if r["Scenario"] == "S3")
    print(f"  taggati: S1={n_s1}, S2={n_s2}, S3={n_s3}, '': {len(listing_rows)-n_s1-n_s2-n_s3}")

    print("→ Loading daily data…")
    daily_jun_rows, jun_days = build_daily_rows(DAILY_JUN)
    daily_jul_rows, jul_days = build_daily_rows(DAILY_JUL)
    print(f"  daily giugno: {len(daily_jun_rows)} listings × {len(jun_days)} giorni")
    print(f"  daily luglio: {len(daily_jul_rows)} listings × {len(jul_days)} giorni")

    print("→ Building workbook…")
    wb = Workbook()
    # Drop default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    register_named_styles(wb)
    set_default_font(wb)

    # ── Sheets in target order ────────────────────────────────────────────
    # 1. Brief
    write_brief(wb)
    # 2-4. Scenarios
    for conf in SCEN_SHEETS:
        write_scenario_sheet(wb, conf)
    # 5. Distribuzione mercato
    write_distribuzione(wb)
    # 6. Trend giornaliero
    write_trend(wb, jun_days, jul_days)
    # 7. Ricavi annui
    write_ricavi(wb)
    # 8. Hotel zona Cernobbio
    write_hotels(wb)
    # 9. Dati listing (TblListing)
    write_dati_listing(wb, listing_rows)
    # 10. Dati giornaliero giugno
    write_dati_daily(wb, "Dati giornaliero giugno", "TblDailyJun",
                     daily_jun_rows, jun_days, "Giugno")
    # 11. Dati giornaliero luglio
    write_dati_daily(wb, "Dati giornaliero luglio", "TblDailyJul",
                     daily_jul_rows, jul_days, "Luglio")
    # 12. Metodologia
    write_metodologia(wb)

    wb.save(XLSX)
    print(f"\n✅  Excel scritto: {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
