"""
Generator v3 — Excel a formule per proiezioni Andrea Furlan.

Architettura:
- Sheet "Parametri": single source of truth (rates, factors, targets, ramp)
- Sheet "Costi setup": valori utente preservati + formule SUM per totali
- Sheet "Mensile Y1/Y2": inputs (mese, gg, stagione, ADR, ramp) + formule per il resto
- Sheet "Sintesi": cross-sheet references ai totali mensili
- Sheet "Breakeven Y1": cumulato via formula
- Sheet "Alternativa MTR": formule riferite a Parametri + Sintesi
- Sheet "Assunzioni & note"

Modifiche v3 vs v2:
- Tutti i totali = formule (SUM/IF/ROUND), non valori precomputati
- Ramp-up Aug-Sep 2026: nuovo annuncio = occupancy ridotta (vedi Parametri)
- Valori "Costi setup" letti dal file esistente (preserva edit utente come Tipico SUAP €170)
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import shutil, os, time

OUT_DESKTOP = r"C:\Users\Andrei\Desktop\Proiezioni    Andrea Furlan.xlsx"
OUT_PRATICA = r"C:\Users\Andrei\Desktop\Pratica Andrea Furlan\proiezioni-andrea-furlan.xlsx"
EXISTING_FILE = OUT_DESKTOP  # source per leggere Costi setup utente

FONT_NAME = "Kanit"
TEAL = "0C7489"
TEAL_BG = "E8F4F6"
DARK = "1F2937"
RED = "DC2626"
GREEN = "059669"
ORANGE = "F59E0B"
BLUE_LINK = "0563C1"
GRAY_BG = "F1F5F9"
YELLOW_BG = "FEF3C7"

# =============================================================================
# Helpers
# =============================================================================
def f(name=FONT_NAME, size=11, bold=False, color=DARK, italic=False, underline=None):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic, underline=underline)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border():
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)

def style_header(cell, color=TEAL):
    cell.font = f(size=11, bold=True, color="FFFFFF")
    cell.fill = fill(color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def style_data(cell, bold=False, total_row=False, eur=True, color=DARK, fmt=None):
    cell.font = f(size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    if total_row:
        cell.fill = fill(TEAL_BG)
        cell.font = f(size=10, bold=True, color=TEAL)
    if fmt:
        cell.number_format = fmt
    elif eur:
        cell.number_format = '#,##0 "€"'

def style_input(cell, bold=False):
    """Input cell visivamente distinguibile (sfondo giallo chiaro)."""
    cell.font = f(size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    cell.fill = fill(YELLOW_BG)

def style_label(cell, bold=False, total=False):
    cell.font = f(size=10, bold=bold or total, color=TEAL if total else DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    cell.border = thin_border()
    if total:
        cell.fill = fill(TEAL_BG)

# =============================================================================
# 1. Read current Costi setup values to preserve user edits
# =============================================================================
def read_setup_values(path):
    """Restituisce dict con voci correnti del file utente. Se file non esiste, defaults."""
    defaults = {
        # sez 1 - burocrazia
        6:  {"voce": "Pratica SUAP (Host Como)", "min": 100, "tipico": 150, "max": 200,
             "note": "Comprende: SCIA su impresainungiorno.gov.it, comunicazione Comune, follow-up CIR Provincia, follow-up CIN ministeriale, supporto SPID proprietario.", "url": None},
        7:  {"voce": "Oneri Comune Cernobbio — diritti segreteria", "min": 100, "tipico": 130, "max": 250,
             "note": "Da confermare con telefonata Sportello Unico (delibera n.161/2019)",
             "url": "https://sportellotelematico.comune.cernobbio.co.it"},
        # sez 2 - dotazioni obbligatorie
        12: {"voce": "Rilevatore CO (monossido) — EN 50291", "min": 19, "tipico": 22, "max": 30,
             "note": "X-Sense / LIORQUE / ELRO — sensore 10 anni",
             "url": "https://www.amazon.it/s?k=rilevatore+monossido+carbonio+EN+50291+X-Sense"},
        13: {"voce": "Rilevatore gas combustibili — EN 50194", "min": 27, "tipico": 30, "max": 40,
             "note": "Metano/GPL, wall-mount certificato",
             "url": "https://www.amazon.it/dp/B0DXYTVFCB"},
        14: {"voce": "Estintore 6kg ABC — EN 3-7 (carica 2026)", "min": 55, "tipico": 60, "max": 80,
             "note": "Certificazione 13A/34A 233BC, ricarica annuale",
             "url": "https://www.amazon.it/dp/B07PDYX8TY"},
        15: {"voce": "Kit pronto soccorso (raccomandato)", "min": 16, "tipico": 20, "max": 25,
             "note": "Non obbligatorio per locazione breve ma utile",
             "url": "https://www.amazon.it/s?k=kit+pronto+soccorso+casa"},
        # sez 3 - opzionali
        20: {"voce": "Foto pro (servizio fotografico)", "min": 0, "tipico": 250, "max": 350,
             "note": "ADR +15-25% vs foto smartphone. Salta se hai già foto pro.", "url": None},
        21: {"voce": "Ring Intercom Video — check-in remoto", "min": 0, "tipico": 70, "max": 100,
             "note": "Apertura porta remota da app, utile se non sei on-site",
             "url": "https://www.amazon.it/dp/B0CKGQ5913"},
        22: {"voce": "Cassaforte digitale per chiavi backup", "min": 0, "tipico": 25, "max": 50,
             "note": "Lockbox 4-cifre per recupero chiavi emergenza",
             "url": "https://www.amazon.it/s?k=key+lockbox+cassaforte+chiavi"},
    }
    if not os.path.exists(path):
        print(f"[INFO] File non trovato, uso defaults: {path}")
        return defaults

    try:
        wb = load_workbook(path, data_only=False)
        if "Costi setup" not in wb.sheetnames:
            return defaults
        ws = wb["Costi setup"]
        out = {}
        for r in [6, 7, 12, 13, 14, 15, 20, 21, 22]:
            voce = ws.cell(row=r, column=1).value
            if not voce:
                out[r] = defaults[r]
                continue
            # Preserva i numeri dell'utente, ma usa SEMPRE voce/note/url dai defaults
            # (la voce/note possono contenere riferimenti commerciali interni vecchi)
            out[r] = {
                "voce": defaults[r]["voce"],
                "min": ws.cell(row=r, column=2).value if ws.cell(row=r, column=2).value is not None else defaults[r]["min"],
                "tipico": ws.cell(row=r, column=3).value if ws.cell(row=r, column=3).value is not None else defaults[r]["tipico"],
                "max": ws.cell(row=r, column=4).value if ws.cell(row=r, column=4).value is not None else defaults[r]["max"],
                "note": defaults[r]["note"],
                "url": defaults[r]["url"],
            }
        print(f"[OK] Letti valori utente da {path}")
        return out
    except Exception as e:
        print(f"[WARN] Errore lettura {path}: {e}. Uso defaults.")
        return defaults

setup_values = read_setup_values(EXISTING_FILE)

# =============================================================================
# 2. Calendar definition (12 mesi Y1 + Y2)
# =============================================================================
# (mese, giorni, ADR, stagione, status, ramp)
# ramp != 1.0 solo per primi 2 mesi operativi Y1 (Ago-Set 2026, nuovo annuncio)
ANNO_1 = [
    ("Giu 2026", 30, 193, "alta", "setup",     1.00),  # SUAP latency
    ("Lug 2026", 31, 210, "alta", "setup",     1.00),
    ("Ago 2026", 31, 256, "alta", "operativo", "RAMP_AUG"),  # nuovo annuncio, no reviews
    ("Set 2026", 30, 204, "alta", "operativo", "RAMP_SEP"),  # secondo mese, reputation in costruzione
    ("Ott 2026", 31, 195, "media", "operativo", 1.00),
    ("Nov 2026", 30, 169, "bassa", "operativo", 1.00),
    ("Dic 2026", 31, 177, "bassa", "operativo", 1.00),
    ("Gen 2027", 31, 166, "bassa", "chiuso",    1.00),
    ("Feb 2027", 28, 177, "bassa", "chiuso",    1.00),
    ("Mar 2027", 31, 187, "bassa", "operativo", 1.00),
    ("Apr 2027", 30, 216, "media", "operativo", 1.00),
    ("Mag 2027", 31, 237, "media", "operativo", 1.00),
]

ANNO_2 = [
    ("Giu 2027", 30, 193, "alta", "operativo", 1.00),
    ("Lug 2027", 31, 210, "alta", "operativo", 1.00),
    ("Ago 2027", 31, 256, "alta", "operativo", 1.00),
    ("Set 2027", 30, 204, "alta", "operativo", 1.00),
    ("Ott 2027", 31, 195, "media", "operativo", 1.00),
    ("Nov 2027", 30, 169, "bassa", "operativo", 1.00),
    ("Dic 2027", 31, 177, "bassa", "operativo", 1.00),
    ("Gen 2028", 31, 166, "bassa", "chiuso",    1.00),
    ("Feb 2028", 29, 177, "bassa", "chiuso",    1.00),
    ("Mar 2028", 31, 187, "bassa", "operativo", 1.00),
    ("Apr 2028", 30, 216, "media", "operativo", 1.00),
    ("Mag 2028", 31, 237, "media", "operativo", 1.00),
]

# =============================================================================
# 3. Build workbook
# =============================================================================
wb = Workbook()

# -----------------------------------------------------------------------------
# Sheet "Parametri" — single source of truth (riferito da tutti gli altri)
# -----------------------------------------------------------------------------
ws_p = wb.active
ws_p.title = "Parametri"

ws_p["A1"] = "Parametri del modello (modificabili — gli altri sheet si ricalcolano)"
ws_p["A1"].font = f(size=15, bold=True, color=TEAL)
ws_p.merge_cells("A1:C1")
ws_p.row_dimensions[1].height = 26

# Header
for i, h in enumerate(["Parametro", "Valore", "Note"], start=1):
    c = ws_p.cell(row=3, column=i, value=h)
    style_header(c)
ws_p.row_dimensions[3].height = 26

# Parametri (riga, voce, valore default, formato, note)
PARAMS = [
    # (row, name, value, fmt, note)
    (4,  "Commissione portali (media ponderata)", 0.135, '0.0%', "Mix: 60% Airbnb (15%) + 30% Booking (16%) + 10% diretto (3%)"),
    (5,  "Commissione Host Como",    0.10,  '0.0%',           "Applicata su (revenue notti + extra ospite). Cleaning a parte."),
    (6,  "Cleaning fee €/turnover",  60,    '#,##0 "€"',       "Fatturata all'ospite, pass-through al servizio pulizie"),
    (7,  "Probabilità 3° ospite",    0.20,  '0%',             "20% delle notti occupate ha 3 ospiti"),
    (8,  "Supplemento 3° ospite €/notte", 50, '#,##0 "€"',     "Sovrapprezzo per la terza persona"),
    (10, "Soggiorno medio ALTA stagione",  3,  '0',           "notti per check-in"),
    (11, "Soggiorno medio MEDIA stagione", 4,  '0',           ""),
    (12, "Soggiorno medio BASSA stagione", 5,  '0',           "Soggiorni più lunghi in bassa stagione"),
    (14, "Occupazione ALTA stagione (× target)",   1.10, '0.00', "Max +10% rispetto al target (Giu-Set)"),
    (15, "Occupazione MEDIA stagione (× target)",  1.00, '0.00', "= target (Mag/Ott/Apr)"),
    (16, "Occupazione BASSA stagione (× target)",  0.75, '0.00', "-25% rispetto al target (Nov/Dic/Mar)"),
    (18, "Riduzione Ago 2026 (lancio nuovo annuncio)", 0.55, '0.00', "Primo mese operativo: no recensioni, visibilità ridotta. Range realistico 0.45-0.70"),
    (19, "Riduzione Set 2026 (costruzione reputazione)", 0.75, '0.00', "Secondo mese: prime recensioni. Range realistico 0.65-0.90"),
    (21, "Target occupazione SCENARIO BASSO",  0.40, '0%',     "Conservativo"),
    (22, "Target occupazione SCENARIO BASE",   0.50, '0%',     "Realistico"),
    (23, "Target occupazione SCENARIO ALTO",   0.60, '0%',     "Ottimistico"),
    (25, "Setup costs (totale tipico)", None, '#,##0 "€"',     "Link automatico al totale 'Tipico' del foglio 'Costi setup'"),
    (26, "Cedolare secca (aliquota)",  0.21, '0.0%',           "21% primo immobile destinato a locazione breve. 26% dal 2° al 4° (DL 145/2023 conv. L 191/2023)."),
    (27, "Quota incassi soggetta a ritenuta OTA", 1.00, '0%',  "Airbnb (dal 2017), Booking (dal 2024), Vrbo/Expedia (dal 2025) trattengono e versano la cedolare. Per booking diretto (no portale): 0. Mix attuale 60+30+10 OTA → 100%."),
]

for row, name, val, fmt, note in PARAMS:
    cl = ws_p.cell(row=row, column=1, value=name)
    style_label(cl)
    cv = ws_p.cell(row=row, column=2)
    if val is not None:
        cv.value = val
    style_input(cv, bold=True)
    cv.number_format = fmt
    cn = ws_p.cell(row=row, column=3, value=note)
    cn.font = f(size=9, italic=True, color="64748B")
    cn.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    cn.border = thin_border()

# Setup cost tipico = formula riferita a Costi setup (sarà C24, vedi sotto)
ws_p["B25"] = "='Costi setup'!C24"
style_input(ws_p["B25"], bold=True)
ws_p["B25"].number_format = '#,##0 "€"'

# Widths
ws_p.column_dimensions["A"].width = 38
ws_p.column_dimensions["B"].width = 14
ws_p.column_dimensions["C"].width = 55

# Legend (riga 29 — sotto i parametri B26/B27)
ws_p["A29"] = "Le celle gialle sono modificabili. Tutti i calcoli negli altri sheet le usano via formule."
ws_p["A29"].font = f(size=10, italic=True, color="64748B")
ws_p.merge_cells("A29:C29")

# Cell references (per riuso nelle formule monthly)
P_PORTAL = "Parametri!$B$4"
P_HC     = "Parametri!$B$5"
P_CLEAN  = "Parametri!$B$6"
P_XPROB  = "Parametri!$B$7"
P_XFEE   = "Parametri!$B$8"
P_AVG_A  = "Parametri!$B$10"
P_AVG_M  = "Parametri!$B$11"
P_AVG_B  = "Parametri!$B$12"
P_OCC_A  = "Parametri!$B$14"
P_OCC_M  = "Parametri!$B$15"
P_OCC_B  = "Parametri!$B$16"
P_RAMP_AUG = "Parametri!$B$18"
P_RAMP_SEP = "Parametri!$B$19"
P_TGT_LOW  = "Parametri!$B$21"
P_TGT_BASE = "Parametri!$B$22"
P_TGT_HIGH = "Parametri!$B$23"
P_SETUP    = "Parametri!$B$25"
P_CEDOLARE = "Parametri!$B$26"
P_OTA_SHARE = "Parametri!$B$27"

# -----------------------------------------------------------------------------
# Sheet "Costi setup" (preserva valori utente + SUM formule)
# -----------------------------------------------------------------------------
ws_s = wb.create_sheet("Costi setup")
ws_s["A1"] = "Costi iniziali — a carico del proprietario"
ws_s["A1"].font = f(size=16, bold=True, color=TEAL)
ws_s.merge_cells("A1:F1")
ws_s.row_dimensions[1].height = 28

ws_s["A2"] = "Tutte le voci sono una tantum. Celle gialle modificabili — i totali si ricalcolano via formule SUM."
ws_s["A2"].font = f(size=10, italic=True, color="64748B")
ws_s.merge_cells("A2:F2")

def section_header(ws, row, title):
    ws.cell(row=row, column=1, value=title).font = f(size=13, bold=True, color=TEAL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 24

def row_headers(ws, row):
    for i, h in enumerate(["Voce", "Min €", "Tipico €", "Max €", "Note", "Link Amazon"], start=1):
        c = ws.cell(row=row, column=i, value=h)
        style_header(c)
    ws.row_dimensions[row].height = 26

def write_item(ws, row, item):
    cl = ws.cell(row=row, column=1, value=item["voce"])
    style_label(cl)
    for cidx, key in enumerate(["min", "tipico", "max"], start=2):
        c = ws.cell(row=row, column=cidx, value=item[key])
        style_input(c)
        c.number_format = '#,##0 "€"'
    nc = ws.cell(row=row, column=5, value=item["note"])
    nc.font = f(size=9, italic=True, color="64748B")
    nc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    nc.border = thin_border()
    if item.get("url"):
        lc = ws.cell(row=row, column=6, value="Apri link")
        lc.hyperlink = item["url"]
        lc.font = f(size=10, color=BLUE_LINK, underline="single")
    else:
        lc = ws.cell(row=row, column=6, value="—")
        lc.font = f(size=10, color="94A3B8")
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = thin_border()

# Sezione 1 — burocrazia (rows 4-7)
section_header(ws_s, 4, "1. Pratiche burocratiche (SUAP + Comune)")
row_headers(ws_s, 5)
write_item(ws_s, 6, setup_values[6])
write_item(ws_s, 7, setup_values[7])

# Sezione 2 — dotazioni (rows 10-15)
section_header(ws_s, 10, "2. Dotazioni sicurezza OBBLIGATORIE (Legge 191/2023)")
row_headers(ws_s, 11)
write_item(ws_s, 12, setup_values[12])
write_item(ws_s, 13, setup_values[13])
write_item(ws_s, 14, setup_values[14])
write_item(ws_s, 15, setup_values[15])

# Sezione 3 — opzionali (rows 18-22)
section_header(ws_s, 18, "3. Opzionali RACCOMANDATI")
row_headers(ws_s, 19)
write_item(ws_s, 20, setup_values[20])
write_item(ws_s, 21, setup_values[21])
write_item(ws_s, 22, setup_values[22])

# Totale row 24 — SUM formula
ws_s.cell(row=24, column=1, value="TOTALE setup (formule SUM)").font = f(size=12, bold=True, color="FFFFFF")
ws_s.cell(row=24, column=1).fill = fill(TEAL)
ws_s.cell(row=24, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_s.cell(row=24, column=1).border = thin_border()

# Formula: SUM(B6:B7) + SUM(B12:B15) + SUM(B20:B22) per ogni colonna
for cidx, col in enumerate(["B", "C", "D"], start=2):
    formula = f"=SUM({col}6:{col}7)+SUM({col}12:{col}15)+SUM({col}20:{col}22)"
    c = ws_s.cell(row=24, column=cidx, value=formula)
    c.font = f(size=12, bold=True, color="FFFFFF")
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    c.number_format = '#,##0 "€"'

nc = ws_s.cell(row=24, column=5, value="Range Minimo → Massimo. Il Tipico è quello usato in 'Parametri!B25'.")
nc.font = f(size=10, italic=True, color="FFFFFF", bold=True)
nc.fill = fill(TEAL)
nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
nc.border = thin_border()
nc6 = ws_s.cell(row=24, column=6, value="")
nc6.fill = fill(TEAL)
nc6.border = thin_border()

# Widths
ws_s.column_dimensions["A"].width = 42
ws_s.column_dimensions["B"].width = 11
ws_s.column_dimensions["C"].width = 12
ws_s.column_dimensions["D"].width = 11
ws_s.column_dimensions["E"].width = 45
ws_s.column_dimensions["F"].width = 17

# -----------------------------------------------------------------------------
# Helper: scrivi sheet mensile con formule
# -----------------------------------------------------------------------------
def write_monthly_sheet(wb, name, calendar, scenario_target_ref, title):
    """
    scenario_target_ref: stringa tipo "Parametri!$B$22" (per Base) o $21/$23
    calendar: lista di tuple (mese, gg, adr, stagione, status, ramp)
    """
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = f(size=15, bold=True, color=TEAL)
    ws.merge_cells("A1:S1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Celle gialle = input (modificabili). Celle bianche = formule che derivano dagli input + Parametri. Cleaning €60/sosta è FUORI dal lordo (pagata dal guest direttamente al cleaner)."
    ws["A2"].font = f(size=9, italic=True, color="64748B")
    ws.merge_cells("A2:S2")

    headers = [
        "Mese", "gg", "Stagione", "Status", "ADR €", "Ramp",
        "Occ %", "Notti", "Avg stay", "Turn",
        "A — Notti", "B — Extra", "TOTAL lordo",
        "Fee Portali", "Fee HC", "Netto pre-cedolare",
        "(–) Cedolare OTA", "NETTO in tasca",
        "Cleaning (fuori lordo)",
    ]
    row_h = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row_h, column=i, value=h)
        style_header(c)
    ws.row_dimensions[row_h].height = 38

    for r_idx, (mese, gg, adr, stagione, status, ramp) in enumerate(calendar):
        rr = row_h + 1 + r_idx

        # Inputs (gialli, modificabili)
        ws.cell(row=rr, column=1, value=mese)
        ws.cell(row=rr, column=1).font = f(size=10, bold=True)
        ws.cell(row=rr, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=rr, column=1).border = thin_border()

        cgg = ws.cell(row=rr, column=2, value=gg); style_input(cgg); cgg.number_format = '0'
        cst = ws.cell(row=rr, column=3, value=stagione); style_input(cst)
        cstat = ws.cell(row=rr, column=4, value=status); style_input(cstat)
        cadr = ws.cell(row=rr, column=5, value=adr); style_input(cadr); cadr.number_format = '#,##0 "€"'

        # Ramp: se è "RAMP_AUG" o "RAMP_SEP", formula riferita a Parametri; altrimenti valore diretto
        cramp = ws.cell(row=rr, column=6)
        if ramp == "RAMP_AUG":
            cramp.value = f"={P_RAMP_AUG}"
        elif ramp == "RAMP_SEP":
            cramp.value = f"={P_RAMP_SEP}"
        else:
            cramp.value = ramp
        style_input(cramp)
        cramp.number_format = '0.00'

        # Cell refs per questa riga (nuovo schema: cleaning fuori dal lordo)
        gg_ref = f"B{rr}"
        st_ref = f"C{rr}"
        stat_ref = f"D{rr}"
        adr_ref = f"E{rr}"
        ramp_ref = f"F{rr}"
        occ_ref = f"G{rr}"
        notti_ref = f"H{rr}"
        avg_ref = f"I{rr}"
        turn_ref = f"J{rr}"
        a_ref = f"K{rr}"
        b_ref = f"L{rr}"
        tot_ref = f"M{rr}"  # TOTAL lordo = K+L (no cleaning)

        # Occ% = IF(status="operativo", target × factor_stagione × ramp, 0)
        occ_formula = (
            f'=IF({stat_ref}="operativo", {scenario_target_ref}*'
            f'IF({st_ref}="alta",{P_OCC_A},IF({st_ref}="media",{P_OCC_M},{P_OCC_B}))*'
            f'{ramp_ref}, 0)'
        )
        cocc = ws.cell(row=rr, column=7, value=occ_formula)
        style_data(cocc, eur=False, fmt='0%')

        # Notti = ROUND(gg * occ, 0)
        cnotti = ws.cell(row=rr, column=8, value=f"=ROUND({gg_ref}*{occ_ref},0)")
        style_data(cnotti, eur=False, fmt='0')

        # Avg stay = lookup
        avg_formula = f'=IF({st_ref}="alta",{P_AVG_A},IF({st_ref}="media",{P_AVG_M},{P_AVG_B}))'
        cavg = ws.cell(row=rr, column=9, value=avg_formula)
        style_data(cavg, eur=False, fmt='0')

        # Turn = IF(notti>0, ROUND(notti/avg, 0), 0)
        cturn = ws.cell(row=rr, column=10, value=f"=IF({notti_ref}>0,ROUND({notti_ref}/{avg_ref},0),0)")
        style_data(cturn, eur=False, fmt='0')

        # K = A = notti × ADR
        ca = ws.cell(row=rr, column=11, value=f"={notti_ref}*{adr_ref}")
        style_data(ca, eur=True)

        # L = B = ROUND(notti × extra_prob × extra_fee, 0)
        cb = ws.cell(row=rr, column=12, value=f"=ROUND({notti_ref}*{P_XPROB}*{P_XFEE},0)")
        style_data(cb, eur=True)

        # M = TOTAL lordo = A + B (cleaning ESCLUSO, gestito direttamente con cleaner)
        ctot = ws.cell(row=rr, column=13, value=f"={a_ref}+{b_ref}")
        style_data(ctot, bold=True, eur=True)

        # N = Fee Portali = -TOTAL × portal_rate
        cport = ws.cell(row=rr, column=14, value=f"=-{tot_ref}*{P_PORTAL}")
        style_data(cport, eur=True, color=RED)

        # O = Fee HC = -TOTAL × hc_rate (stesso base dei portali — coerente col modello user)
        chc = ws.cell(row=rr, column=15, value=f"=-{tot_ref}*{P_HC}")
        style_data(chc, eur=True, color=RED)

        # P = Netto pre-cedolare = TOTAL + Portal + HC (Portal/HC negativi)
        cnet = ws.cell(row=rr, column=16, value=f"={tot_ref}+N{rr}+O{rr}")
        cnet.font = f(size=10, bold=True, color=DARK)
        cnet.fill = fill(GRAY_BG)
        cnet.alignment = Alignment(horizontal="center", vertical="center")
        cnet.border = thin_border()
        cnet.number_format = '#,##0 "€"'

        # Q = Cedolare OTA = -TOTAL × cedolare × quota_OTA
        # (OTA trattengono direttamente alla fonte sul canone, escluso cleaning)
        ccd = ws.cell(row=rr, column=17, value=f"=-{tot_ref}*{P_CEDOLARE}*{P_OTA_SHARE}")
        style_data(ccd, eur=True, color=RED)

        # R = NETTO in tasca = Netto pre-cedolare + Cedolare (cedolare negativa)
        cnetfinal = ws.cell(row=rr, column=18, value=f"=P{rr}+Q{rr}")
        cnetfinal.font = f(size=10, bold=True, color=TEAL)
        cnetfinal.fill = fill(TEAL_BG)
        cnetfinal.alignment = Alignment(horizontal="center", vertical="center")
        cnetfinal.border = thin_border()
        cnetfinal.number_format = '#,##0 "€"'

        # S = Cleaning (fuori lordo) — informativa: turn × cleaning_fee
        # NON impatta proprietario (guest paga DIRETTAMENTE all'impresa pulizie)
        cclean_info = ws.cell(row=rr, column=19, value=f"={turn_ref}*{P_CLEAN}")
        cclean_info.font = f(size=10, italic=True, color="64748B")
        cclean_info.fill = fill(GRAY_BG)
        cclean_info.alignment = Alignment(horizontal="center", vertical="center")
        cclean_info.border = thin_border()
        cclean_info.number_format = '#,##0 "€"'

    # Total row (formula SUM)
    tot_row = row_h + 1 + len(calendar)
    ws.cell(row=tot_row, column=1, value="TOT 12 mesi").font = f(size=11, bold=True, color="FFFFFF")
    ws.cell(row=tot_row, column=1).fill = fill(TEAL)
    ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=tot_row, column=1).border = thin_border()
    # Dash per colonne non sommabili (gg, stagione, status, ADR, ramp, occ%, avg)
    for cidx in [2, 3, 4, 5, 6, 7, 9]:
        c = ws.cell(row=tot_row, column=cidx, value="—")
        c.font = f(size=11, bold=True, color="FFFFFF")
        c.fill = fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    # SUM per colonne sommabili (nuovo schema: cleaning fuori dal lordo)
    first_data_row = row_h + 1
    last_data_row = row_h + len(calendar)
    sum_cols = {
        8:  '0',                 # Notti
        10: '0',                 # Turn
        11: '#,##0 "€"',         # K — A notti
        12: '#,##0 "€"',         # L — B extra
        13: '#,##0 "€"',         # M — TOTAL lordo (K+L)
        14: '#,##0 "€"',         # N — Fee Portali
        15: '#,##0 "€"',         # O — Fee HC
        16: '#,##0 "€"',         # P — Netto pre-cedolare
        17: '#,##0 "€"',         # Q — Cedolare OTA
        18: '#,##0 "€"',         # R — NETTO in tasca
        19: '#,##0 "€"',         # S — Cleaning (fuori lordo, info)
    }
    for cidx, fmt in sum_cols.items():
        col_letter = get_column_letter(cidx)
        f_str = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        c = ws.cell(row=tot_row, column=cidx, value=f_str)
        c.font = f(size=11, bold=True, color="FFFFFF")
        c.fill = fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        c.number_format = fmt

    # Widths (19 colonne, generose per evitare wrap testo header)
    # A  B  C  D  E  F  G  H  I  J   K   L   M    N      O      P      Q       R       S
    widths = [12, 7, 11, 12, 10, 8, 9, 9, 11, 9, 12, 11, 13, 13, 13, 18, 17, 16, 19]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws, tot_row  # tot_row = riga dei totali (usabile in Sintesi)


sheet_y1_base, tot_row_y1_base = write_monthly_sheet(wb, "Anno 1 - Base", ANNO_1, P_TGT_BASE,
    "Anno 1 — Scenario Base (50%) — avviamento Giu-Lug, riduzione lancio Ago-Set, chiusura Gen-Feb")
sheet_y1_low, tot_row_y1_low = write_monthly_sheet(wb, "Anno 1 - Basso", ANNO_1, P_TGT_LOW,
    "Anno 1 — Scenario Basso (40%) — ipotesi conservativa")
sheet_y1_high, tot_row_y1_high = write_monthly_sheet(wb, "Anno 1 - Alto", ANNO_1, P_TGT_HIGH,
    "Anno 1 — Scenario Alto (60%) — ipotesi ottimistica")
sheet_y2_base, tot_row_y2_base = write_monthly_sheet(wb, "Anno 2 - Base", ANNO_2, P_TGT_BASE,
    "Anno 2 a regime — Scenario Base (50%) — solo chiusura Gen-Feb")
sheet_y2_low, tot_row_y2_low = write_monthly_sheet(wb, "Anno 2 - Basso", ANNO_2, P_TGT_LOW,
    "Anno 2 a regime — Scenario Basso (40%)")
sheet_y2_high, tot_row_y2_high = write_monthly_sheet(wb, "Anno 2 - Alto", ANNO_2, P_TGT_HIGH,
    "Anno 2 a regime — Scenario Alto (60%)")

# -----------------------------------------------------------------------------
# Sheet "Sintesi" — solo cross-sheet references
# -----------------------------------------------------------------------------
ws_sint = wb.create_sheet("Sintesi", 0)  # primo sheet
ws_sint["A1"] = "Andrea Furlan — Cernobbio, Via della Libertà 7"
ws_sint["A1"].font = f(size=18, bold=True, color=TEAL)
ws_sint.merge_cells("A1:E1")
ws_sint.row_dimensions[1].height = 30

ws_sint["A2"] = "Sintesi conservativa — tutti i numeri derivano dai sheet mensili via formule"
ws_sint["A2"].font = f(size=11, color=DARK)
ws_sint.merge_cells("A2:E2")

ws_sint["A3"] = "Anno 1: 8 mesi operativi (avviamento Giu-Lug 2026, chiusura inverno Gen-Feb 2027, riduzione lancio Ago-Set). Anno 2 a regime: 10 mesi operativi."
ws_sint["A3"].font = f(size=9, italic=True, color="64748B")
ws_sint.merge_cells("A3:E3")
ws_sint.row_dimensions[3].height = 22

def write_sintesi_section(ws, start_row, year_label, sheet_low, sheet_base, sheet_high, tot_row_low, tot_row_base, tot_row_high):
    """Costruisce sezione sintesi con formule riferite agli sheet mensili."""
    ws.cell(row=start_row, column=1, value=year_label).font = f(size=13, bold=True, color=TEAL)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    ws.row_dimensions[start_row].height = 24

    headers = ["Voce", "Scenario Basso (40%)", "Scenario Base (50%)", "Scenario Alto (60%)", "Note"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row + 1, column=i, value=h)
        style_header(c)
    ws.row_dimensions[start_row + 1].height = 28

    # (label, col_letter_in_monthly_sheet, fmt, eur, total, note)
    # Schema nuovo: cleaning FUORI dal lordo. Lordo = A+B = notti+extra.
    # Fee HC e Fee Portali ENTRAMBE su lordo (stesso base). Cedolare 21% su lordo.
    rows_def = [
        ("Notti occupate",      "H",  '0',              False, False, "Somma 12 mesi"),
        ("Turnover",            "J",  '0',              False, False, "Cambi ospite (soggiorno medio 3-5 notti)"),
        ("",                    None, None,             False, False, ""),
        ("A — Revenue notti",   "K",  '#,##0 "€"',       True,  False, "ADR × notti — incasso prenotazione"),
        ("B — Extra ospite",    "L",  '#,##0 "€"',       True,  False, "20% delle notti ha 3° ospite × €50"),
        ("TOTAL lordo (A+B)",   "M",  '#,##0 "€"',       True,  True,  "Lordo proprietario — base per portali, HC e cedolare"),
        ("",                    None, None,             False, False, ""),
        ("Fee portali 13,5%",   "N",  '#,##0 "€"',       True,  False, "Mix portali ponderato (Airbnb 15% + Booking 18% + diretto 3%)"),
        ("Fee Host Como 10%",   "O",  '#,##0 "€"',       True,  False, "Su lordo (stesso base dei portali) — rate intro"),
        ("",                    None, None,             False, False, ""),
        ("Netto pre-cedolare",  "P",  '#,##0 "€"',       True,  False, "Lordo – fee portali – fee HC"),
        ("(–) Cedolare 21% trattenuta OTA", "Q", '#,##0 "€"', True, False, "Airbnb/Booking/Expedia trattengono alla fonte e versano in F24"),
        ("NETTO IN TASCA proprietario", "R", '#,##0 "€"', True, True,  "Cash effettivamente incassato (dopo cedolare già pagata)"),
        ("",                    None, None,             False, False, ""),
        ("Cleaning gestita separatamente (info)", "S", '#,##0 "€"', True, False, "Guest paga €60/sosta direttamente all'impresa pulizie — fuori dal lordo proprietario"),
    ]

    for r_idx, (label, col, fmt, eur, total, note) in enumerate(rows_def):
        rr = start_row + 2 + r_idx
        cl = ws.cell(row=rr, column=1, value=label)
        if label == "":
            cl.border = Border()
        else:
            style_label(cl, total=total)
        if col is None:
            for cidx in range(2, 6):
                cb = ws.cell(row=rr, column=cidx)
                cb.border = Border()
            continue
        # 3 formule cross-sheet
        for cidx, (sname, trow) in enumerate([
            (sheet_low.title, tot_row_low),
            (sheet_base.title, tot_row_base),
            (sheet_high.title, tot_row_high),
        ], start=2):
            formula = f"='{sname}'!{col}{trow}"
            c = ws.cell(row=rr, column=cidx, value=formula)
            style_data(c, total_row=total, eur=eur, fmt=fmt)
            if isinstance(label, str) and label.startswith("Fee") or label.startswith("Costo"):
                c.font = f(size=10, bold=False, color=RED)
        # Note
        nc = ws.cell(row=rr, column=5, value=note)
        nc.font = f(size=9, italic=True, color="64748B")
        nc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        nc.border = thin_border()

    # Netto mensile medio (formula = NETTO IN TASCA / 12)
    # NETTO IN TASCA è all'index 12 in rows_def → riga = start_row + 2 + 12
    rr_avg = start_row + 2 + len(rows_def)
    cl = ws.cell(row=rr_avg, column=1, value="Netto in tasca / mese medio")
    style_label(cl, total=True)
    netto_intasca_row = start_row + 2 + 12
    for cidx in range(2, 5):
        col_letter = get_column_letter(cidx)
        c = ws.cell(row=rr_avg, column=cidx, value=f"={col_letter}{netto_intasca_row}/12")
        style_data(c, total_row=True, eur=True, fmt='#,##0 "€"')
    nc = ws.cell(row=rr_avg, column=5, value="Netto in tasca annuo (dopo cedolare) ÷ 12")
    nc.font = f(size=9, italic=True, color="64748B")
    nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    nc.border = thin_border()

    return rr_avg


end_y1 = write_sintesi_section(ws_sint, 5, "Anno 1 — con avviamento iniziale e chiusura inverno",
                                sheet_y1_low, sheet_y1_base, sheet_y1_high,
                                tot_row_y1_low, tot_row_y1_base, tot_row_y1_high)

end_y2 = write_sintesi_section(ws_sint, end_y1 + 3, "Anno 2 a regime — solo chiusura inverno",
                                sheet_y2_low, sheet_y2_base, sheet_y2_high,
                                tot_row_y2_low, tot_row_y2_base, tot_row_y2_high)

# Widths Sintesi
ws_sint.column_dimensions["A"].width = 33
ws_sint.column_dimensions["B"].width = 17
ws_sint.column_dimensions["C"].width = 17
ws_sint.column_dimensions["D"].width = 17
ws_sint.column_dimensions["E"].width = 50

# Chart Sintesi: confronto NETTO IN TASCA Y1 vs Y2 (riferimenti a celle Sintesi)
chart_row = end_y2 + 3
ws_sint.cell(row=chart_row, column=1, value="Confronto Netto in tasca Anno 1 vs Anno 2 (dopo cedolare)").font = f(size=12, bold=True, color=TEAL)

# NETTO IN TASCA è all'index 12 di rows_def → riga = start_row + 2 + 12
y1_netto_row = 5 + 2 + 12   # = 19
y2_netto_row = end_y1 + 3 + 2 + 12

# Tabella dati per chart
chart_data_row = chart_row + 2
for cidx, h in enumerate(["Scenario", "Anno 1", "Anno 2"], start=1):
    c = ws_sint.cell(row=chart_data_row, column=cidx, value=h)
    style_header(c)
ws_sint.row_dimensions[chart_data_row].height = 26

for i, (lbl, col) in enumerate([("Basso (40%)", "B"), ("Base (50%)", "C"), ("Alto (60%)", "D")]):
    cl = ws_sint.cell(row=chart_data_row + 1 + i, column=1, value=lbl)
    cl.font = f(size=10, bold=True)
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cl.border = thin_border()
    c1 = ws_sint.cell(row=chart_data_row + 1 + i, column=2, value=f"={col}{y1_netto_row}")
    c1.number_format = '#,##0 "€"'
    c1.font = f(size=10)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = thin_border()
    c2 = ws_sint.cell(row=chart_data_row + 1 + i, column=3, value=f"={col}{y2_netto_row}")
    c2.number_format = '#,##0 "€"'
    c2.font = f(size=10)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = thin_border()

chart = BarChart()
chart.type = "col"
chart.style = 11
chart.title = "Netto in tasca (post-cedolare) — Anno 1 vs Anno 2"
chart.y_axis.title = "Euro / anno"
chart.x_axis.title = "Scenario"
data = Reference(ws_sint, min_col=2, min_row=chart_data_row, max_col=3, max_row=chart_data_row + 3)
chart.add_data(data, titles_from_data=True)
cats = Reference(ws_sint, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + 3)
chart.set_categories(cats)
chart.height = 9
chart.width = 18
chart.dataLabels = DataLabelList(showVal=True)
ws_sint.add_chart(chart, f"F{chart_row}")

# -----------------------------------------------------------------------------
# Sheet "Breakeven Y1" — cumulato con formule
# -----------------------------------------------------------------------------
ws_be = wb.create_sheet("Breakeven Anno 1")
ws_be["A1"] = "Punto di pareggio Anno 1 — quando il proprietario recupera l'investimento iniziale"
ws_be["A1"].font = f(size=16, bold=True, color=TEAL)
ws_be.merge_cells("A1:E1")
ws_be.row_dimensions[1].height = 28

ws_be["A2"] = "L'investimento setup esce a Maggio 2026. Da Giugno in poi: cumulato = cumulato mese precedente + NETTO IN TASCA del mese (Scenario Base 50%, dopo cedolare già trattenuta da OTA). Tutti i numeri sono formule."
ws_be["A2"].font = f(size=10, italic=True, color="64748B")
ws_be.merge_cells("A2:E2")

headers = ["Mese", "Status", "Netto mese", "Cumulato netto", "Note"]
for i, h in enumerate(headers, start=1):
    c = ws_be.cell(row=4, column=i, value=h)
    style_header(c)
ws_be.row_dimensions[4].height = 28

# Row 5: Mag 2026 — setup esborso (-Parametri setup)
ws_be.cell(row=5, column=1, value="Mag 2026").font = f(size=10, bold=True)
ws_be.cell(row=5, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_be.cell(row=5, column=1).border = thin_border()
ws_be.cell(row=5, column=2, value="Setup (out)").font = f(size=10, bold=True, color=ORANGE)
ws_be.cell(row=5, column=2).alignment = Alignment(horizontal="center", vertical="center")
ws_be.cell(row=5, column=2).border = thin_border()
c3 = ws_be.cell(row=5, column=3, value=f"=-{P_SETUP}")
c3.font = f(size=10, color=RED)
c3.alignment = Alignment(horizontal="center", vertical="center")
c3.border = thin_border()
c3.number_format = '#,##0 "€"'
c4 = ws_be.cell(row=5, column=4, value="=C5")  # primo cumul = setup
c4.font = f(size=10, bold=True, color=RED)
c4.alignment = Alignment(horizontal="center", vertical="center")
c4.border = thin_border()
c4.number_format = '#,##0 "€"'
c4.fill = fill(YELLOW_BG)
ws_be.cell(row=5, column=5, value="Il proprietario sostiene l'investimento iniziale").font = f(size=9, italic=True, color="64748B")
ws_be.cell(row=5, column=5).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_be.cell(row=5, column=5).border = thin_border()

# Rows 6-17: i 12 mesi Y1 Base
for i, (mese, gg, adr, stagione, status, ramp) in enumerate(ANNO_1):
    rr = 6 + i
    # Mese
    ws_be.cell(row=rr, column=1, value=mese).font = f(size=10, bold=True)
    ws_be.cell(row=rr, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_be.cell(row=rr, column=1).border = thin_border()
    # Status
    status_color = {"setup": ORANGE, "chiuso": "94A3B8", "operativo": GREEN}[status]
    ws_be.cell(row=rr, column=2, value=status.capitalize()).font = f(size=10, bold=True, color=status_color)
    ws_be.cell(row=rr, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws_be.cell(row=rr, column=2).border = thin_border()
    # Netto mese = riferimento al NETTO IN TASCA in 'Anno 1 - Base' (colonna R, dopo cedolare)
    monthly_row = 5 + i
    c3 = ws_be.cell(row=rr, column=3, value=f"='Anno 1 - Base'!R{monthly_row}")
    c3.font = f(size=10, color=(GREEN if status == "operativo" else "94A3B8"))
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.border = thin_border()
    c3.number_format = '#,##0 "€"'
    # Cumulato = cumulato precedente + netto mese
    c4 = ws_be.cell(row=rr, column=4, value=f"=D{rr-1}+C{rr}")
    c4.font = f(size=11, bold=True, color=TEAL)
    c4.alignment = Alignment(horizontal="center", vertical="center")
    c4.border = thin_border()
    c4.number_format = '#,##0 "€"'
    # Note: indica breakeven se cumulato passa 0 (calcolato a posteriori, ma posso usare formula IF)
    # Formula: =IF(AND(D{rr}>=0, D{rr-1}<0), "BREAKEVEN ✓", "")
    cnote = ws_be.cell(row=rr, column=5, value=f'=IF(AND(D{rr}>=0,D{rr-1}<0),"BREAKEVEN ✓","")')
    cnote.font = f(size=10, italic=True, bold=True, color=GREEN)
    cnote.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cnote.border = thin_border()

# Widths
ws_be.column_dimensions["A"].width = 13
ws_be.column_dimensions["B"].width = 14
ws_be.column_dimensions["C"].width = 16
ws_be.column_dimensions["D"].width = 20
ws_be.column_dimensions["E"].width = 40

# Chart breakeven (line)
chart = LineChart()
chart.title = "Cumulato netto Anno 1 — Scenario Base (50%)"
chart.style = 12
chart.y_axis.title = "Cumulato €"
chart.x_axis.title = "Mese"
data_ref = Reference(ws_be, min_col=4, min_row=4, max_row=17)
chart.add_data(data_ref, titles_from_data=True)
cats = Reference(ws_be, min_col=1, min_row=5, max_row=17)
chart.set_categories(cats)
chart.height = 10
chart.width = 22
ws_be.add_chart(chart, "G4")

# Sintesi breakeven box (formule)
ws_be.cell(row=20, column=1, value="Sintesi punto di pareggio").font = f(size=13, bold=True, color=TEAL)
ws_be.merge_cells("A20:E20")

# Cells sintesi
ws_be.cell(row=21, column=1, value="Investimento iniziale (esborso)").font = f(size=10, bold=True)
ws_be.cell(row=21, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_be.cell(row=21, column=1).border = thin_border()
c = ws_be.cell(row=21, column=2, value=f"=C5")
c.font = f(size=11, bold=True, color=RED); c.alignment = Alignment(horizontal="center"); c.border = thin_border(); c.number_format = '#,##0 "€"'
ws_be.merge_cells("B21:C21")

ws_be.cell(row=22, column=1, value="Netto cumulato a fine Anno 1").font = f(size=10, bold=True)
ws_be.cell(row=22, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_be.cell(row=22, column=1).border = thin_border()
c = ws_be.cell(row=22, column=2, value=f"=D17")
c.font = f(size=11, bold=True, color=TEAL); c.alignment = Alignment(horizontal="center"); c.border = thin_border(); c.number_format = '#,##0 "€"'
ws_be.merge_cells("B22:C22")

# -----------------------------------------------------------------------------
# Sheet "Affitto medio termine" (formule)
# -----------------------------------------------------------------------------
ws_mtr = wb.create_sheet("Affitto medio termine")
ws_mtr["A1"] = "Alternativa: affitto medio termine Gennaio-Febbraio invece di chiusura"
ws_mtr["A1"].font = f(size=15, bold=True, color=TEAL)
ws_mtr.merge_cells("A1:D1")
ws_mtr.row_dimensions[1].height = 28

ws_mtr["A2"] = "Scenario: affittare a tenant unico (lavoratore in trasferta, expat) per 2 mesi invernali. No portali, no fee Airbnb, 1 pulizia inizio + 1 fine."
ws_mtr["A2"].font = f(size=10, italic=True, color="64748B")
ws_mtr["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws_mtr.merge_cells("A2:D2")
ws_mtr.row_dimensions[2].height = 40

# Input cells gialle
for i, h in enumerate(["Voce", "Importo", "Note"], start=1):
    c = ws_mtr.cell(row=4, column=i, value=h)
    style_header(c)
ws_mtr.row_dimensions[4].height = 26

# Inputs (gialle)
ws_mtr.cell(row=5, column=1, value="Canone mensile €").font = f(size=10, bold=True)
ws_mtr.cell(row=5, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_mtr.cell(row=5, column=1).border = thin_border()
crent = ws_mtr.cell(row=5, column=2, value=1200); style_input(crent, bold=True); crent.number_format = '#,##0 "€"'
ws_mtr.cell(row=5, column=3, value="Stima conservativa Cernobbio (zona lago) — modificabile").font = f(size=9, italic=True, color="64748B")
ws_mtr.cell(row=5, column=3).border = thin_border()

ws_mtr.cell(row=6, column=1, value="Durata (mesi)").font = f(size=10, bold=True)
ws_mtr.cell(row=6, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_mtr.cell(row=6, column=1).border = thin_border()
cdur = ws_mtr.cell(row=6, column=2, value=2); style_input(cdur, bold=True); cdur.number_format = '0'
ws_mtr.cell(row=6, column=3, value="Gennaio + Febbraio = i 2 mesi con domanda turistica minima").font = f(size=9, italic=True, color="64748B")
ws_mtr.cell(row=6, column=3).border = thin_border()

# LORDO = formula
ws_mtr.cell(row=7, column=1, value="LORDO affitto").font = f(size=11, bold=True, color=TEAL)
ws_mtr.cell(row=7, column=1).fill = fill(TEAL_BG)
ws_mtr.cell(row=7, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_mtr.cell(row=7, column=1).border = thin_border()
clordo = ws_mtr.cell(row=7, column=2, value="=B5*B6")
clordo.font = f(size=11, bold=True, color=TEAL); clordo.fill = fill(TEAL_BG)
clordo.alignment = Alignment(horizontal="center", vertical="center"); clordo.border = thin_border()
clordo.number_format = '#,##0 "€"'
ws_mtr.cell(row=7, column=3, value="= Canone × Durata").font = f(size=9, italic=True, color="64748B")
ws_mtr.cell(row=7, column=3).border = thin_border()

# HC fee = -lordo × HC rate (riferito a Parametri)
ws_mtr.cell(row=8, column=1, value="(–) Fee Host Como").font = f(size=10, bold=True)
ws_mtr.cell(row=8, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_mtr.cell(row=8, column=1).border = thin_border()
chc = ws_mtr.cell(row=8, column=2, value=f"=-B7*{P_HC}")
chc.font = f(size=10, color=RED); chc.alignment = Alignment(horizontal="center", vertical="center"); chc.border = thin_border()
chc.number_format = '#,##0 "€"'
ws_mtr.cell(row=8, column=3, value=f"= LORDO × Fee HC (Parametri B5)").font = f(size=9, italic=True, color="64748B")
ws_mtr.cell(row=8, column=3).border = thin_border()

# Cleaning cost = -2 cleaning fees
ws_mtr.cell(row=9, column=1, value="(–) Pulizie (ingresso + uscita)").font = f(size=10, bold=True)
ws_mtr.cell(row=9, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_mtr.cell(row=9, column=1).border = thin_border()
cclean = ws_mtr.cell(row=9, column=2, value=f"=-2*{P_CLEAN}")
cclean.font = f(size=10, color=RED); cclean.alignment = Alignment(horizontal="center", vertical="center"); cclean.border = thin_border()
cclean.number_format = '#,##0 "€"'
ws_mtr.cell(row=9, column=3, value="= 2 × cleaning fee (Parametri B6)").font = f(size=9, italic=True, color="64748B")
ws_mtr.cell(row=9, column=3).border = thin_border()

# NETTO MTR
ws_mtr.cell(row=10, column=1, value="NETTO proprietario").font = f(size=11, bold=True, color=TEAL)
ws_mtr.cell(row=10, column=1).fill = fill(TEAL_BG)
ws_mtr.cell(row=10, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_mtr.cell(row=10, column=1).border = thin_border()
cnet = ws_mtr.cell(row=10, column=2, value="=B7+B8+B9")
cnet.font = f(size=11, bold=True, color=TEAL); cnet.fill = fill(TEAL_BG)
cnet.alignment = Alignment(horizontal="center", vertical="center"); cnet.border = thin_border()
cnet.number_format = '#,##0 "€"'
ws_mtr.cell(row=10, column=3, value="Cash extra rispetto a chiudere").font = f(size=9, italic=True, color="64748B")
ws_mtr.cell(row=10, column=3).border = thin_border()

# Confronto Y1 con/senza affitto medio termine
ws_mtr.cell(row=13, column=1, value="Confronto Anno 1 — con/senza affitto medio termine").font = f(size=13, bold=True, color=TEAL)
ws_mtr.merge_cells("A13:D13")
ws_mtr.row_dimensions[13].height = 24

for i, h in enumerate(["Scenario", "Senza affitto medio termine", "Con affitto medio termine", "Delta"], start=1):
    c = ws_mtr.cell(row=15, column=i, value=h)
    style_header(c)
ws_mtr.row_dimensions[15].height = 26

# Riferimenti a Sintesi NETTO IN TASCA Y1 (riga 19, colonne B=Basso, C=Base, D=Alto)
# Nota: il netto MTR (B10) è pre-cedolare; per coerenza viene mostrato come delta lordo,
# considerando che la cedolare sui contratti diretti (non OTA) viene pagata in F24 a fine anno
for i, (lbl, col) in enumerate([("Basso (40%)", "B"), ("Base (50%)", "C"), ("Alto (60%)", "D")]):
    rr = 16 + i
    ws_mtr.cell(row=rr, column=1, value=lbl).font = f(size=10, bold=True)
    ws_mtr.cell(row=rr, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_mtr.cell(row=rr, column=1).border = thin_border()
    # Senza MTR = Sintesi!{col}19 (Y1 NETTO IN TASCA row)
    c1 = ws_mtr.cell(row=rr, column=2, value=f"=Sintesi!{col}19")
    c1.font = f(size=10); c1.number_format = '#,##0 "€"'
    c1.alignment = Alignment(horizontal="center", vertical="center"); c1.border = thin_border()
    # Con MTR = Senza + NETTO MTR × (1 - cedolare) per portarlo in "tasca" consistente
    c2 = ws_mtr.cell(row=rr, column=3, value=f"={get_column_letter(2)}{rr}+B10*(1-{P_CEDOLARE})")
    c2.font = f(size=10, bold=True, color=TEAL); c2.number_format = '#,##0 "€"'
    c2.alignment = Alignment(horizontal="center", vertical="center"); c2.border = thin_border()
    # Delta = Con - Senza
    c3 = ws_mtr.cell(row=rr, column=4, value=f"=C{rr}-B{rr}")
    c3.font = f(size=10, bold=True, color=GREEN); c3.number_format = '#,##0 "€"'
    c3.alignment = Alignment(horizontal="center", vertical="center"); c3.border = thin_border()

ws_mtr.column_dimensions["A"].width = 35
ws_mtr.column_dimensions["B"].width = 22
ws_mtr.column_dimensions["C"].width = 22
ws_mtr.column_dimensions["D"].width = 14

# -----------------------------------------------------------------------------
# Sheet "Assunzioni & note"
# -----------------------------------------------------------------------------
ws_a = wb.create_sheet("Assunzioni & note")
ws_a["A1"] = "Assunzioni e note metodologiche"
ws_a["A1"].font = f(size=16, bold=True, color=TEAL)
ws_a.merge_cells("A1:C1")
ws_a.row_dimensions[1].height = 26

ws_a["A2"] = "Tutti i parametri numerici sono modificabili dal sheet 'Parametri'. Questo foglio è solo documentazione."
ws_a["A2"].font = f(size=10, italic=True, color="64748B")
ws_a.merge_cells("A2:C2")

assumptions = [
    ("Foglio", "Contenuto"),
    ("Sintesi", "Riepilogo annuale per 3 scenari (Anno 1 con avviamento, Anno 2 a regime). Mostra TOTAL booking → fee portali → fee HC → cleaning → Netto pre-cedolare → Cedolare 21% → NETTO IN TASCA. Tutti i numeri sono il risultato di formule che derivano dai fogli mensili."),
    ("Parametri", "Parametri numerici del modello: commissioni, occupazione target, stagionalità, riduzione primi mesi di lancio, aliquota cedolare, quota incassi soggetta a ritenuta OTA. Modificabili direttamente — gli altri fogli si ricalcolano."),
    ("Costi setup", "Voci di spesa iniziale a carico del proprietario, suddivise in: pratiche burocratiche, dotazioni di sicurezza obbligatorie (Legge 191/2023), opzionali raccomandati. Range Min/Tipico/Max."),
    ("Mensile Anno 1 (Base/Basso/Alto)", "12 mesi dettagliati Giu 2026 - Mag 2027 con setup Giu-Lug, chiusura inverno Gen-Feb e riduzione primi 2 mesi operativi per nuovo annuncio. Colonne: A notti + B extra = TOTAL lordo → fee portali → fee HC → netto pre-cedolare → cedolare 21% → NETTO IN TASCA. Cleaning €60/sosta in colonna info S (fuori lordo, pagata dal guest direttamente al cleaner)."),
    ("Mensile Anno 2 (Base/Basso/Alto)", "12 mesi a regime Giu 2027 - Mag 2028, solo chiusura inverno Gen-Feb."),
    ("Breakeven Anno 1", "Cumulato mese per mese con setup esborso a Maggio 2026. Usa NETTO IN TASCA (cash effettivo dopo cedolare già trattenuta da OTA). Indica il mese di pareggio (scenario Base)."),
    ("Alternativa affitto medio termine", "Scenario alternativo: invece di chiudere Gen-Feb, affittare a tenant unico per 2 mesi a canone fisso. Confronto su NETTO IN TASCA per coerenza."),
    ("", ""),
    ("Modello cleaning fuori dal lordo", "Spiegazione"),
    ("Trattamento spese pulizia", "La cleaning fee (€60/sosta) è fatturata SEPARATAMENTE dal canone di prenotazione. Il guest paga: (a) il prezzo notti+extra al proprietario tramite OTA, (b) i €60 cleaning direttamente all'impresa pulizie partner. Quindi NON impattano il proprietario: nessun ricavo aggiuntivo, nessun costo da sostenere, nessuna commissione portali/HC su di essi, nessuna cedolare sul cleaning."),
    ("Implicazione su lordo e cedolare", "Il LORDO proprietario è solo (A notti + B extra) — esclusa cleaning. Tutte le formule (fee portali, fee HC, cedolare secca) si applicano a questo lordo, NON al cleaning. La colonna S 'Cleaning (fuori lordo)' in ogni foglio mensile è puramente informativa: indica quanto va all'impresa pulizie nel mese."),
    ("", ""),
    ("Cedolare secca trattenuta dalle OTA — come funziona", "Spiegazione"),
    ("Cosa è la cedolare secca", "Imposta sostitutiva IRPEF al 21% sui redditi da locazione breve (≤30 gg). Aliquota 26% dal secondo al quarto immobile (DL 145/2023 conv. L 191/2023). Per Andrea Furlan: 21% perché primo immobile."),
    ("Chi la trattiene oggi (2026)", "Airbnb (dal 2017), Booking.com (dal 2024), Vrbo / Expedia Group (dal giugno 2025). Le piattaforme agiscono come sostituto d'imposta: trattengono il 21% sul corrispettivo lordo e versano direttamente al fisco italiano via F24 entro il 16 del mese successivo."),
    ("Cosa riceve il proprietario", "Riceve direttamente nel conto il NETTO (lordo - commissione portale - cedolare 21%). Non deve fare F24, è già pagata. A marzo dell'anno successivo riceve dal portale la Certificazione Unica (CU) con totale lordo + ritenuta operata."),
    ("Cosa fa il proprietario in dichiarazione", "Dichiara il reddito lordo da locazione nel 730 / Modello Redditi → indica regime cedolare secca → compensa con il credito d'imposta da CU. Se ha solo un immobile e tutto è già stato trattenuto al 21%, saldo netto = 0."),
    ("Base imponibile (su cosa è il 21%)", "Canone lordo = (notti × ADR) + extra ospite. Cleaning ESCLUSA perché fatturata direttamente dal cleaner al guest (non è canone del proprietario). Nessuna deduzione forfettaria del 5% (vale solo per regime ordinario IRPEF)."),
    ("Booking diretto / canale proprio", "Se il proprietario riceve prenotazioni dirette (senza OTA), la cedolare NON è trattenuta — il proprietario la paga in F24 a fine anno. Parametro 'Quota incassi soggetta a ritenuta OTA' (Parametri B27) = 100% se tutto via OTA, < 100% se c'è anche diretto."),
    ("Implicazione sul cash flow", "Per Andrea: tutto il fatturato 12m è via OTA (Airbnb + Booking + Expedia), quindi cedolare è interamente trattenuta alla fonte mensilmente. Cash flow mensile = 'NETTO IN TASCA' della tabella mensile. Nessun pagamento aggiuntivo F24 a fine anno."),
    ("Fonti normative", "DL 50/2017 art. 4 (introduzione ritenuta 21%) · DL 145/2023 conv. L 191/2023 (aliquota 26% dal 2°-4° immobile, dotazioni sicurezza) · Sentenza CGUE C-83/21 (12/2022, legittimità ritenuta su piattaforme estere)."),
    ("", ""),
    ("Cosa NON è incluso nel modello", "Note"),
    ("Tassa di soggiorno comunale", "Riscossa direttamente dall'ospite e versata al Comune dal gestionale. Non incide sul netto del proprietario."),
    ("Spese fisse dell'immobile", "Utenze, IMU, spese condominiali, manutenzione ordinaria. Costi strutturali a carico del proprietario indipendentemente dall'affitto breve."),
    ("Inflazione anno su anno", "Per l'Anno 2 sono mantenuti gli stessi ADR dell'Anno 1 (impostazione conservativa)."),
    ("Capex futuri", "Sostituzione dotazioni di sicurezza, eventuale refresh foto (stimato €100-200/anno, non incluso)."),
    ("", ""),
    ("Logica del lancio nuovo annuncio (Ago-Set 2026)", "Note"),
    ("Agosto 2026", "Primo mese operativo del nuovo annuncio: senza recensioni e con visibilità ridotta sui portali. Range realistico di occupazione 45-70% rispetto al target stagionale."),
    ("Settembre 2026", "Secondo mese: arrivano le prime recensioni e la visibilità migliora. Range 65-90% del target stagionale."),
    ("Da Ottobre 2026 in poi", "Annuncio considerato a regime: stagionalità standard applicata senza riduzione di lancio."),
]

SECTION_HEADERS = {
    "Foglio",
    "Modello cleaning fuori dal lordo",
    "Cedolare secca trattenuta dalle OTA — come funziona",
    "Cosa NON è incluso nel modello",
    "Logica del lancio nuovo annuncio (Ago-Set 2026)",
}
for r_idx, row_data in enumerate(assumptions, start=4):
    for c_idx, val in enumerate(row_data, start=1):
        c = ws_a.cell(row=r_idx, column=c_idx, value=val)
        is_section_header = val in SECTION_HEADERS or (c_idx == 2 and row_data[0] in SECTION_HEADERS)
        if is_section_header:
            style_header(c)
        else:
            c.font = f(size=10, bold=(c_idx == 1 and val != ""))
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            if val != "":
                c.border = thin_border()
            ws_a.row_dimensions[r_idx].height = max(ws_a.row_dimensions[r_idx].height or 0, 30)

ws_a.column_dimensions["A"].width = 38
ws_a.column_dimensions["B"].width = 80
ws_a.row_dimensions[4].height = 28

# =============================================================================
# Workbook metadata (client-facing)
# =============================================================================
wb.properties.creator = "Host Como"
wb.properties.lastModifiedBy = "Host Como"
wb.properties.title = "Proiezione economica — Andrea Furlan, Cernobbio"
wb.properties.subject = "Proiezione 24 mesi STR + Setup costs + Alternativa MTR"
wb.properties.description = "Modello a formule live per la gestione affitti brevi dell'immobile di Via della Libertà 7, Cernobbio."
wb.properties.keywords = "Host Como, Cernobbio, STR, proiezione, breakeven"
try:
    wb.properties.company = "Host Como"
except Exception:
    pass

# =============================================================================
# Save (retry if locked)
# =============================================================================
def save_with_retry(wb, path):
    for attempt in range(3):
        try:
            wb.save(path)
            return path
        except PermissionError:
            if attempt < 2:
                time.sleep(1)
            else:
                base = os.path.splitext(path)[0]
                tmp_path = f"{base}-v3.xlsx"
                wb.save(tmp_path)
                print(f"[WARN] {path} locked, saved as: {tmp_path}")
                return tmp_path
    return path

result = save_with_retry(wb, OUT_DESKTOP)
out_dir = os.path.dirname(OUT_PRATICA)
if os.path.isdir(out_dir):
    try:
        shutil.copy(result, OUT_PRATICA)
        print(f"[OK] Copied to: {OUT_PRATICA}")
    except PermissionError:
        base = os.path.splitext(OUT_PRATICA)[0]
        tmp_p = f"{base}-v3.xlsx"
        shutil.copy(result, tmp_p)
        print(f"[WARN] Pratica locked, saved as: {tmp_p}")
else:
    print(f"[SKIP] Cartella Pratica non esiste, salto la copia: {out_dir}")
print(f"[OK] Saved: {result}")
print(f"     Sheets: {wb.sheetnames}")
print(f"     Architecture: formule live (Parametri + Costi setup + 6 mensili + Sintesi + Breakeven + MTR)")
